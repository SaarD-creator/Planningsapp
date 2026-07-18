# fix open spots maar verdere testen nodig!
# toevoegingen Antwerpen compleet
# overall verbetering pauzeplanning
# pauzevlindercheck is toegevoegd maar weinig getest
# oplossing voor Antwerpen met introductie vinkje
# gaten gevuld in post processing lange blokken
# oude pauzeplanning weg!
# mooiere blokken met tweede plekken
# nieuwe verdeling van de blokken! drempel voorlopig op 60% voor ingesloten shiften + 18% voor vermijden éénuursblokken
# kleine fix bij last minute qua layout + ontdekt dat afkapping uren niet altijd top werkt (PV staat niet bij extra of op planning)
# nieuwe logica voor studenten die langer werken dan effectieve uren op planning
# Last minute planning is vaak niet top
# overschakeling compleettt
# splitsing volgens ideaalmomenten
# samengevoegde attracties als 1 --> fix in plaatsing! maar nog vaak 2x drie uur bij 1 attractie
# niet meer laden na elke verandering
# Werkblad heropleidingen! Plus op de planning
# LM: afkappingen is in orde (automatisch op planning) Maar: lange blokken (5 uur aan een stuk) kunnen precies voorkomen...
# post-processing laat geen switches toe die een 1-uursblok achterlaten (ook niet in LM)
# max van vier uur aaneensluitend is weg, maar lossere regels in post-processing (wel vaak vier uursblokken)
# last minute afwezigen ziet er al goed uit!
# uitschakelen pauzevlinderuren werkt!
# volgorde verdeling met pauzevlinders laatst!
# pauzes kloppen!! 1h15 min pauze voor minderjarige lange werkers --> alleen nog niet top op korte dagen (mogelijks gefixt)
# nieuw werkblad analyse
# zelfde versie als 3.5 maar pauzevlinders zijn ook volgens volgorde uit gekozen nummertje
#betere verdeling 3 uur blokken, maar te veel 6 uur bij zelfde attractie & 1+3 logica voor 9u30 ipv 3+1 & 2+2 logica voor 4 uur opt einde

#uitschakelen attracties op bepaalde uren lijkt te werken!
#samenvoegen attracties per uur werkttttt!!! Kleine bug is er uit gehaald
#hele dag bij attractie werkt
# probleem met twee


import streamlit as st
import random
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime
import re


# -----------------------------
# Excelbestand uploaden
# -----------------------------
uploaded_file = st.file_uploader("Upload Excel bestand", type=["xlsx"])

if not uploaded_file:
    st.session_state.pop("lm_base_bytes", None)   # ← reset bij nieuw bestand
    st.warning("Upload eerst het Excelbestand met de gegevens om verder te gaan.")
    st.stop()

file_bytes = uploaded_file.read()

# Workbook met berekende waarden voor de gewone planning
wb = load_workbook(BytesIO(file_bytes), data_only=True)
ws = wb["Input"]


ws_speciaal = wb["Input_"]
ws_aanpassingen = wb["Aanpassingen"]
ws_instellingen = wb["Instellingen"]

# --- VINKJE B2 (Instellingen): forceer perfect exhaustieve verdeling ---
_vinkje_y17 = ws_instellingen.cell(row=2, column=2).value   # B2
FORCEER_EXHAUSTIEF = _vinkje_y17 in [1, True, "WAAR", "X"]
# ------------------------------------------------------------------------

# --- VINKJE B3 (Instellingen): pauze pas vanaf MEER dan 4 uur werk (i.p.v. vanaf 4 uur) ---
_vinkje_b3 = ws_instellingen.cell(row=3, column=2).value   # B3
PAUZE_STRIKT_BOVEN_4U = _vinkje_b3 in [1, True, "WAAR", "X"]
# ------------------------------------------------------------------------
# ------------------------------------------------------------------------

# -----------------------------
# Datum op basis van W4 in Input_
# -----------------------------
_vandaag_datum = datetime.date.today()
_morgen_datum = _vandaag_datum + datetime.timedelta(days=1)
_overmorgen_datum = _vandaag_datum + datetime.timedelta(days=2)
_w4 = str(ws_speciaal.cell(4, 23).value or "").strip().lower()
if _w4 == "morgen":
    vandaag = _morgen_datum.strftime("%d-%m-%Y")
elif _w4 == "overmorgen":
    vandaag = _overmorgen_datum.strftime("%d-%m-%Y")
else:
    vandaag = _vandaag_datum.strftime("%d-%m-%Y")
vandaag_altijd_vandaag = _vandaag_datum.strftime("%d-%m-%Y")  # altijd vandaag, voor last-minute

_w5 = str(ws_speciaal.cell(5, 23).value or "").strip().lower()
RUSTIG_MODUS = (_w5 == "rustig")


def parse_uur_waarde(val):
    """
    Zet een celwaarde om naar een uur als float.
    Werkt met: int (10), float (17.5), str ('10', '17,5', '17:30'),
               datetime.time (10:00:00 → 10.0, 17:30:00 → 17.5)
    Geeft None terug als de waarde niet parseerbaar is.
    """
    if val is None:
        return None
    import datetime as _dt
    if isinstance(val, _dt.time):
        return val.hour + val.minute / 60 + val.second / 3600
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            return int(parts[0]) + int(parts[1]) / 60
        except:
            return None
    s = s.replace(",", ".").replace("u", "").strip()
    try:
        return float(s)
    except:
        return None
        

def formatteer_uur(u):
    """10 → '10u', 13.5 → '13u30', 9.25 → '9u15'"""
    uren = int(u)
    minuten = round((u - uren) * 60)
    if minuten == 0:
        return f"{uren}u"
    return f"{uren}u{minuten:02d}"


# Kolom→uur mapping vanuit Input_ rij 2 (I2:S2)
# Nodig voor samenvoeg-attracties en dichte uren, die vóór open_uren gelezen worden
col_to_uur_speciaal = {}
for _kol in range(9, 20):
    _uur = parse_uur_waarde(ws_speciaal.cell(2, _kol).value)
    if _uur is not None and _uur != 0:
        col_to_uur_speciaal[_kol] = _uur
        

def parse_blok_duur(label_str):
    """Parst '0,5h', '0,75h', '1,5h' → float. Geeft 1.0 als niet parseerbaar."""
    if not label_str:
        return 1.0
    s = str(label_str).strip().replace(",", ".").replace("h", "").strip()
    try:
        return float(s)
    except ValueError:
        return 1.0



# Blokduur per uur (uit I1:S1 van Input_)
# Kolommen zonder label krijgen standaard 1.0 uur
blok_durations = {}
for _kol in range(9, 20):  # kolom I t/m S
    _uur = col_to_uur_speciaal.get(_kol)
    if _uur is None:
        continue
    _label = ws_speciaal.cell(1, _kol).value
    blok_durations[_uur] = parse_blok_duur(_label)


ws_studenten = wb["Studenten"]


# -----------------------------

# Hulpfuncties
# -----------------------------
def max_consecutive_hours(urenlijst):
    if not urenlijst:
        return 0
    urenlijst = sorted(set(urenlijst))
    maxr = huidig = 1
    for i in range(1, len(urenlijst)):
        huidig = huidig + 1 if urenlijst[i] == urenlijst[i-1] + 1 else 1
        maxr = max(maxr, huidig)
    return maxr



def compute_ideal_moments():
    """
    Nieuwe ideaalmomenten o.b.v. de echte shiften van de studenten.
    - shift = aaneensluitend werkinterval (PV-uren eruit geknipt)
    - stap 1: split de drukste shift in blokken van <= 3 uur
    - stap 2/3: kies per span de opsplitsing met de meeste begin/eind-dekking
    - recursie: elk overgebleven stuk > 3u wordt opnieuw opgesplitst
    """
    # shiften verzamelen (zelfde uren-filter als assign_student)
    shifts = defaultdict(int)
    for s in studenten_workend:
        uren = sorted(u for u in s["uren_beschikbaar"] if u in open_uren)
        if s["is_pauzevlinder"]:
            uren = [u for u in uren if u not in required_pauze_hours]
        for run in contiguous_runs(uren):
            shifts[(run[0], run[-1] + 1)] += 1   # (start, eind-marker)

    # Aparte shiften-telling ZONDER pauzevlinders, uitsluitend voor de "forceer exhaustief"-berekening
    shifts_zonder_pv = defaultdict(int)
    for s in studenten_workend:
        if s.get("is_pauzevlinder"):
            continue
        uren = sorted(u for u in s["uren_beschikbaar"] if u in open_uren)
        for run in contiguous_runs(uren):
            shifts_zonder_pv[(run[0], run[-1] + 1)] += 1

    if not shifts:
        return set()

    open_start = min(open_uren)
    open_end = max(open_uren) + 1

    # begin/eind-histogram (geteld in #studenten)
    hist = defaultdict(int)
    for (start, eind), aantal in shifts.items():
        hist[start] += aantal
        hist[eind] += aantal
        
    top_count = max(shifts.values())
    DEKKING_DREMPEL = 0.18   # een begin/eind-moment telt pas mee als >= 18% van de drukste shift het deelt

    # alle geordende opsplitsingen van n in exact k delen, elk 1..3
    def _composities(n, k):
        if k == 1:
            if 1 <= n <= 3:
                yield (n,)
            return
        for eerste in range(1, 4):
            rest = n - eerste
            if rest < k - 1:
                break
            if rest > (k - 1) * 3:
                continue
            for staart in _composities(rest, k - 1):
                yield (eerste,) + staart

    def _kies_cuts(a, b):
        n = int(round(b - a))
        k = -(-n // 3)                            # ceil(n/3)
        beste, beste_sleutel = None, None
        for comp in _composities(n, k):
            cuts, u = set(), a
            for L in comp:
                cuts.add(u); u += L
            cuts.add(b)
            sleutel = (
                -sum(1 for L in comp if L == 1),                                                       # 1) minste EIGEN 1-uursblokken
                sum(hist.get(h, 0) for h in cuts if hist.get(h, 0) >= DEKKING_DREMPEL * top_count),     # 2) dekking breekt gelijkstand (3-2 vs 2-3)
                comp,
            )
            if beste_sleutel is None or sleutel > beste_sleutel:
                beste, beste_sleutel = cuts, sleutel
        return beste

    # Stap 1
    # -- volledig grid bouwen vanuit één sturende shift --
    def _bouw_grid(sturende):
        grid = {open_start, open_end}
        grid |= _kies_cuts(sturende[0], sturende[1])
        veranderd = True
        while veranderd:
            veranderd = False
            for g1, g2 in zip(sorted(grid), sorted(grid)[1:]):
                if g2 - g1 > 3:
                    grid |= _kies_cuts(g1, g2)
                    veranderd = True
                    break
        return grid

    # -- kwaliteit van een grid over de hele populatie --
    def _kwaliteit(grid):
        enen = wissels = 0
        for (a, b), aantal in shifts.items():
            run = [h for h in open_uren if a <= h < b]
            bl, huidig = [], 1
            for i in range(1, len(run)):
                if run[i] in grid:
                    bl.append(huidig); huidig = 1
                else:
                    huidig += 1
            bl.append(huidig)
            enen += aantal * sum(1 for x in bl if x == 1)
            wissels += aantal * (len(bl) - 1)
        return (enen, wissels)

    # stap 1: welke EENHEID stuurt? Een eenheid = losse shift OF complementair half-paar
    # (open_start..m + m..open_end), dat als één telt met de som van beide helften.
    DREMPEL = 0.30   # aandeel van de niet-PV-shiften waarboven de populairste eenheid stuurt

    # noemer: alle shiften zonder de pauzevlinders
    niet_pv_totaal = sum(
        len(contiguous_runs(sorted(u for u in s["uren_beschikbaar"] if u in open_uren)))
        for s in studenten_workend if not s.get("is_pauzevlinder")
    ) or 1

    def _half_grid(paar):
        (a, m), (_, b) = paar
        return {open_start, open_end} | _kies_cuts(a, m) | _kies_cuts(m, b)

    eenheden = [(aantal, "single", se) for se, aantal in shifts.items()]
    _m_kandidaten = ({e for (s, e) in shifts if s == open_start}     # eindes van shiften die om open_start starten
                     & {s for (s, e) in shifts if e == open_end})    # starts van shiften die op open_end eindigen
    paren = [
        ((open_start, m), (m, open_end), shifts[(open_start, m)] + shifts[(m, open_end)])
        for m in _m_kandidaten
    ]
    beste_paar = max(paren, key=lambda p: p[2]) if paren else None

    # --- Alles hieronder ENKEL voor FORCEER_EXHAUSTIEF: zelfde berekening, maar met de
    #     pauzevlinders volledig uit de shiften-telling gehaald (shifts_zonder_pv). ---
    if FORCEER_EXHAUSTIEF:
        _m_kandidaten_ex = ({e for (s, e) in shifts_zonder_pv if s == open_start}
                             & {s for (s, e) in shifts_zonder_pv if e == open_end})
        paren_ex = [
            ((open_start, m), (m, open_end), shifts_zonder_pv[(open_start, m)] + shifts_zonder_pv[(m, open_end)])
            for m in _m_kandidaten_ex
        ]
        beste_paar_ex = max(paren_ex, key=lambda p: p[2]) if paren_ex else None

        # Y17 aangevinkt -> altijd volgens de perfect exhaustieve shiften (pauzevlinders niet meegeteld)
        if beste_paar_ex is not None:
            return _half_grid((beste_paar_ex[0], beste_paar_ex[1]))

        # Y17 aangevinkt maar geen echt complementair paar aanwezig (zonder pauzevlinders):
        # zoek een halve-dag-shift die vanaf open_start vertrekt of op open_end eindigt,
        # waarbij het resterende stuk van de dag minstens 4 uur bedraagt. Die shift wordt
        # dan gebruikt om een geforceerd (virtueel) complementair paar te bouwen.
        _halve_dag_kandidaten = []
        for (s, e), aantal in shifts_zonder_pv.items():
            if s == open_start and (open_end - e) >= 4:
                _halve_dag_kandidaten.append((aantal, e))    # m = einde van de shift
            if e == open_end and (s - open_start) >= 4:
                _halve_dag_kandidaten.append((aantal, s))    # m = start van de shift
        if _halve_dag_kandidaten:
            _, _m_gedwongen = max(_halve_dag_kandidaten, key=lambda x: x[0])
            return _half_grid(((open_start, _m_gedwongen), (_m_gedwongen, open_end)))
        # geen geschikte halve-dag-shift gevonden (zonder pauzevlinders) -> val terug op de gewone logica hieronder

    if beste_paar:
        eenheden.append((beste_paar[2], "pair", (beste_paar[0], beste_paar[1])))

    def _eenheid_grid(u):
        return _half_grid(u[2]) if u[1] == "pair" else _bouw_grid(u[2])

    # hoogste aantal; bij gelijkspel de eenheid met de minste totale 1-uursblokken
    leider_aantal, leider_soort, leider_data = max(
        eenheden, key=lambda u: (u[0], -_kwaliteit(_eenheid_grid(u))[0])
    )

    if leider_aantal / niet_pv_totaal >= DREMPEL:
        # duidelijke leider stuurt (half-paar via propere helften, losse shift optimaal gesplitst)
        return _half_grid(leider_data) if leider_soort == "pair" else _bouw_grid(leider_data)

    # geen leider boven de drempel -> stuur op het totale aantal 1-uursblokken
    top_count = max(shifts.values())
    kandidaat_grids = [_bouw_grid(se) for se in shifts if shifts[se] >= 0.60 * top_count]
    if beste_paar:
        kandidaat_grids.append(_half_grid((beste_paar[0], beste_paar[1])))
    return min(kandidaat_grids, key=_kwaliteit)
    

def partition_run_lengths(run_hours, ideal_moments=None):
    """
    Knipt de run van een student op de ideaalmomenten.
    Een nieuw blok begint telkens als een uur een ideaalmoment is (behalve het eerste uur).
    """
    if not run_hours:
        return []
    ideal = ideal_moments or set()
    blokken, huidig = [], 1
    for i in range(1, len(run_hours)):
        if run_hours[i] in ideal:
            blokken.append(huidig)
            huidig = 1
        else:
            huidig += 1
    blokken.append(huidig)
    return blokken


def contiguous_runs(sorted_hours):
    runs = []
    if not sorted_hours:
        return runs
    # Gebruik positie in open_uren, niet uur+1
    blok_index = {u: i for i, u in enumerate(sorted(open_uren))}
    run = [sorted_hours[0]]
    for h in sorted_hours[1:]:
        if blok_index.get(h, -99) == blok_index.get(run[-1], -98) + 1:
            run.append(h)
        else:
            runs.append(run)
            run = [h]
    runs.append(run)
    return runs

# Helpers die in meerdere delen gebruikt worden
def normalize_attr(name):
    """Normaliseer attractienaam zodat 'X 2' telt als 'X'; trim & lower-case voor vergelijking."""
    if not name:
        return ""
    s = str(name).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        s = parts[0]
    return s.strip().lower()

def parse_header_uur(header):
    """Map headertekst (bv. '14u', '14:00', '14:30') naar het hele uur (14)."""
    if not header:
        return None
    s = str(header).strip()
    try:
        if "u" in s:
            return int(s.split("u")[0])
        if ":" in s:
            uur, _min = s.split(":")
            return int(uur)
        return int(s)
    except:
        return None

# -----------------------------
# Studenten inlezen
# -----------------------------
studenten = []

# Lees attractienamen uit rij 1 van 'Studenten', kolommen G t/m X (7-24)
attractie_namen_studenten = [
    ws_studenten.cell(1, kol).value
    for kol in range(7, 25)
]

for rij in range(2, 500):
    naam = ws_studenten.cell(rij, 5).value  # kolom E
    if not naam:
        continue

    # Werkuren uit C (beginuur) en D (einduur)
    _begin = parse_uur_waarde(ws_studenten.cell(rij, 3).value)
    _eind  = parse_uur_waarde(ws_studenten.cell(rij, 4).value)
    if _begin is None or _eind is None:
        uren_beschikbaar = []
    else:
        uren_beschikbaar = [
            u for u in col_to_uur_speciaal.values()
            if _begin <= u and u + blok_durations.get(u, 1.0) <= _eind
        ]

    # Attracties uit kolommen G-X (7-24), namen staan in rij 1
    attracties = [
        attractie_namen_studenten[kol - 7]
        for kol in range(7, 25)
        if ws_studenten.cell(rij, kol).value in [1, True, "WAAR", "X"]
        and attractie_namen_studenten[kol - 7]
    ]

    # Aantal attracties uit kolom Z (26)
    try:
        aantal_attracties = int(ws_studenten.cell(rij, 26).value) if ws_studenten.cell(rij, 26).value else len(attracties)
    except:
        aantal_attracties = len(attracties)

    studenten.append({
        "naam": naam,
        "uren_beschikbaar": sorted(uren_beschikbaar),
        "attracties": attracties,
        "aantal_attracties": aantal_attracties,
        "is_pauzevlinder": False,
        "pv_number": None,
        "assigned_attracties": set(),
        "assigned_hours": [],
        "begin_uur": _begin,
        "eind_uur": _eind,
    })
dichte_uren_per_attr = defaultdict(set)
# Input_: rijen 17 t/m 22, vakjes in I-S (kol 9-19), attractienaam in T (kol 20)

for rij in range(17, 23):  # rij 17 t/m 22
    attr_naam_raw = ws_speciaal.cell(rij, 20).value  # kolom T
    if attr_naam_raw:
        attr_naam = normalize_attr(attr_naam_raw)
        for col_idx in range(9, 20):  # kolom I t/m S
            val = ws_speciaal.cell(rij, col_idx).value
            if val in [1, True, "WAAR", "X"]:
                uur = col_to_uur_speciaal.get(col_idx)
                if uur:
                    dichte_uren_per_attr[attr_naam].add(uur)

# -----------------------------
# Samenvoeg-attracties (per uur)
# -----------------------------


# In DEEL 1 bij "Samenvoeg-attracties (per uur)"
uur_samenvoegingen = defaultdict(list)
# Input_: rijen 10 t/m 15, vakjes in I-S (kol 9-19), attractienamen in T-U-V (kol 20-22)

for rij in range(10, 16):  # rij 10 t/m 15
    groep = []
    for col in range(20, 23):  # kolom T, U, V
        val = ws_speciaal.cell(rij, col).value
        if val:
            groep.append(str(val).strip())

    if len(groep) > 1:
        for col_idx in range(9, 20):  # kolom I t/m S
            if ws_speciaal.cell(rij, col_idx).value in [1, True, "WAAR", "X"]:
                uur = col_to_uur_speciaal.get(col_idx)
                if uur:
                    uur_samenvoegingen[uur].append(groep)


# -----------------------------
# Alle mogelijke samengevoegde attracties (namen)
# -----------------------------

samengevoegde_attracties = set()

for groepen in uur_samenvoegingen.values():
    for groep in groepen:
        samengevoegde_attracties.add(" + ".join(groep))



# -----------------------------
# Voeg samengestelde attracties toe aan individuele studenten
# -----------------------------
for s in studenten:
    huidige = set(s["attracties"])
    for sameng in samengevoegde_attracties:
        onderdelen = [a.strip() for a in sameng.split("+")]
        if all(o in huidige for o in onderdelen):
            s["attracties"].append(sameng)  # voeg de samengestelde attractie toe




# -----------------------------
# Openingsuren
# -----------------------------
# Openingsuren lezen vanuit Input_, I2:S2 (kolommen 9-19)
open_uren = []
uur_labels = {}

for kol in range(9, 20):
    val = ws_speciaal.cell(2, kol).value
    uur = parse_uur_waarde(val)
    if uur is None or uur == 0:
        continue
    open_uren.append(uur)
    label = ws_speciaal.cell(1, kol).value
    if label and str(label).strip():
        uur_labels[uur] = str(label).strip()

if not open_uren:
    open_uren = list(range(10, 19))
open_uren = sorted(set(open_uren))
        


# -----------------------------
# Sorteervolgorde studenten
# Eerst op aantal attracties,
# daarna op vaste tie-break regel uit BU2
# -----------------------------
bu2_waarde = ws_speciaal.cell(3, 23).value  # W3 in Input_
try:
    tie_break_mode = int(bu2_waarde)
except:
    tie_break_mode = 1

if tie_break_mode not in [1, 2, 3, 4, 5]:
    tie_break_mode = 1


def naam_tie_break_key(naam_raw):
    naam = str(naam_raw).strip().lower()

    if tie_break_mode == 1:
        # gewone alfabetische volgorde
        return naam

    elif tie_break_mode == 2:
        # omgekeerde alfabetische volgorde
        return "".join(chr(255 - ord(c)) for c in naam)

    elif tie_break_mode == 3:
        # eerst op aantal letters, daarna alfabetisch
        return (len(naam), naam)

    elif tie_break_mode == 4:
        # alfabetisch op basis van laatste letters
        return naam[::-1]

    elif tie_break_mode == 5:
        # omgekeerde van mode 4
        return "".join(chr(255 - ord(c)) for c in naam[::-1])

    return naam




# -----------------------------

# Pauzevlinders
# -----------------------------
# Input_: C14 t/m C18 (kolom 3, rijen 14-18)
pauzevlinder_namen = [
    ws_speciaal.cell(rij, 3).value
    for rij in range(14, 19)
    if ws_speciaal.cell(rij, 3).value
]

# Pauzevlinderuren lezen vanuit Input_, I3:S3 (kolommen 9-19)
required_pauze_hours = []
for kol in range(9, 20):
    val = ws_speciaal.cell(3, kol).value
    try:
        uur = int(val)
    except (TypeError, ValueError):
        continue
    if uur and uur in open_uren:
        required_pauze_hours.append(uur)
required_pauze_hours = sorted(set(required_pauze_hours))


for idx,pvnaam in enumerate(pauzevlinder_namen,start=1):
    for s in studenten:
        if s["naam"]==pvnaam:
            s["is_pauzevlinder"]=True
            s["pv_number"]=idx
            s["uren_beschikbaar"]=[u for u in s["uren_beschikbaar"] if u not in required_pauze_hours]
            break

# Maak 'selected' lijst van pauzevlinders (dicts met naam en attracties)
selected = [s for s in studenten if s.get("is_pauzevlinder")]
selected = sorted(selected, key=lambda s: naam_tie_break_key(s["naam"]))

# -----------------------------
# Attracties & aantallen (raw)
# -----------------------------
aantallen_raw = {}
attracties_te_plannen = []
for rij in range(3, 21):  # E3:F20 in Aanpassingen
    naam = ws_aanpassingen.cell(rij, 5).value  # kolom E
    if naam:
        try:
            aantal = int(ws_aanpassingen.cell(rij, 6).value)  # kolom F
        except:
            aantal = 0
        max_toegestaan = 1 if RUSTIG_MODUS else 2
        aantallen_raw[naam] = max(0, min(max_toegestaan, aantal))
        if aantallen_raw[naam] >= 1:
            attracties_te_plannen.append(naam)

# Priority order for second spots (column BA, rows 5-11)
second_priority_order = [
    ws_aanpassingen.cell(rij, 9).value  # kolom I
    for rij in range(3, 13)             # I3:I12
    if ws_aanpassingen.cell(rij, 9).value
]


# -----------------------------
# Attractielijst uitbreiden met samengevoegde attracties (globaal)
# -----------------------------

for nieuwe in samengevoegde_attracties:
    if nieuwe not in attracties_te_plannen:
        attracties_te_plannen.append(nieuwe)
    aantallen_raw[nieuwe] = 1


# -----------------------------
# Actieve attracties per uur (ivm samenvoegingen)
# -----------------------------

actieve_attracties_per_uur = {}
# Gebruik de raw aantallen als basis
aantallen = {uur: {a: aantallen_raw.get(a, 1) for a in attracties_te_plannen} for uur in open_uren}

for uur in open_uren:
    actief = set()
    # 1. Voeg eerst alle individuele attracties toe die NIET dicht zijn
    for a in attracties_te_plannen:
        if " + " in a: continue # Sla samengevoegde namen hier nog even over
        
        if uur in dichte_uren_per_attr.get(normalize_attr(a), set()):
            aantallen[uur][a] = 0
        else:
            actief.add(a)

    # 2. Verwerk de samenvoegingen voor dit specifieke uur
    huidige_groepen = uur_samenvoegingen.get(uur, [])
    for groep in huidige_groepen:
        samengevoegde_naam = " + ".join(groep)
        
        # Voeg de samengevoegde attractie toe aan de planning
        actief.add(samengevoegde_naam)
        aantallen[uur][samengevoegde_naam] = 1
        
        # VERWIJDER de onderdelen uit de actieve lijst (voorkomt dubbele telling)
        for onderdeel in groep:
            if onderdeel in actief:
                actief.remove(onderdeel)
            aantallen[uur][onderdeel] = 0

    actieve_attracties_per_uur[uur] = actief



### -----------------------------
### Compute aantallen per hour + red spots (GEÏNTEGREERD)
### -----------------------------
red_spots = {uur: set() for uur in open_uren}          
second_spot_blocked = {uur: set() for uur in open_uren}  

for uur in open_uren:
    # 1. Hoeveel studenten zijn er dit uur echt beschikbaar? [1]
    student_count = sum(
        1 for s in studenten
        if uur in s["uren_beschikbaar"] and not (
            s["is_pauzevlinder"] and uur in required_pauze_hours
        )
    )
    
    # 2. Hoeveel attracties moeten dit uur minimaal 1 persoon hebben? [1]
    # We kijken naar de actieve lijst van dat uur (rekening houdend met uitschakelingen/samenvoegingen)
    base_spots = sum(1 for a in actieve_attracties_per_uur[uur] if aantallen[uur].get(a, 0) >= 1)
    
    # 3. Bereken het overschot
    extra_spots = student_count - base_spots

    # 4. Verdeel de tweede plekken op basis van de prioriteitslijst uit Excel (BA5:BA11) [2]
    for attr in second_priority_order:
        # Check of de attractie dit uur actief is én of hij normaal 2 personen nodig heeft [2, 3]
        if attr in actieve_attracties_per_uur[uur] and aantallen_raw.get(attr) == 2:
            if extra_spots > 0:
                # Er is nog een student over voor een tweede plek
                aantallen[uur][attr] = 2
                extra_spots -= 1
            else:
                # Geen studenten meer over? Blokkeer de tweede plek voor dit uur
                second_spot_blocked[uur].add(attr)
                aantallen[uur][attr] = 1  # Forceer het aantal voor dit uur naar 1


# -----------------------------
# Red spots for samengestelde attracties
# -----------------------------

for uur in open_uren:
    # Groepen die dit uur samengevoegd zijn
    groepen = uur_samenvoegingen.get(uur, [])

    # Samengestelde attracties die DIT uur actief zijn
    samengestelde = set(" + ".join(g) for g in groepen)

    # Losse attracties die in een samenvoeging zitten
    losse_in_samenvoeging = set(a for g in groepen for a in g)

    # 1️⃣ Samenvoeging actief → losse attracties verbieden
    for attr in losse_in_samenvoeging:
        red_spots[uur].add(attr)

    # 2️⃣ Samenvoeging NIET actief → samenvoeging verbieden
    for samengestelde_attr in samengevoegde_attracties:
        if samengestelde_attr not in samengestelde:
            red_spots[uur].add(samengestelde_attr)


# -----------------------------
# Studenten die effectief inzetbaar zijn
# -----------------------------
studenten_workend = [
    s for s in studenten if any(u in open_uren for u in s["uren_beschikbaar"])
]


ideaalmomenten = compute_ideal_moments()  


# -----------------------------
# Blacklist van attracties per student (BB16:BG79)
# -----------------------------
student_blacklist = defaultdict(set)

for rij in range(3, 26):  # O3:P25 in Aanpassingen
    naam = ws_aanpassingen.cell(rij, 15).value  # kolom O
    if not naam:
        continue
    naam = str(naam).strip()
    attr = ws_aanpassingen.cell(rij, 16).value  # kolom P (1 attractie)
    if attr:
        student_blacklist[naam].add(str(attr).strip().lower())


# Sorteer attracties op "kritieke score" (hoeveel studenten ze kunnen doen)
def kritieke_score(attr, studenten_list):
    return sum(1 for s in studenten_list if attr in s["attracties"])

attracties_te_plannen.sort(key=lambda a: kritieke_score(a, studenten_workend))



# -----------------------------
# Assign per student
# -----------------------------
assigned_map = defaultdict(list)  # (uur, attr) -> list of student-names
per_hour_assigned_counts = {uur: {a: 0 for a in attracties_te_plannen} for uur in open_uren}
extra_assignments = defaultdict(list)


# -----------------------------
# Overbodige pauzevlinderuren → laatste PV naar extra
# -----------------------------
import math

aantal_pv = len(selected)
aantal_pauze_uren = len(required_pauze_hours)

def pp2_bepaal_pv_voor_afknip(selected_lijst):
    """
    Bepaal welke pauzevlinder afgeknipt wordt:
    1) degene met het minst aantal attracties (de 'moeilijkste') -- wint
       altijd als dit uniek is, ongeacht naam.
    2) bij gelijkstand: als één van de gelijk-staande kandidaten
       "Vlinder..." heet, wordt die afgeknipt.
    3) bij gelijkstand zonder "Vlinder"-naam: gewoon de laatste
       pauzevlinder van de volledige lijst (het oorspronkelijke gedrag).
    """
    if not selected_lijst:
        return None

    min_aantal = min(len(pv.get("attracties", [])) for pv in selected_lijst)
    kandidaten = [pv for pv in selected_lijst if len(pv.get("attracties", [])) == min_aantal]

    if len(kandidaten) == 1:
        return kandidaten[0]

    vlinder_kandidaten = [
        pv for pv in kandidaten
        if str(pv.get("naam", "")).strip().lower().startswith("vlinder")
    ]
    if vlinder_kandidaten:
        return vlinder_kandidaten[-1]

    return selected_lijst[-1]


afgekapte_pv_uren = set()  # nieuw: bijhouden welke uren afgekapt worden

if aantal_pv > 0 and aantal_pauze_uren > 0:
    plaatsen_pauzeplanning = (aantal_pauze_uren * 4 - 1) * aantal_pv

    try:
        lange_pauzes = int(ws_speciaal.cell(15, 6).value) if ws_speciaal.cell(15, 6).value else 0  # F15
    except:
        lange_pauzes = 0
    try:
        korte_pauzes = int(ws_speciaal.cell(16, 6).value) if ws_speciaal.cell(16, 6).value else 0  # F16
    except:
        korte_pauzes = 0

    pauze_kwartieren = 2 * lange_pauzes + korte_pauzes
    open_spots = plaatsen_pauzeplanning - pauze_kwartieren
    min_open_spots_per_pv = 1 if len(open_uren) <= 6 else 3

    _afknip_kandidaat = pp2_bepaal_pv_voor_afknip(selected)

    # Probeer de strengste begrenzing eerst; val pas terug op een lossere
    # begrenzing als de strengere zichzelf niet waarmaakt (te weinig uren
    # om af te knippen om die begrenzing te rechtvaardigen).
    _marge_tiers = [
        (0, 4),                      # marge 0, vereist >=4 afgeknipte uren
        (2, 2),                      # marge 2, vereist >=2 afgeknipte uren
        (min_open_spots_per_pv, 0),  # normale marge, geen minimum vereist
    ]

    for marge_kandidaat, _vereiste_uren in _marge_tiers:
        marge_totaal = 0
        for _pv_marge in selected:
            if _afknip_kandidaat is not None and _pv_marge["naam"] == _afknip_kandidaat["naam"]:
                marge_totaal += marge_kandidaat
            else:
                marge_totaal += min_open_spots_per_pv

        beschikbaar = open_spots - marge_totaal
        if beschikbaar >= 3:
            overbodige_uren = 1 + max(0, math.floor((beschikbaar - 3) / 4))
        else:
            overbodige_uren = 0

        pv_pauze_uren = sorted(required_pauze_hours, reverse=True)
        uren_te_verschuiven = min(overbodige_uren, len(pv_pauze_uren))

        if uren_te_verschuiven >= _vereiste_uren:
            break  # deze marge houdt zichzelf overeind -> gebruiken

    if overbodige_uren > 0:
        laatste_pv = _afknip_kandidaat

        # Check: valt deze PV VOLLEDIG weg, is er dan met 1 PV minder
        # (en dus ook 1 marge minder) nog voldoende capaciteit voor de
        # rest om de effectieve pauzebehoefte te dekken? Zo ja: knip
        # meteen de volledige shift af i.p.v. enkel een deel.
        if uren_te_verschuiven < len(pv_pauze_uren):
            aantal_pv_zonder_laatste = aantal_pv - 1
            if aantal_pv_zonder_laatste > 0:
                plaatsen_zonder_laatste = (aantal_pauze_uren * 4 - 1) * aantal_pv_zonder_laatste
                marge_zonder_laatste = aantal_pv_zonder_laatste * min_open_spots_per_pv
                nodig_zonder_laatste = pauze_kwartieren + marge_zonder_laatste
                if plaatsen_zonder_laatste >= nodig_zonder_laatste:
                    uren_te_verschuiven = len(pv_pauze_uren)

        for i in range(uren_te_verschuiven):
            uur = pv_pauze_uren[i]
            extra_assignments[uur].append(laatste_pv["naam"])
            afgekapte_pv_uren.add(uur)


def herbereken_afgekapte_pv_uren(absentees_set=None, base_maps=None):
    """
    Herleidt afgekapte_pv_uren op basis van de huidige selected-lijst.
    Raakt extra_assignments niet aan.
    In last-minute context: geef absentees_set en base_maps mee voor
    correcte pauzetelling (i.p.v. stale BP2/BQ2 cellen).
    """
    global afgekapte_pv_uren
    afgekapte_pv_uren = set()

    _aantal_pv = len(selected)
    _aantal_pauze_uren = len(required_pauze_hours)
    if _aantal_pv == 0 or _aantal_pauze_uren == 0:
        return

    _plaatsen = (_aantal_pauze_uren * 4 - 1) * _aantal_pv

    if absentees_set is not None and base_maps is not None:
        _lange, _korte = lm5_bereken_pauze_counts(absentees_set, base_maps)
    else:
        try:
            _lange = int(ws_speciaal.cell(15, 6).value) if ws_speciaal.cell(15, 6).value else 0
        except:
            _lange = 0
        try:
            _korte = int(ws_speciaal.cell(16, 6).value) if ws_speciaal.cell(16, 6).value else 0
        except:
            _korte = 0

    _open_spots = _plaatsen - (2 * _lange + _korte)
    _min_open_spots_per_pv = 1 if len(open_uren) <= 6 else 3

    _afknip_kandidaat_herb = pp2_bepaal_pv_voor_afknip(selected)

    _marge_tiers = [
        (0, 4),
        (2, 2),
        (_min_open_spots_per_pv, 0),
    ]

    for _marge_kandidaat, _vereiste_uren in _marge_tiers:
        _marge_totaal = 0
        for _pv_marge in selected:
            if _afknip_kandidaat_herb is not None and _pv_marge["naam"] == _afknip_kandidaat_herb["naam"]:
                _marge_totaal += _marge_kandidaat
            else:
                _marge_totaal += _min_open_spots_per_pv

        _beschikbaar = _open_spots - _marge_totaal
        if _beschikbaar >= 3:
            _overbodige = 1 + max(0, math.floor((_beschikbaar - 3) / 4))
        else:
            _overbodige = 0

        _pv_pauze_uren = sorted(required_pauze_hours, reverse=True)
        _uren_te_verschuiven = min(_overbodige, len(_pv_pauze_uren))

        if _uren_te_verschuiven >= _vereiste_uren:
            break

    if _overbodige > 0:

        if _uren_te_verschuiven < len(_pv_pauze_uren):
            _aantal_pv_zonder_laatste = _aantal_pv - 1
            if _aantal_pv_zonder_laatste > 0:
                _plaatsen_zonder_laatste = (_aantal_pauze_uren * 4 - 1) * _aantal_pv_zonder_laatste
                _marge_zonder_laatste = _aantal_pv_zonder_laatste * _min_open_spots_per_pv
                _nodig_zonder_laatste = (2 * _lange + _korte) + _marge_zonder_laatste
                if _plaatsen_zonder_laatste >= _nodig_zonder_laatste:
                    _uren_te_verschuiven = len(_pv_pauze_uren)

        for uur in _pv_pauze_uren[:_uren_te_verschuiven]:
            afgekapte_pv_uren.add(uur)




MAX_CONSEC = 4
MAX_PER_STUDENT_ATTR = 6


# -----------------------------
# Vaste dag-attracties (BG–BI)
# -----------------------------

vaste_plaatsingen = []  # lijst van dicts: {naam, attractie}

for rij in range(3, 6):  # R3:T5 in Aanpassingen
    if ws_aanpassingen.cell(rij, 18).value in [1, True, "WAAR", "X"]:  # kolom R
        naam = ws_aanpassingen.cell(rij, 19).value       # kolom S
        attractie = ws_aanpassingen.cell(rij, 20).value  # kolom T
        if naam and attractie:
            vaste_plaatsingen.append({
                "naam": str(naam).strip(),
                "attractie": str(attractie).strip()
            })

# -----------------------------
# Vaste plaatsingen toepassen
# -----------------------------

for vp in vaste_plaatsingen:
    student = next((s for s in studenten if s["naam"] == vp["naam"]), None)
    if not student:
        continue

    attr = vp["attractie"]

    # effectieve werkuren van deze student
    uren = [
        u for u in student["uren_beschikbaar"]
        if u in open_uren
        and not (student["is_pauzevlinder"] and u in required_pauze_hours)
    ]

    for uur in uren:
        # attractie moet dit uur actief zijn
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            continue

        # rode attracties overslaan
        if attr in red_spots.get(uur, set()):
            continue

        # capaciteit check
        max_spots = aantallen[uur].get(attr, 1)
        if attr in second_spot_blocked.get(uur, set()):
            max_spots = 1

        if per_hour_assigned_counts[uur][attr] >= max_spots:
            continue

        # plaats student
        assigned_map[(uur, attr)].append(student["naam"])
        per_hour_assigned_counts[uur][attr] += 1
        student["assigned_hours"].append(uur)
        student["assigned_attracties"].add(attr)

    # student mag niet meer door de normale planner
    student["uren_beschikbaar"] = []





def student_tie_break_key(student):
    return naam_tie_break_key(student["naam"])

def is_werkende_pauzevlinder(s):
    """Geeft 1 als de student een pauzevlinder is die nog wél werkt, anders 0."""
    if s["is_pauzevlinder"] and any(u in open_uren for u in s["uren_beschikbaar"]):
        return 1
    return 0

studenten_sorted = sorted(
    studenten_workend,
    key=lambda s: (is_werkende_pauzevlinder(s), s["aantal_attracties"], student_tie_break_key(s))
)

# -----------------------------
# Voorbereiden: expand naar posities per uur
# -----------------------------
positions_per_hour = {uur: [] for uur in open_uren}
for uur in open_uren:
    for attr in actieve_attracties_per_uur[uur]:
        max_pos = aantallen[uur].get(attr, 1)
        for pos in range(1, max_pos+1):
            # sla rode posities over
            if attr in second_spot_blocked[uur] and pos == 2:
                continue
            positions_per_hour[uur].append((attr, pos))
# -----------------------------
# occupied_positions vullen op basis van bestaande assigned_map
# -----------------------------
occupied_positions = {uur: {} for uur in open_uren}

for (uur, attr), namen in assigned_map.items():
    for idx, naam in enumerate(namen, start=1):
        occupied_positions[uur][(attr, idx)] = naam


# -----------------------------
# Hulpfunctie: kan blok geplaatst worden?
# -----------------------------
def can_place_block(student, block_hours, attr):
    for h in block_hours:
        # check of attractie beschikbaar is in dit uur
        if (attr, 1) not in positions_per_hour[h] and (attr, 2) not in positions_per_hour[h]:
            return False
        # alle posities bezet?
        taken1 = (attr,1) in occupied_positions[h]
        taken2 = (attr,2) in occupied_positions[h]
        if taken1 and taken2:
            return False
    return True

# -----------------------------
# Plaats blok
# -----------------------------
def place_block(student, block_hours, attr):
    for h in block_hours:
        # kies positie: eerst pos1, anders pos2
        if (attr,1) in positions_per_hour[h] and (attr,1) not in occupied_positions[h]:
            pos = 1
        else:
            pos = 2
        occupied_positions[h][(attr,pos)] = student["naam"]
        assigned_map[(h, attr)].append(student["naam"])
        student["assigned_hours"].append(h)
    student["assigned_attracties"].add(attr)


# =============================
# FLEXIBELE BLOKKEN & PLAATSLOGICA
# =============================

def student_kan_attr(student, attr):
    if " + " not in attr:
        # check blacklist
        if attr.lower() in student_blacklist.get(student["naam"], set()):
            return False
        return attr in student["attracties"]
    onderdelen = [a.strip() for a in attr.split("+")]
    # check elk onderdeel tegen blacklist
    for o in onderdelen:
        if o.lower() in student_blacklist.get(student["naam"], set()):
            return False
    return all(o in student["attracties"] for o in onderdelen)


def _max_spots_for(attr, uur):
    """Houd rekening met red_spots: 2e plek dicht als het rood is."""
    max_spots = aantallen[uur].get(attr, 1)
    if attr in second_spot_blocked.get(uur, set()):
        max_spots = 1
    return max_spots

def _has_capacity(attr, uur):
    if attr in red_spots.get(uur, set()):
        return False
    return per_hour_assigned_counts[uur][attr] < _max_spots_for(attr, uur)


def uren_bij_basis_attr(student, attr):
    """
    Tel alle uren die een student bij een basisattractie heeft gestaan,
    inclusief uren bij de samengevoegde versie of losse onderdelen.
    - "A"   → telt ook uren bij "A+B"
    - "A+B" → telt ook uren bij "A" en bij "B"
    """
    gerelateerd = {attr}

    if " + " in attr:
        for onderdeel in attr.split(" + "):
            gerelateerd.add(onderdeel.strip())
    else:
        for sameng in samengevoegde_attracties:
            onderdelen = [o.strip() for o in sameng.split(" + ")]
            if attr.strip() in onderdelen:
                gerelateerd.add(sameng)

    uren = set()
    for h in student["assigned_hours"]:
        for rel_attr in gerelateerd:
            if student["naam"] in assigned_map.get((h, rel_attr), []):
                uren.add(h)
                break
    return uren


def _try_place_block_on_attr(student, block_hours, attr):
    # Capaciteit check
    for h in block_hours:
        if not _has_capacity(attr, h):
            return False

    # ── AANGEPAST: tel ook uren bij verwante attracties mee ──
    uren_bij_attr = uren_bij_basis_attr(student, attr)

    # Check max 6 unieke uren per basisattractie per dag
    nieuwe_uren = set(block_hours)
    totaal_uren = uren_bij_attr | nieuwe_uren
    if len(totaal_uren) > 6:
        return False

    # Plaatsen
    for h in block_hours:
        assigned_map[(h, attr)].append(student["naam"])
        per_hour_assigned_counts[h][attr] += 1
        student["assigned_hours"].append(h)

    student["assigned_attracties"].add(attr)
    return True



def _try_place_block_any_attr(student, block_hours):
    def candidate_score(attr):
        schaarste = sum(1 for s in studenten_workend if attr in s["attracties"])

        # ── AANGEPAST: tel ook uren bij verwante attracties mee ──
        bestaande_uren = uren_bij_basis_attr(student, attr)
        totaal_na_plaatsing = len(bestaande_uren | set(block_hours))
        reeds_gebruikt = attr in student["assigned_attracties"] or bool(bestaande_uren)

        breedte_profiel = student.get("aantal_attracties", len(student.get("attracties", [])))

        fairness_straf = 0
        if totaal_na_plaatsing > 4:
            if breedte_profiel >= 6:    fairness_straf = 100
            elif breedte_profiel >= 5:  fairness_straf = 60
            elif breedte_profiel >= 4:  fairness_straf = 25

        hergebruik_straf = 1 if reeds_gebruikt else 0
        huidige_uren_op_attr = len(bestaande_uren)

        return (fairness_straf, hergebruik_straf, huidige_uren_op_attr, schaarste, attr)

    candidate_attrs = [
        a for a in attracties_te_plannen
        if student_kan_attr(student, a)
    ]
    candidate_attrs.sort(key=candidate_score)

    for attr in candidate_attrs:
        if _try_place_block_on_attr(student, block_hours, attr):
            return True

    return False
    


def _try_place_block_samenvoeging_transitie(student, block_hours, respecteer_fairness=True):
    if len(block_hours) < 2:
        return False

    breedte_profiel = student.get("aantal_attracties", len(student.get("attracties", [])))

    def fairness_ok(basis_attr, nieuwe_uren):
        if not respecteer_fairness:
            return True
        bestaande = uren_bij_basis_attr(student, basis_attr)
        totaal = len(bestaande | set(nieuwe_uren))
        if totaal <= 4:
            return True
        # Zelfde drempels als candidate_score
        if breedte_profiel >= 6: return False
        if breedte_profiel >= 5: return False
        if breedte_profiel >= 4: return False
        return True  # weinig opties → toch toestaan

    # ── Geval 1: eerste uur/uren zijn samenvoeging ──
    eerste_uur = block_hours[0]
    for sameng_attr in actieve_attracties_per_uur.get(eerste_uur, set()):
        if " + " not in sameng_attr:
            continue
        if not student_kan_attr(student, sameng_attr):
            continue

        # Hoeveel opeenvolgende uren is de samenvoeging actief vanaf het begin?
        sameng_uren = []
        for h in block_hours:
            if sameng_attr in actieve_attracties_per_uur.get(h, set()):
                sameng_uren.append(h)
            else:
                break

        rest_uren = [h for h in block_hours if h not in sameng_uren]
        if not rest_uren:
            continue

        if not all(_has_capacity(sameng_attr, h) for h in sameng_uren):
            continue

        onderdelen = [o.strip() for o in sameng_attr.split("+")]

        for onderdeel in onderdelen:
            if not student_kan_attr(student, onderdeel):
                continue
            if not all(_has_capacity(onderdeel, h) for h in rest_uren):
                continue

            if len(uren_bij_basis_attr(student, onderdeel) | set(block_hours)) > 6:
                continue
            if not fairness_ok(onderdeel, block_hours):
                continue

            for h in sameng_uren:
                assigned_map[(h, sameng_attr)].append(student["naam"])
                per_hour_assigned_counts[h][sameng_attr] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(sameng_attr)

            for h in rest_uren:
                assigned_map[(h, onderdeel)].append(student["naam"])
                per_hour_assigned_counts[h][onderdeel] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(onderdeel)

            return True

    # ── Geval 2: laatste uur/uren zijn samenvoeging ──
    laatste_uur = block_hours[-1]
    for sameng_attr in actieve_attracties_per_uur.get(laatste_uur, set()):
        if " + " not in sameng_attr:
            continue
        if not student_kan_attr(student, sameng_attr):
            continue

        sameng_uren = []
        for h in reversed(block_hours):
            if sameng_attr in actieve_attracties_per_uur.get(h, set()):
                sameng_uren.insert(0, h)
            else:
                break

        vroege_uren = [h for h in block_hours if h not in sameng_uren]
        if not vroege_uren:
            continue

        if not all(_has_capacity(sameng_attr, h) for h in sameng_uren):
            continue

        onderdelen = [o.strip() for o in sameng_attr.split("+")]

        for onderdeel in onderdelen:
            if not student_kan_attr(student, onderdeel):
                continue
            if not all(_has_capacity(onderdeel, h) for h in vroege_uren):
                continue

            # ── AANGEPAST: gebruik uren_bij_basis_attr + fairness check ──
            if len(uren_bij_basis_attr(student, onderdeel) | set(block_hours)) > 6:
                continue
            if not fairness_ok(onderdeel, block_hours):
                continue

            for h in vroege_uren:
                assigned_map[(h, onderdeel)].append(student["naam"])
                per_hour_assigned_counts[h][onderdeel] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(onderdeel)

            for h in sameng_uren:
                assigned_map[(h, sameng_attr)].append(student["naam"])
                per_hour_assigned_counts[h][sameng_attr] += 1
                student["assigned_hours"].append(h)
            student["assigned_attracties"].add(sameng_attr)

            return True

    return False


def _place_block_with_fallback(student, hours_seq, preferred_sizes=None, reset_sizes=None):
    if not hours_seq:
        return []

    if preferred_sizes is None:
        preferred_sizes = [3, 2, 4, 1]

    # Na het eerste blok terugvallen op reset_sizes (of gewone volgorde als geen reset opgegeven)
    next_sizes = reset_sizes if reset_sizes is not None else preferred_sizes

    for size in preferred_sizes:
        if len(hours_seq) < size:
            continue
        block = hours_seq[:size]

        eerste_uur = block[0]
        laatste_uur = block[-1]

        heeft_samenvoeging = size > 1 and (
            any(" + " in attr and student_kan_attr(student, attr)
                for attr in actieve_attracties_per_uur.get(eerste_uur, set()))
            or any(" + " in attr and student_kan_attr(student, attr)
                   for attr in actieve_attracties_per_uur.get(laatste_uur, set()))
        )

        if heeft_samenvoeging:
            sameng_kandidaten = [
                attr
                for uur in [eerste_uur, laatste_uur]
                for attr in actieve_attracties_per_uur.get(uur, set())
                if " + " in attr and student_kan_attr(student, attr)
            ]
            beste_sameng_score = (
                min(kritieke_score(a, studenten_workend) for a in sameng_kandidaten)
                if sameng_kandidaten else float("inf")
            )
            reguliere_kandidaten = [
                a for a in attracties_te_plannen
                if student_kan_attr(student, a)
                and all(_has_capacity(a, h) for h in block)
            ]
            beste_regulier_score = (
                min(kritieke_score(a, studenten_workend) for a in reguliere_kandidaten)
                if reguliere_kandidaten else float("inf")
            )

            if beste_sameng_score <= beste_regulier_score:
                # Stap 1: transitie mét fairness
                if _try_place_block_samenvoeging_transitie(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 2: normaal mét fairness
                if _try_place_block_any_attr(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 3: transitie zónder fairness (laatste redmiddel voor deze size)
                if _try_place_block_samenvoeging_transitie(student, block, respecteer_fairness=False):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
            else:
                # Stap 1: normaal mét fairness
                if _try_place_block_any_attr(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 2: transitie mét fairness
                if _try_place_block_samenvoeging_transitie(student, block):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
                # Stap 3: transitie zónder fairness (laatste redmiddel voor deze size)
                if _try_place_block_samenvoeging_transitie(student, block, respecteer_fairness=False):
                    return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)
        else:
            # Stap 1: normaal mét fairness
            if _try_place_block_any_attr(student, block):
                return _place_block_with_fallback(student, hours_seq[size:], preferred_sizes=next_sizes)

    # Noodventiel: eerste uur tijdelijk overslaan
    return [hours_seq[0]] + _place_block_with_fallback(student, hours_seq[1:], preferred_sizes=next_sizes)



    
# -----------------------------
# Nieuwe assign_student
# -----------------------------


def assign_student(s):
    uren = sorted(u for u in s["uren_beschikbaar"] if u in open_uren)
    if s["is_pauzevlinder"]:
        uren = [u for u in uren if u not in required_pauze_hours]

    if not uren:
        return

    runs = contiguous_runs(uren)

    for run in runs:
        blokken = partition_run_lengths(run, ideal_moments=ideaalmomenten)

        idx = 0
        for b in blokken:
            block_hours = run[idx: idx + b]
            idx += b
            degr = list(range(b, 0, -1))          # b=3 -> [3, 2, 1]
            unplaced = _place_block_with_fallback(
                s, block_hours, preferred_sizes=degr, reset_sizes=degr
            )
            for h in unplaced:
                extra_assignments[h].append(s["naam"])


for s in studenten_sorted:
    assign_student(s)
    
# -----------------------------
# Post-processing: lege plekken opvullen door doorschuiven
# -----------------------------

def doorschuif_leegplek(uur, attr, pos_idx, student_naam, stap, max_stappen=5):
    if stap > max_stappen:
        return False
    namen = assigned_map.get((uur, attr), [])
    naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
    if naam:
        return False

    kandidaten = []
    for b_attr in attracties_te_plannen:
        b_namen = assigned_map.get((uur, b_attr), [])
        for b_pos, b_naam in enumerate(b_namen):
            if not b_naam or b_naam == student_naam:
                continue
            cand_student = next((s for s in studenten_workend if s["naam"] == b_naam), None)
            if not cand_student:
                continue
            # Mag deze student de lege attractie doen?
            if attr not in cand_student["attracties"]:
                continue
            # Mag de extra de vrijgekomen plek doen?
            extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
            if not extra_student:
                continue
            if b_attr not in extra_student["attracties"]:
                continue
            # Score: zo min mogelijk 1-uursblokken creëren
            uren_cand = sorted([u for u in cand_student["assigned_hours"] if u != uur] + [uur])
            uren_extra = sorted(extra_student["assigned_hours"] + [uur])
            def count_1u_blokken(uren):
                if not uren:
                    return 0
                runs = contiguous_runs(uren)
                return sum(1 for r in runs if len(r) == 1)
            score = count_1u_blokken(uren_cand) + count_1u_blokken(uren_extra)
            kandidaten.append((score, b_attr, b_pos, b_naam, cand_student))
    kandidaten.sort()

    for score, b_attr, b_pos, b_naam, cand_student in kandidaten:
        extra_student = next((s for s in studenten_workend if s["naam"] == student_naam), None)
        if not extra_student:
            continue
        # Voer de swap uit
        assigned_map[(uur, b_attr)][b_pos] = student_naam
        extra_student["assigned_hours"].append(uur)
        extra_student["assigned_attracties"].add(b_attr)
        per_hour_assigned_counts[uur][b_attr] += 0  # netto gelijk
        assigned_map[(uur, attr)].insert(pos_idx-1, b_naam)
        assigned_map[(uur, attr)] = assigned_map[(uur, attr)][:aantallen[uur].get(attr, 1)]
        cand_student["assigned_hours"].remove(uur)
        cand_student["assigned_attracties"].discard(b_attr)
        cand_student["assigned_hours"].append(uur)
        cand_student["assigned_attracties"].add(attr)
        per_hour_assigned_counts[uur][attr] += 0  # netto gelijk
        # Check of alles klopt (geen dubbele, geen restricties overtreden)
        # (optioneel: extra checks toevoegen)
        return True
    return False

max_iterations = 10
for _ in range(max_iterations):
    changes_made = False
    for uur in open_uren:
        for attr in actieve_attracties_per_uur[uur]:
            max_pos = aantallen[uur].get(attr, 1)
            if attr in red_spots.get(uur, set()):
                max_pos = 1
            for pos_idx in range(1, max_pos+1):
                namen = assigned_map.get((uur, attr), [])
                naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""
                if naam:
                    continue
                # Probeer voor alle extra's op dit uur
                extras_op_uur = list(extra_assignments[uur])  # kopie ivm mutatie
                for extra_naam in extras_op_uur:
                    extra_student = next((s for s in studenten_workend if s["naam"] == extra_naam), None)
                    if not extra_student:
                        continue
                    if attr in extra_student["attracties"]:
                        # Kan direct geplaatst worden, dus hoort niet bij dit scenario
                        continue
                    # Probeer doorschuiven
                    if doorschuif_leegplek(uur, attr, pos_idx, extra_naam, 1, max_iterations):
                        extra_assignments[uur].remove(extra_naam)
                        changes_made = True
                        break  # stop met deze plek, ga naar volgende lege plek
    if not changes_made:
        break



# -----------------------------
# Post-processing: wissel laatste blok van 2 of 3 uren
# als iemand 5 of 6 uur op 1 attractie staat
# -----------------------------

vaste_studenten = {vp["naam"] for vp in vaste_plaatsingen}

def get_student_by_name(naam):
    return next((s for s in studenten_workend if s["naam"] == naam), None)

def get_student_attr_on_hour(student_naam, uur):
    for attr in actieve_attracties_per_uur.get(uur, set()):
        if student_naam in assigned_map.get((uur, attr), []):
            return attr
    return None

def get_hours_on_attr(student, attr):
    uren = []
    for uur in sorted(set(student["assigned_hours"])):
        if student["naam"] in assigned_map.get((uur, attr), []):
            uren.append(uur)
    return sorted(uren)

def get_runs_on_attr(student, attr):
    uren = get_hours_on_attr(student, attr)
    return contiguous_runs(uren)

def count_attr_switches(student):
    uur_attr = []
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            uur_attr.append((uur, attr))

    if not uur_attr:
        return 0

    switches = 0
    prev_attr = uur_attr[0][1]
    for _, attr in uur_attr[1:]:
        if attr != prev_attr:
            switches += 1
        prev_attr = attr
    return switches

def remove_assignment(student, uur, attr):
    namen = assigned_map.get((uur, attr), [])
    if student["naam"] in namen:
        namen.remove(student["naam"])
    if uur in student["assigned_hours"]:
        student["assigned_hours"].remove(uur)

def add_assignment(student, uur, attr):
    assigned_map[(uur, attr)].append(student["naam"])
    student["assigned_hours"].append(uur)
    student["assigned_attracties"].add(attr)

def rebuild_student_attrs(student):
    attrs = set()
    for uur in sorted(set(student["assigned_hours"])):
        attr = get_student_attr_on_hour(student["naam"], uur)
        if attr:
            attrs.add(attr)
    student["assigned_attracties"] = attrs

def is_valid_attr_for_student_on_hours(student, attr, uren):
    # vaste dagplaatsingen niet aanpassen
    if student["naam"] in vaste_studenten:
        return False

    # student moet attractie kunnen doen
    if not student_kan_attr(student, attr):
        return False

    # attractie moet op al die uren actief en geldig zijn
    for uur in uren:
        if attr not in actieve_attracties_per_uur.get(uur, set()):
            return False
        if attr in red_spots.get(uur, set()):
            return False

    return True

def respects_student_attr_rules(student, attr):
    uren = get_hours_on_attr(student, attr)
    if len(uren) > 6:
        return False
    return True

def can_swap_exact_block(student_a, attr_a, block_hours, student_b, attr_b):
    # zelfde student of zelfde attractie heeft geen zin
    if student_a["naam"] == student_b["naam"]:
        return False
    if attr_a == attr_b:
        return False

    # beide richtingen moeten kunnen
    if not is_valid_attr_for_student_on_hours(student_a, attr_b, block_hours):
        return False
    if not is_valid_attr_for_student_on_hours(student_b, attr_a, block_hours):
        return False

    # student_b moet op exact deze uren ook éénzelfde blok hebben op attr_b
    for uur in block_hours:
        if student_b["naam"] not in assigned_map.get((uur, attr_b), []):
            return False
        # en niet tegelijk nog ergens anders zitten
        current_attr = get_student_attr_on_hour(student_b["naam"], uur)
        if current_attr != attr_b:
            return False

    # student_a moet natuurlijk ook exact daar staan
    for uur in block_hours:
        if student_a["naam"] not in assigned_map.get((uur, attr_a), []):
            return False
        current_attr = get_student_attr_on_hour(student_a["naam"], uur)
        if current_attr != attr_a:
            return False

    return True

def count_problem_attrs(student):
    """
    Tel voor hoeveel attracties deze student meer dan 4 uur ingepland staat.
    """
    count = 0
    for attr in list(student["assigned_attracties"]):
        if len(get_hours_on_attr(student, attr)) > 4:
            count += 1
    return count

def total_overflow_hours(student):
    """
    Tel hoeveel uren boven de limiet van 4 uur deze student in totaal heeft.
    Voorbeeld:
    - 5 uur op een attractie => +1
    - 6 uur op een attractie => +2
    """
    overflow = 0
    for attr in list(student["assigned_attracties"]):
        uren = len(get_hours_on_attr(student, attr))
        if uren > 4:
            overflow += (uren - 4)
    return overflow

def can_use_block_as_swap_target(student, attr, block_hours):
    """
    Check of student op exact deze uren op exact dezelfde attractie staat.
    """
    for uur in block_hours:
        if student["naam"] not in assigned_map.get((uur, attr), []):
            return False
        huidige_attr = get_student_attr_on_hour(student["naam"], uur)
        if huidige_attr != attr:
            return False
    return True

def try_swap_specific_block(student, attr, block_hours, sta_toe_overflow_andere=False, voorkeur_nieuwe_attractie=False):
    """
    Probeer één specifiek blok (eerste OF laatste) van student/attr te wisselen.
    Alleen als:
    - het blok 2 of 3 uur lang is
    - de andere student op exact die uren ook één blok op één attractie heeft
    - alle regels geldig blijven
    - geen geïsoleerde 1-uursblokken ontstaan
    - max 2 extra wissels in totaal
    - het totaal aantal >4u-problemen niet stijgt (of, als sta_toe_overflow_andere
      True is: enkel als laatste redmiddel, en enkel als het bij de andere
      student opgesplitst blijft i.p.v. een nieuw lang aaneengesloten blok)

    voorkeur_nieuwe_attractie: als True, worden kandidaten die voor BEIDE studenten
    een attractie opleveren die ze die dag nog niet deden eerst geprobeerd. Andere
    (niet-ideale) kandidaten blijven nog steeds mogelijk als fallback, enkel later
    in de volgorde.
    """
    if len(block_hours) not in [2, 3]:
        return False

    orig_switches_a = count_attr_switches(student)
    orig_problem_count_a = count_problem_attrs(student)
    orig_overflow_a = total_overflow_hours(student)
    orig_max_blok_a = max_consecutive_hours(get_hours_on_attr(student, attr))

    eerste_uur = block_hours[0]
    kandidaten = []

    for andere_student in studenten_workend:
        if andere_student["naam"] == student["naam"]:
            continue
        if andere_student["naam"] in vaste_studenten:
            continue

        attr_b = get_student_attr_on_hour(andere_student["naam"], eerste_uur)
        if not attr_b or attr_b == attr:
            continue

        # Andere student moet exact op dit hele blok op dezelfde attractie staan
        if not can_use_block_as_swap_target(andere_student, attr_b, block_hours):
            continue

        # Beide studenten moeten elkaars attractie op die uren mogen doen
        if not is_valid_attr_for_student_on_hours(student, attr_b, block_hours):
            continue
        if not is_valid_attr_for_student_on_hours(andere_student, attr, block_hours):
            continue

        kandidaten.append((andere_student["naam"], attr_b, andere_student))

    if voorkeur_nieuwe_attractie:
        # ideale kandidaten (nieuwe attractie voor BEIDE studenten die dag) eerst proberen;
        # de rest blijft achteraan in de lijst staan als fallback
        def _is_ideaal(kand):
            _, attr_b, andere_student = kand
            student_krijgt_nieuwe_attr = attr_b not in student["assigned_attracties"]
            andere_krijgt_nieuwe_attr = attr not in andere_student["assigned_attracties"]
            return 0 if (student_krijgt_nieuwe_attr and andere_krijgt_nieuwe_attr) else 1
        kandidaten.sort(key=_is_ideaal)

    for _, attr_b, andere_student in kandidaten:
        orig_switches_b = count_attr_switches(andere_student)
        orig_problem_count_b = count_problem_attrs(andere_student)
        orig_overflow_b = total_overflow_hours(andere_student)

        def aantal_geisoleerde_uren(s, a):
            runs = get_runs_on_attr(s, a)
            return sum(1 for r in runs if len(r) == 1)

        # Isolatie VOOR de wissel meten. Een al bestaand los uur, dat
        # niets met deze wissel te maken heeft, mag een verder prima
        # wissel niet blokkeren -- enkel een NIEUW los uur telt.
        orig_iso_student_attr  = aantal_geisoleerde_uren(student, attr)
        orig_iso_student_attrb = aantal_geisoleerde_uren(student, attr_b)
        orig_iso_andere_attr   = aantal_geisoleerde_uren(andere_student, attr)
        orig_iso_andere_attrb  = aantal_geisoleerde_uren(andere_student, attr_b)

        # --- tijdelijke swap uitvoeren ---
        for uur in block_hours:
            remove_assignment(student, uur, attr)
            remove_assignment(andere_student, uur, attr_b)

        for uur in block_hours:
            add_assignment(student, uur, attr_b)
            add_assignment(andere_student, uur, attr)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

        valid = True

        # Enkel weigeren bij een NIEUW geïsoleerd 1-uursblok
        if aantal_geisoleerde_uren(student, attr) > orig_iso_student_attr:
            valid = False
        if aantal_geisoleerde_uren(student, attr_b) > orig_iso_student_attrb:
            valid = False
        if aantal_geisoleerde_uren(andere_student, attr) > orig_iso_andere_attr:
            valid = False
        if aantal_geisoleerde_uren(andere_student, attr_b) > orig_iso_andere_attrb:
            valid = False
        # Regels voor beide studenten / beide attracties
        for s, a in [
            (student, attr),
            (student, attr_b),
            (andere_student, attr),
            (andere_student, attr_b),
        ]:
            if not respects_student_attr_rules(s, a):
                valid = False

        # Max 2 extra wissels in totaal
        new_switches_a = count_attr_switches(student)
        new_switches_b = count_attr_switches(andere_student)
        extra_wissels = (new_switches_a - orig_switches_a) + (new_switches_b - orig_switches_b)

        if extra_wissels > 2:
            valid = False

        # Problemen na swap
        new_problem_count_a = count_problem_attrs(student)
        new_problem_count_b = count_problem_attrs(andere_student)
        new_overflow_a = total_overflow_hours(student)
        new_overflow_b = total_overflow_hours(andere_student)
        new_max_blok_a = max_consecutive_hours(get_hours_on_attr(student, attr))

        orig_total_problem_count = orig_problem_count_a + orig_problem_count_b
        new_total_problem_count = new_problem_count_a + new_problem_count_b

        orig_total_overflow = orig_overflow_a + orig_overflow_b
        new_total_overflow = new_overflow_a + new_overflow_b

        # Is het eventuele nieuwe probleem bij de andere student wél opgesplitst
        # (geen nieuw lang aaneengesloten blok), en lost het bij 'student' het
        # aaneengesloten blok effectief op? Dan mag dit tellen als laatste redmiddel.
        max_blok_andere_op_attr = max_consecutive_hours(get_hours_on_attr(andere_student, attr))
        laatste_redmiddel_ok = (
            sta_toe_overflow_andere
            and new_max_blok_a <= 4
            and max_blok_andere_op_attr <= 4
        )

        # Geen nieuw probleem creëren, TENZIJ dit het laatste redmiddel is
        if new_total_problem_count > orig_total_problem_count and not laatste_redmiddel_ok:
            valid = False

        # Geen grotere overschrijding creëren, TENZIJ dit het laatste redmiddel is
        if (
            new_total_problem_count == orig_total_problem_count
            and new_total_overflow > orig_total_overflow
            and not laatste_redmiddel_ok
        ):
            valid = False

        # Moet minstens iets verbeteren
        verbetering = (
            (new_total_problem_count < orig_total_problem_count)
            or (
                new_total_problem_count == orig_total_problem_count
                and new_total_overflow < orig_total_overflow
            )
            or (
                new_total_problem_count == orig_total_problem_count
                and new_total_overflow == orig_total_overflow
                and new_max_blok_a < orig_max_blok_a
            )
            or laatste_redmiddel_ok
        )

        if not verbetering:
            valid = False

        if valid:
            return True

        # --- rollback ---
        for uur in block_hours:
            remove_assignment(student, uur, attr_b)
            remove_assignment(andere_student, uur, attr)

        for uur in block_hours:
            add_assignment(student, uur, attr)
            add_assignment(andere_student, uur, attr_b)

        rebuild_student_attrs(student)
        rebuild_student_attrs(andere_student)

    return False

def try_swap_last_or_first_block(student, attr):
    """
    Probeer het laatste of eerste blok op deze attractie te wisselen.
    Als een run langer is dan 3, pak dan de laatste 3 of laatste 2 uur eruit.
    Alleen relevant als student >4 uur op deze attractie staat.

    Fase 1: strikte ruil (mag NERGENS een nieuw >4u-probleem geven).
    Fase 2 (enkel als fase 1 nergens lukt): laatste redmiddel, waarbij de
    andere student wel over de 4 uur mag gaan, MITS opgesplitst (geen
    nieuw lang aaneengesloten blok bij hem/haar).
    """
    uren_op_attr = get_hours_on_attr(student, attr)
    if len(uren_op_attr) <= 4:
        return False

    runs = get_runs_on_attr(student, attr)
    if not runs:
        return False

    laatste_run = runs[-1]
    eerste_run  = runs[0]

    def kandidaat_blokken(run):
        """Geef blokken van 2 of 3 uur terug vanuit deze run (einde eerst)."""
        blokken = []
        if len(run) >= 3:
            blokken.append(run[-3:])  # laatste 3
        if len(run) >= 2:
            blokken.append(run[-2:])  # laatste 2
        if len(run) >= 3:
            blokken.append(run[:3])   # eerste 3
        if len(run) >= 2:
            blokken.append(run[:2])   # eerste 2
        return blokken

    # ── Fase 1: strikte ruil ──
    for blok in kandidaat_blokken(laatste_run):
        if try_swap_specific_block(student, attr, blok):
            return True

    if eerste_run != laatste_run:
        for blok in kandidaat_blokken(eerste_run):
            if try_swap_specific_block(student, attr, blok):
                return True

    # ── Fase 2: laatste redmiddel ──
    for blok in kandidaat_blokken(laatste_run):
        if try_swap_specific_block(student, attr, blok, sta_toe_overflow_andere=True):
            return True

    if eerste_run != laatste_run:
        for blok in kandidaat_blokken(eerste_run):
            if try_swap_specific_block(student, attr, blok, sta_toe_overflow_andere=True):
                return True

    return False


# Iteratief toepassen tot er niets meer verandert
max_block_swap_passes = 15
for _ in range(max_block_swap_passes):
    wijziging = False

    for student in studenten_workend:
        probleem_attracties = [
            a for a in list(student["assigned_attracties"])
            if len(get_hours_on_attr(student, a)) > 4
        ]

        # Eerst de zwaarste problemen proberen
        probleem_attracties.sort(
            key=lambda a: (
                -len(get_hours_on_attr(student, a)),
                -max(get_hours_on_attr(student, a))
            )
        )

        for attr in probleem_attracties:
            if try_swap_last_or_first_block(student, attr):
                wijziging = True
                break

    if not wijziging:
        break


# -----------------------------
# Post-processing (ENKEL bij FORCEER_EXHAUSTIEF, Instellingen!B2):
# probeer blokken van EXACT 4 uur aan één stuk ook op te splitsen in 2x 2 uur.
# Staat het vakje niet aan, dan verandert hier niets.
# -----------------------------
if FORCEER_EXHAUSTIEF:

    def try_split_exact_4h_block(student, attr):
        """
        Probeert een aaneengesloten blok van EXACT 4 uur op deze attractie op te
        splitsen door de eerste of laatste 2 uur te wisselen met een andere student.
        Voorkeur: de wissel geeft zowel 'student' als de wisselpartner een attractie
        die ze die dag nog niet gedaan hebben (via voorkeur_nieuwe_attractie=True).
        Lukt dat nergens, dan wordt elke andere geldige 2-uurs-wissel aanvaard --
        dat is nog steeds beter dan het blok van 4 uur aan één stuk te laten staan.
        """
        runs = get_runs_on_attr(student, attr)
        for run in runs:
            if len(run) != 4:
                continue
            for blok in (run[-2:], run[:2]):
                if try_swap_specific_block(student, attr, blok, voorkeur_nieuwe_attractie=True):
                    return True
        return False

    # Iteratief toepassen tot er niets meer verandert
    max_4u_split_passes = 15
    for _ in range(max_4u_split_passes):
        wijziging_4u = False

        for student in studenten_workend:
            vier_uur_attracties = [
                a for a in list(student["assigned_attracties"])
                if any(len(r) == 4 for r in get_runs_on_attr(student, a))
            ]

            for attr in vier_uur_attracties:
                if try_split_exact_4h_block(student, attr):
                    wijziging_4u = True
                    break

        if not wijziging_4u:
            break


# -----------------------------
# Volgorde attracties uit Input!BL16:BL33
# -----------------------------
input_volgorde = []
for rij in range(3, 21):  # C3:C20 in Aanpassingen
    waarde = ws_aanpassingen.cell(rij, 3).value  # kolom C
    if waarde:
        input_volgorde.append(str(waarde).strip())

# Prioriteit per naam = positie in de VOLLEDIGE Aanpassingen-lijst,
# ook al is die attractie vandaag niet los actief (enkel via een
# combinatie zoals "Klimmen + Archery").
prioriteit_per_naam = {naam: i for i, naam in enumerate(input_volgorde)}
default_prioriteit = len(input_volgorde) + 1000

def attractie_prioriteit(attr):
    attr_str = str(attr)
    if attr_str in prioriteit_per_naam:
        return prioriteit_per_naam[attr_str]
    if " + " in attr_str:
        onderdelen = [x.strip() for x in attr_str.split("+")]
        indices = [prioriteit_per_naam[o] for o in onderdelen if o in prioriteit_per_naam]
        if indices:
            # net na het laatste onderdeel in de volledige inputvolgorde
            return max(indices) + 0.5
    return default_prioriteit

# -----------------------------
# Alle attracties die minstens één keer actief zijn (voor output),
# gesorteerd op basis van de Aanpassingen-volgorde
# -----------------------------
alle_actieve_attracties = set()
for uur in open_uren:
    alle_actieve_attracties |= actieve_attracties_per_uur.get(uur, set())

alle_actieve_attracties = sorted(alle_actieve_attracties, key=attractie_prioriteit)

def stabiliseer_assigned_map_voor_output():
    gesorteerde_uren = sorted(open_uren)

    def blijft_aantal_uren(naam, attr, vanaf_uur):
        # aantal aaneensluitende uren vanaf vanaf_uur dat naam op deze attractie staat
        teller = 0
        for u in gesorteerde_uren:
            if u < vanaf_uur:
                continue
            if naam in assigned_map.get((u, attr), []):
                teller += 1
            else:
                break
        return teller

    for attr in alle_actieve_attracties:
        vorige = []  # namen van vorig uur op plek-volgorde: vorige[0]=plek1, [1]=plek2
        for uur in gesorteerde_uren:
            max_pos = aantallen[uur].get(attr, 1)
            if attr in second_spot_blocked.get(uur, set()):
                max_pos = 1
            namen = list(dict.fromkeys(n for n in assigned_map.get((uur, attr), []) if n))

            slots = [None] * max_pos
            # 1) wie vorig uur een plek had, houdt die plek (indien nog geldig & vrij)
            for naam in namen:
                if naam in vorige:
                    plek = vorige.index(naam)
                    if plek < max_pos and slots[plek] is None:
                        slots[plek] = naam
            # 2) nieuwe namen: wie het langst blijft eerst -> krijgt de laagste vrije plek
            nieuwe = [n for n in namen if n not in slots]
            nieuwe.sort(key=lambda n: -blijft_aantal_uren(n, attr, uur))
            for naam in nieuwe:
                for i in range(max_pos):
                    if slots[i] is None:
                        slots[i] = naam
                        break

            assigned_map[(uur, attr)] = [n for n in slots if n]
            vorige = slots

stabiliseer_assigned_map_voor_output()


# -----------------------------

# Excel output
# -----------------------------
wb_out = Workbook()
ws_out = wb_out.active
ws_out.title = "Planning"

gray_fill = PatternFill(start_color="808080", fill_type="solid")

# Witte fill voor headers en attracties
white_fill = PatternFill(start_color="FFFFFF", fill_type="solid")
pv_fill = PatternFill(start_color="FFF2CC", fill_type="solid")
extra_fill = PatternFill(start_color="FCE4D6", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# Felle, maar lichte pastelkleuren (gelijkmatige felheid, veel variatie)
studenten_namen = sorted({s["naam"] for s in studenten})
# Pauzevlinders krijgen ook een kleur uit het schema
alle_namen = studenten_namen + [pv for pv in pauzevlinder_namen if pv not in studenten_namen]
# Unieke kleuren genereren: als er te weinig kleuren zijn, maak er meer met lichte variatie
base_colors = [
    "FFB3BA", "FFDFBA", "FFFFBA", "BAFFC9", "BAE1FF", "E0BBE4", "957DAD", "D291BC", "FEC8D8", "FFDFD3",
    "B5EAD7", "C7CEEA", "FFDAC1", "E2F0CB", "F6DFEB", "F9E2AE", "B6E2D3", "B6D0E2", "F6E2B3", "F7C5CC",
    "F7E6C5", "C5F7D6", "C5E6F7", "F7F6C5", "F7C5F7", "C5C5F7", "C5F7F7", "F7C5C5", "C5F7C5", "F7E2C5",
    "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "F7F7C5", "C5F7F7", "F7C5F7", "C5C5F7", "F7C5C5",
    "C5F7C5", "F7E2C5", "E2F7C5", "C5F7E2", "E2C5F7", "C5E2F7", "F7C5E2", "E2C5F7", "C5F7E2", "E2F7C5"
]
import colorsys
def pastel_variant(hex_color, variant):
    # hex_color: 'RRGGBB', variant: int
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    # kleine variatie in lichtheid en saturatie
    l = min(1, l + 0.03 * (variant % 3))
    s = max(0.3, s - 0.04 * (variant % 5))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return f"{int(r2*255):02X}{int(g2*255):02X}{int(b2*255):02X}"

unique_colors = []
needed = len(alle_namen)
variant = 0
while len(unique_colors) < needed:
    for base in base_colors:
        if len(unique_colors) >= needed:
            break
        # voeg lichte variatie toe als nodig
        color = pastel_variant(base, variant) if variant > 0 else base
        if color not in unique_colors:
            unique_colors.append(color)
    variant += 1

student_kleuren = dict(zip(alle_namen, unique_colors))

ws_out.cell(1, 1, vandaag).font = Font(bold=True)
ws_out.cell(1, 1).fill = white_fill


for col_idx, uur in enumerate(sorted(open_uren), start=2):
    label = uur_labels.get(uur)
    header_text = f"{formatteer_uur(uur)} ({label})" if label else formatteer_uur(uur)
    ws_out.cell(1, col_idx, header_text).font = Font(bold=True)
    ws_out.cell(1, col_idx).fill = white_fill
    ws_out.cell(1, col_idx).alignment = center_align
    ws_out.cell(1, col_idx).border = thin_border

    

rij_out = 2
for attr in alle_actieve_attracties:
    # 1. Bepaal hoeveel rijen deze attractie nodig heeft (1 of 2 plekken)
    max_pos = max(
        max(aantallen[uur].get(attr, 1) for uur in open_uren),
        max(per_hour_assigned_counts[uur].get(attr, 0) for uur in open_uren)
    )

    for pos_idx in range(1, max_pos + 1):
        # --- LAYOUT: Naam gevolgd door spatie en nummer (zonder haakjes) ---
        display_name = f"{attr} {pos_idx}" if max_pos > 1 else attr
        ws_out.cell(rij_out, 1, display_name).font = Font(bold=True)
        ws_out.cell(rij_out, 1).fill = white_fill
        ws_out.cell(rij_out, 1).border = thin_border

        for col_idx, uur in enumerate(sorted(open_uren), start=2):
            cell = ws_out.cell(rij_out, col_idx)

            # Haal de studentnaam op voor dit uur en deze positie
            namen = assigned_map.get((uur, attr), [])
            naam = namen[pos_idx-1] if pos_idx-1 < len(namen) else ""

            # --- LOGICA VOOR GRIJS KLEUREN ---
            current_attr_norm = normalize_attr(attr)
            is_samengesteld = " + " in attr
            groepen_dit_uur = uur_samenvoegingen.get(uur, [])
            
            moet_grijs = False

            # A. Check of de attractie dit uur gesloten is
            if uur in dichte_uren_per_attr.get(current_attr_norm, set()):
                moet_grijs = True

            # B. Check voor samengestelde attracties (bv. 'A + B')
            elif is_samengesteld:
                # De samengevoegde rij is grijs als deze specifieke groep dit uur NIET actief is
                onderdelen_set = {normalize_attr(x.strip()) for x in attr.split("+")}
                actief_als_groep = any({normalize_attr(g) for g in groep} == onderdelen_set for groep in groepen_dit_uur)
                if not actief_als_groep:
                    moet_grijs = True

            # C. Check voor individuele attracties (bv. 'A')
            else:
                # De individuele rij wordt grijs als de attractie opgaat in een samenvoeging
                is_onderdeel_van_samenvoeging = any(current_attr_norm in [normalize_attr(g) for g in groep] for groep in groepen_dit_uur)
                if is_onderdeel_van_samenvoeging:
                    moet_grijs = True

            # D. Check of de tweede plek geblokkeerd is (red spots)
            if pos_idx == 2 and attr in second_spot_blocked.get(uur, set()):
                moet_grijs = True

            # --- Cel invullen en opmaken ---
            cell.value = naam
            cell.alignment = center_align
            cell.border = thin_border

            if moet_grijs:
                cell.fill = gray_fill  # Grijs uit je bronnen
            elif naam and naam in student_kleuren:
                cell.fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
            else:
                cell.fill = white_fill

        rij_out += 1
        
# Pauzevlinders
rij_out += 1
pauzevlinder_namen_sorted = [pv["naam"] for pv in selected]
_afgeknipte_pv_voor_weergave = pp2_bepaal_pv_voor_afknip(selected)
_afgeknipte_pv_naam_weergave = _afgeknipte_pv_voor_weergave["naam"] if _afgeknipte_pv_voor_weergave else None

for pv_idx, pvnaam in enumerate(pauzevlinder_namen_sorted, start=1):
    ws_out.cell(rij_out, 1, f"Pauzevlinder {pv_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        # Afgeknipte PV: toon niet bij de afgekapte uren
        if pvnaam == _afgeknipte_pv_naam_weergave and uur in afgekapte_pv_uren:
            naam = ""
        else:
            naam = pvnaam if uur in required_pauze_hours else ""
        ws_out.cell(rij_out, col_idx, naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if naam and naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid")
    rij_out += 1

# Extra's per rij
rij_out += 1
extras_flat = []
for uur in sorted(open_uren):
    for naam in extra_assignments[uur]:
        if naam not in extras_flat:
            extras_flat.append(naam)
for extra_idx, naam in enumerate(extras_flat, start=1):
    ws_out.cell(rij_out, 1, f"Extra {extra_idx}").font = Font(bold=True)
    ws_out.cell(rij_out, 1).fill = white_fill
    ws_out.cell(rij_out, 1).border = thin_border
    for col_idx, uur in enumerate(sorted(open_uren), start=2):
        # Toon naam alleen als deze extra op dit uur is ingepland
        cell_naam = naam if naam in extra_assignments[uur] else ""
        ws_out.cell(rij_out, col_idx, cell_naam).alignment = center_align
        ws_out.cell(rij_out, col_idx).border = thin_border
        if cell_naam and cell_naam in student_kleuren:
            ws_out.cell(rij_out, col_idx).fill = PatternFill(start_color=student_kleuren[cell_naam], fill_type="solid")
    rij_out += 1

# Kolombreedte
for col in range(1, len(open_uren) + 2):
    ws_out.column_dimensions[get_column_letter(col)].width = 18

# ---- student_totalen beschikbaar maken voor volgende delen ----
from collections import defaultdict
# Bouw col→uur mapping vanuit de headerrij van ws_out
_col_uur_out = {}
for _c in range(2, ws_out.max_column + 1):
    _uur = parse_header_uur(ws_out.cell(1, _c).value)
    if _uur is not None:
        _col_uur_out[_c] = _uur

student_totalen = defaultdict(float)
for row in ws_out.iter_rows(min_row=2):
    for cel in row[1:]:
        naam = cel.value
        if naam and str(naam).strip() != "":
            _uur = _col_uur_out.get(cel.column)
            student_totalen[str(naam).strip()] += blok_durations.get(_uur, 1.0)


# -----------------------------
# Analyse-sheet maken indien nodig
# Alleen als er nog extra's zijn terwijl er elders echte lege plekken zijn
# -----------------------------

def heeft_echte_lege_plek():
    """
    True als er minstens 1 echte lege plek bestaat op de planning:
    - attractie is actief op dat uur
    - niet gesloten / niet red spot
    - geen geblokkeerde 2e plek
    - plaats is binnen de capaciteit
    - er staat nog niemand op die plek
    """
    for uur in open_uren:
        for attr in actieve_attracties_per_uur.get(uur, set()):
            if attr in red_spots.get(uur, set()):
                continue

            max_pos = aantallen[uur].get(attr, 1)
            if attr in second_spot_blocked.get(uur, set()):
                max_pos = 1

            namen = assigned_map.get((uur, attr), [])
            for pos_idx in range(1, max_pos + 1):
                naam = namen[pos_idx - 1] if pos_idx - 1 < len(namen) else ""
                if not naam:
                    return True
    return False


def heeft_extra_studenten():
    return any(len(namen) > 0 for namen in extra_assignments.values())


def student_is_aanwezig_op_uur_zonder_pauzevlinder(student, uur):
    """
    Student telt mee in analyse voor dit uur als:
    - student effectief aanwezig is op dit uur
      (ingepland of extra)
    - en NIET als pauzevlinder bezig is op dit uur
    """
    naam = student["naam"]

    # Pauzevlinder tijdens pauzevlinderuur telt niet mee
    if student.get("is_pauzevlinder") and uur in required_pauze_hours:
        return False

    if uur in set(student.get("assigned_hours", [])):
        return True

    if naam in extra_assignments.get(uur, []):
        return True

    return False


def student_kan_attr_in_analyse(student, attr):
    """
    Voor analyse:
    - respecteer blacklist
    - samengevoegde attractie mag enkel als student alle onderdelen kan
    """
    naam = student["naam"]

    if " + " not in attr:
        return attr.lower() not in student_blacklist.get(naam, set()) and attr in student.get("attracties", [])

    onderdelen = [a.strip() for a in attr.split("+")]
    for onderdeel in onderdelen:
        if onderdeel.lower() in student_blacklist.get(naam, set()):
            return False

    return all(onderdeel in student.get("attracties", []) for onderdeel in onderdelen)


def actieve_analyse_attracties_op_uur(uur, actieve_set=None):
    """
    Geeft attracties terug in de volgorde van Aanpassingen!C3:C20,
    maar aangepast aan het specifieke uur:
    - losse attracties als ze actief zijn
    - samengevoegde attracties enkel als ze dat uur actief samengevoegd zijn
    Optioneel: geef actieve_set mee om de globale actieve_attracties_per_uur te overschrijven.
    """
    if actieve_set is None:
        actieve_set = actieve_attracties_per_uur.get(uur, set())

    input_volgorde_lokaal = []
    for rij in range(3, 21):  # C3:C20 in Aanpassingen
        attr = ws_aanpassingen.cell(rij, 3).value
        if attr:
            input_volgorde_lokaal.append(str(attr).strip())

    resultaat = []
    gebruikte = set()

    # Eerst gewone attracties in inputvolgorde
    for attr in input_volgorde_lokaal:
        if attr in actieve_set and attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

        # Kijk of een samengestelde attractie met dit onderdeel actief is op dit uur
        for actief_attr in actieve_set:
            if " + " not in str(actief_attr):
                continue
            onderdelen = [x.strip() for x in str(actief_attr).split("+")]
            if attr in onderdelen and actief_attr not in gebruikte:
                if all(o in input_volgorde_lokaal for o in onderdelen):
                    laatst_idx = max(input_volgorde_lokaal.index(o) for o in onderdelen)
                    huidig_idx = input_volgorde_lokaal.index(attr)
                    if huidig_idx == laatst_idx:
                        resultaat.append(actief_attr)
                        gebruikte.add(actief_attr)

    # Daarna nog eventuele actieve attracties die niet in BL-lijst zaten
    for attr in actieve_set:
        if attr not in gebruikte:
            resultaat.append(attr)
            gebruikte.add(attr)

    return resultaat


def maak_analyse_sheet(wb_arg, am_arg, ea_arg, st_arg, actieve_attracties_override=None):
    # Verwijder oud sheet
    if "Analyse" in wb_arg.sheetnames:
        del wb_arg["Analyse"]

    ws_analyse = wb_arg.create_sheet(title="Analyse")
    analyse_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    witte_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Gebruik override als die meegegeven is (last-minute), anders globale data
    def actieve_set_voor_uur(uur):
        if actieve_attracties_override is not None:
            return actieve_attracties_override.get(uur, set())
        return actieve_attracties_per_uur.get(uur, set())

    # Herdefinieer de hulpfunctie lokaal zodat ze de juiste data gebruikt
    def is_aanwezig(student, uur):
        naam = student["naam"]
        if student.get("is_pauzevlinder") and uur in required_pauze_hours:
            return False
        if uur in set(student.get("assigned_hours", [])):
            return True
        if naam in ea_arg.get(uur, []):
            return True
        return False

    titel = "Hier zie je per uur welke studenten aanwezig zijn en welke attracties ze kunnen:"
    ws_analyse.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    titel_cel = ws_analyse.cell(1, 1, titel)
    titel_cel.font = Font(bold=True, size=12)
    titel_cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    titel_cel.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    titel_cel.border = thin_border

    start_rij = 3
    for uur in sorted(open_uren):
        actieve_set = actieve_set_voor_uur(uur)
        analyse_attracties_uur = actieve_analyse_attracties_op_uur(uur, actieve_set)

        analyse_studenten_uur = sorted(
            [s for s in st_arg if is_aanwezig(s, uur)],
            key=lambda s: (
                sum(1 for attr in analyse_attracties_uur if student_kan_attr_in_analyse(s, attr)),
                naam_tie_break_key(s["naam"])
            )
        )

        if not analyse_studenten_uur or not analyse_attracties_uur:
            continue

        ws_analyse.cell(start_rij, 1, formatteer_uur(uur)).font = Font(bold=True)
        ws_analyse.cell(start_rij, 1).fill = analyse_header_fill
        ws_analyse.cell(start_rij, 1).alignment = center_align
        ws_analyse.cell(start_rij, 1).border = thin_border
        ws_analyse.cell(start_rij, 2, "Student").font = Font(bold=True)
        ws_analyse.cell(start_rij, 2).fill = analyse_header_fill
        ws_analyse.cell(start_rij, 2).alignment = center_align
        ws_analyse.cell(start_rij, 2).border = thin_border

        start_col_attr = 3
        for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
            cel = ws_analyse.cell(start_rij, idx, attr)
            cel.font = Font(bold=True)
            cel.fill = analyse_header_fill
            cel.alignment = center_align
            cel.border = thin_border

        rij = start_rij + 1
        for s in analyse_studenten_uur:
            naam = s["naam"]
            ws_analyse.cell(rij, 1, rij - start_rij).alignment = center_align
            ws_analyse.cell(rij, 1).border = thin_border
            ws_analyse.cell(rij, 1).fill = witte_fill
            naam_cel = ws_analyse.cell(rij, 2, naam)
            naam_cel.alignment = center_align
            naam_cel.border = thin_border
            student_fill = PatternFill(start_color=student_kleuren[naam], fill_type="solid") if naam in student_kleuren else witte_fill
            naam_cel.fill = student_fill

            for idx, attr in enumerate(analyse_attracties_uur, start=start_col_attr):
                cel = ws_analyse.cell(rij, idx)
                cel.alignment = center_align
                cel.border = thin_border
                cel.font = Font(color="000000")
                if student_kan_attr_in_analyse(s, attr):
                    cel.value = attr
                    cel.fill = student_fill
                else:
                    cel.value = ""
                    cel.fill = witte_fill
            rij += 1

        ws_analyse.column_dimensions["A"].width = 8
        ws_analyse.column_dimensions["B"].width = 24
        for idx in range(start_col_attr, start_col_attr + len(analyse_attracties_uur)):
            ws_analyse.column_dimensions[get_column_letter(idx)].width = 13.5
        start_rij = rij + 3


maak_analyse_sheet(wb_out, assigned_map, extra_assignments, studenten)

#DEEL 2
#oooooooooooooooooooo
#oooooooooooooooooooo

# -----------------------------
# DEEL 2: Pauzevlinder overzicht
# -----------------------------
ws_pauze = wb_out.create_sheet(title="Pauzevlinders")

light_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# -----------------------------
# Rij 1: Uren
# -----------------------------
# Gebruik compute_pauze_hours/open_uren als basis voor de pauzeplanning-urenrij
uren_rij1 = []
from datetime import datetime, timedelta
if required_pauze_hours:
    start_uur = min(required_pauze_hours)
    eind_uur = max(required_pauze_hours)
    tijd = datetime(2020,1,1,start_uur,0)
    # Laatste pauze mag een kwartier vóór het einde starten
    laatste_pauze = datetime(2020,1,1,eind_uur,30)
    while tijd <= laatste_pauze:
        uren_rij1.append(f"{tijd.hour}u" if tijd.minute==0 else f"{tijd.hour}u{tijd.minute:02d}")
        tijd += timedelta(minutes=15)
else:
    # fallback: gebruik open_uren
    for uur in sorted(open_uren):
        uren_rij1.append(formatteer_uur(uur))

# Schrijf uren in rij 1, start in kolom B
for col_idx, uur in enumerate(uren_rij1, start=2):
    c = ws_pauze.cell(1, col_idx, uur)
    c.fill = light_fill
    c.alignment = center_align
    c.border = thin_border

### Zet de datum van vandaag in cel A1 van de pauzeplanning
a1 = ws_pauze.cell(1, 1, vandaag)
a1.font = Font(bold=True)
a1.fill = light_fill
a1.alignment = center_align
a1.border = thin_border

# -----------------------------
# Pauzevlinders en namen
# -----------------------------
rij_out = 2
for pv_idx, pv in enumerate(selected, start=1):
    # Titel: Pauzevlinder X
    title_cell = ws_pauze.cell(rij_out, 1, f"Pauzevlinder {pv_idx}")
    title_cell.font = Font(bold=True)
    title_cell.fill = light_fill
    title_cell.alignment = center_align
    title_cell.border = thin_border

    # Naam eronder (zelfde stijl en kleur)
    naam_cel = ws_pauze.cell(rij_out + 1, 1, pv["naam"])
    naam_cel.fill = light_fill
    naam_cel.alignment = center_align
    naam_cel.border = thin_border

    rij_out += 3  # lege rij ertussen

# -----------------------------
# Kolombreedte voor overzicht
# -----------------------------

# Automatisch de breedte van kolom A instellen op basis van de langste tekst
max_len_colA = 0
for row in range(1, ws_pauze.max_row + 1):
    val = ws_pauze.cell(row, 1).value
    if val:
        max_len_colA = max(max_len_colA, len(str(val)))
# Voeg wat extra ruimte toe
ws_pauze.column_dimensions['A'].width = max(12, max_len_colA + 2)

for col in range(2, len(uren_rij1) + 2):
    ws_pauze.column_dimensions[get_column_letter(col)].width = 10

# Gebruik exact dezelfde open_uren en headers als in deel 1 voor de pauzeplanning
uren_rij1 = []
for uur in sorted(open_uren):
    # Zoek de originele header uit ws_out (de hoofdplanning)
    for col in range(2, ws_out.max_column + 1):
        header = ws_out.cell(1, col).value
        if header and str(header).startswith(str(uur)):
            uren_rij1.append(header)
            break

# Opslaan met dezelfde unieke naam

# Maak in-memory bestand
output = BytesIO()





#oooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


#DEEL 3
#oooooooooooooooooooo
#oooooooooooooooooooo

import random
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side, PatternFill
import datetime

# -----------------------------
# DEEL 3: Extra naam voor pauzevlinders die langer dan 6 uur werken
# -----------------------------

# Sheet referenties
ws_planning = wb_out["Planning"]

def _get_theo_werkuren(naam):
    """Theoretische werkuren van een student, afgelezen uit het Planning-sheet."""
    uren = set()
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        uur = parse_header_uur(header)
        if uur is None:
            continue
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                uren.add(uur)
                break
    return sorted(uren)

def werkduur_voor_pauze(naam):
    """
    Geeft de werkduur terug die gebruikt wordt voor pauzebeslissingen.

    Gebruikt de echte duur (kolom C/D uit Studenten-sheet) als:
      - de theoretische uren zowel 12u als 13u bevatten (= 12-14u periode gedekt)
      - de echte duur verschilt van de theoretische duur

    In alle andere gevallen: gebruik de theoretische duur (student_totalen).
    """
    theo_duur = student_totalen.get(naam, 0)

    student = next((s for s in studenten if s["naam"] == naam), None)
    if student is None:
        return theo_duur

    begin_uur = student.get("begin_uur")
    eind_uur  = student.get("eind_uur")
    if begin_uur is None or eind_uur is None:
        return theo_duur

    echte_duur = eind_uur - begin_uur

    theo_uren = _get_theo_werkuren(naam)
    if 12 not in theo_uren or 13 not in theo_uren:
        return theo_duur

    return echte_duur


def vind_attractie_op_uur(naam, uur):
    """Geef attractienaam (exact zoals in Planning-kolom A) waar student staat op dit uur; None als niet gevonden."""
    for col in range(2, ws_planning.max_column + 1):
        header = ws_planning.cell(1, col).value
        col_uur = parse_header_uur(header)
        if col_uur != uur:
            continue
        for row in range(2, ws_planning.max_row + 1):
            if ws_planning.cell(row, col).value == naam:
                return ws_planning.cell(row, 1).value
    return None
    



##### EXTRA INFO TOEVOEGEN AAN PAUZEPLANNING (A12 e.v.)
##### -------------------------------------------------------------
ws_pauze_sheet = wb_out["Pauzevlinders"]
witte_fill = PatternFill(start_color="FFFFFF", fill_type="solid")


# -------------------------------------



#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# DEEL 5: PP optie 2 + Feedback optie 2
# ──────────────────────────────────────────────────────────────────────
def maak_pp2_sheets(wb_arg, am_arg):
    global ws_planning, student_totalen

    # Globals tijdelijk omwisselen
    _ws_planning_bak     = ws_planning
    _student_totalen_bak = student_totalen

    ws_planning = wb_arg["Planning"]

    # Bouw student_totalen vanuit de Planning-sheet van wb_arg,
    # net zoals de globale versie dat doet vanuit ws_out.
    # Dit is belangrijk voor pauzevlinders: hun pauzevlinderuren
    # staan NIET in assigned_map, maar wél in de Planning-sheet.
    _ws_plan_tmp = wb_arg["Planning"]

    # Bouw col→uur mapping vanuit de headerrij van _ws_plan_tmp
    _col_uur_tmp = {}
    for _c in range(2, _ws_plan_tmp.max_column + 1):
        _uur = parse_header_uur(_ws_plan_tmp.cell(1, _c).value)
        if _uur is not None:
            _col_uur_tmp[_c] = _uur
    
    student_totalen = defaultdict(float)
    for row in _ws_plan_tmp.iter_rows(min_row=2):
        for cel in row[1:]:
            naam = cel.value
            if naam and str(naam).strip() != "":
                _uur = _col_uur_tmp.get(cel.column)
                student_totalen[str(naam).strip()] += blok_durations.get(_uur, 1.0)

    for sheet_name in ["Pauzeplanning", "Feedback PP"]:
        if sheet_name in wb_arg.sheetnames:
            wb_arg.remove(wb_arg[sheet_name])

    ws_pauze_basis = wb_arg["Pauzevlinders"]
    ws_pp2 = wb_arg.copy_worksheet(ws_pauze_basis)
    ws_pp2.title = "Pauzeplanning"
    roze_fill = PatternFill(start_color="FFD6E7", end_color="FFD6E7", fill_type="solid")
    conflict_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ── hierna de rest van DEEL 5 geïndenteerd ──

    # -----------------------------
    # Helpers
    # -----------------------------
    
    def pp2_is_minderjarig(naam):
        return "-18" in str(naam)
    
    
    def pp2_parse_kwartier_header(header):
        """
        Zet '12u', '12u15', '12u30', '12u45' om naar minuten sinds 00:00.
        """
        if not header:
            return None
        s = str(header).strip().lower()
        if "u" not in s:
            return None
        parts = s.split("u", 1)
        try:
            uur = int(parts[0])
            mins = int(parts[1]) if parts[1] != "" else 0
            return uur * 60 + mins
        except:
            return None



    def pp2_tijdvenster_pauze(cols, ws_sheet, speling_minuten=30):
        """
        Geeft (start_min, eind_min) in absolute minuten voor een pauze, incl. speling.
        cols = [col] voor een kort kwartier, of [col1, col2] voor een lang halfuur.
        """
        start_min = pp2_parse_kwartier_header(ws_sheet.cell(1, cols[0]).value)
        eind_min  = pp2_parse_kwartier_header(ws_sheet.cell(1, cols[-1]).value)
        if start_min is None or eind_min is None:
            return None
        eind_min += 15
        return start_min, eind_min + speling_minuten

    def pp2_basis_attractie_naam(attr):
        """
        Strip een volgnummer van een attractie met meerdere plaatsen
        (bv. 'Zipline 1' -> 'Zipline', 'Zipline 1 + Golf' -> 'Zipline + Golf'),
        zodat kwalificatie gecheckt wordt op de attractie zelf, niet de
        specifieke plaats.
        """
        attr = str(attr).strip()
        if " + " in attr:
            delen = [re.sub(r"\s+\d+$", "", d.strip()) for d in attr.split(" + ")]
            return " + ".join(delen)
        return re.sub(r"\s+\d+$", "", attr)

    
    def pp2_attracties_in_venster(naam, start_min, eind_min):
        """
        Attracties die 'naam' bezet tussen start_min en eind_min.
        'Extra'-uren tellen niet mee: die kan elke pauzevlinder overnemen.
        """
        attracties = set()
        for uur in sorted(open_uren):
            blok_start = uur * 60
            blok_eind  = blok_start + blok_durations.get(uur, 1.0) * 60
            if blok_eind <= start_min or blok_start >= eind_min:
                continue
            attr = vind_attractie_op_uur(naam, uur)
            if not attr or str(attr).startswith("Extra") or attr == "Pauzevlinder-vervanging":
                continue
            attracties.add(pp2_basis_attractie_naam(attr))
        return attracties


    def pp2_pv_kan_overname(pv, attracties_set):
        """True als deze PV alle attracties in de set aankan. Lege set => altijd True."""
        if not attracties_set:
            return True
        return all(student_kan_attr_in_analyse(pv, attr) for attr in attracties_set)


    def pp2_eerste_vrije_blok_op_rij(naam, ws_sheet, pv_row, pauze_cols):
        """
        Het allereerste tijdgeldige, vrije halfuur-blok op déze ene PV-rij,
        zonder rekening te houden met kwalificatie — dit is bewust de
        'eigen eerstvolgende plek' van deze PV, niets verder gezocht.
        """
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
        for col1, col2 in blokken:
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
            return (col1, col2)
        return None


    def pp2_schaarste_pv(pv):
        """Lager = 'moeilijkere' pauzevlinder (kan minder attracties)."""
        return len(pv.get("attracties", []))



    def pp2_eerste_vrij_kwartier_op_rij(naam, ws_sheet, pv_row, pauze_cols, open_spots_set):
        """
        Eerste tijdgeldige, vrije KORTE (1 kwartier) plek op deze PV-rij.
        'Plakken': nooit verder gezocht dan deze ene, eerstvolgende plek.
        """
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
            return col
        return None



    def pp2_attracties_in_venster_gesplitst(naam, cols, ws_sheet, speling_minuten=30):
        """
        Zoals pp2_attracties_in_venster, maar gesplitst in:
        - huidige: attracties tijdens de pauze zelf
        - volgende: attracties enkel bereikbaar dankzij de speling erna
        """
        start_min = pp2_parse_kwartier_header(ws_sheet.cell(1, cols[0]).value)
        eind_pauze_min = pp2_parse_kwartier_header(ws_sheet.cell(1, cols[-1]).value) + 15
        if start_min is None or eind_pauze_min is None:
            return set(), set()
        eind_venster_min = eind_pauze_min + speling_minuten

        huidige, volgende = set(), set()
        for uur in sorted(open_uren):
            blok_start = uur * 60
            blok_eind = blok_start + blok_durations.get(uur, 1.0) * 60
            if blok_eind <= start_min or blok_start >= eind_venster_min:
                continue
            attr = vind_attractie_op_uur(naam, uur)
            if not attr or str(attr).startswith("Extra") or attr == "Pauzevlinder-vervanging":
                continue
            attr = pp2_basis_attractie_naam(attr)
            if blok_start < eind_pauze_min:
                huidige.add(attr)
            else:
                volgende.add(attr)
        return huidige, volgende


    def pp2_bouw_conflict_reden(naam, cols, ws_sheet, pv_rows):
        """
        Begrijpbare uitleg waarom deze pauze niet bij een gekwalificeerde
        PV geplaatst kon worden.
        """
        huidige, volgende = pp2_attracties_in_venster_gesplitst(naam, cols, ws_sheet)

        per_pv = []
        for pv, pv_row in pv_rows:
            mist_huidig = sorted(a for a in huidige if not student_kan_attr_in_analyse(pv, a))
            mist_volgend = sorted(a for a in volgende if not student_kan_attr_in_analyse(pv, a))
            per_pv.append((pv["naam"], mist_huidig, mist_volgend))

        # Iedereen faalt op exact dezelfde huidige attractie(s), niets bij 'volgend'?
        alle_huidig = [tuple(h) for _n, h, v in per_pv if h and not v]
        if len(alle_huidig) == len(per_pv) and len(set(alle_huidig)) == 1 and alle_huidig[0]:
            attrs_str = " en ".join(alle_huidig[0])
            return f"omdat geen enkele pauzevlinder de attractie '{attrs_str}' kan overnemen"

        stukken = []
        for pv_naam, mist_huidig, mist_volgend in per_pv:
            if not mist_huidig and not mist_volgend:
                continue
            delen = []
            if mist_huidig:
                delen.append(f"de attractie ('{', '.join(mist_huidig)}') niet kan overnemen")
            if mist_volgend:
                delen.append(f"de attractie die erop volgt ('{', '.join(mist_volgend)}') niet kan overnemen")
            stukken.append(f"{pv_naam} {' en '.join(delen)}")
        if not stukken:
            return "omdat er op dat moment nergens een vrije plek meer was bij een pauzevlinder"

        return "omdat " + "; ".join(stukken)


    def pp2_verzamel_rode_pauze_redenen(ws_sheet, pv_rows, pauze_cols):
        """
        Doorzoekt de volledige Pauzeplanning-sheet op conflict-gekleurde
        (rode) pauzes en bouwt per stuk een leesbare uitleg.
        """
        redenen = []
        conflict_rgb = conflict_fill.start_color.rgb

        for pv, pv_row in pv_rows:
            cols_lijst = list(pauze_cols)
            i = 0
            while i < len(cols_lijst):
                col = cols_lijst[i]
                cel = ws_sheet.cell(pv_row, col)
                rgb = cel.fill.start_color.rgb if cel.fill and cel.fill.start_color else None

                if rgb == conflict_rgb and cel.value:
                    naam = cel.value
                    gebruikte_cols = [col]
                    if i + 1 < len(cols_lijst):
                        volgende_col = cols_lijst[i + 1]
                        volgende_cel = ws_sheet.cell(pv_row, volgende_col)
                        volgende_rgb = (
                            volgende_cel.fill.start_color.rgb
                            if volgende_cel.fill and volgende_cel.fill.start_color else None
                        )
                        if volgende_cel.value == naam and volgende_rgb == conflict_rgb:
                            gebruikte_cols.append(volgende_col)
                            i += 1

                    reden = pp2_bouw_conflict_reden(naam, gebruikte_cols, ws_sheet, pv_rows)
                    redenen.append((naam, f"staat in het rood {reden}."))

                i += 1

        return redenen
        

    def pp2_verzamel_opties_alle_pvs_kort(naam, ws_sheet, pv_rows, pauze_cols, open_spots_set):
        """
        Verzamelt per PV zijn eigen eerstvolgende korte plek, gesorteerd op
        (vroegste tijd, moeilijkste PV) -- zelfde principe als bij lange pauzes.
        """
        opties = []
        for pv, pv_row in pv_rows:
            col = pp2_eerste_vrij_kwartier_op_rij(naam, ws_sheet, pv_row, pauze_cols, open_spots_set)
            if col is not None:
                start_min = pp2_parse_kwartier_header(ws_sheet.cell(1, col).value)
                opties.append((start_min, pp2_schaarste_pv(pv), pv, pv_row, col))
        opties.sort(key=lambda o: (o[0], o[1]))
        return opties


    def pp2_zoek_laatste_kort_kwartier_vanaf(naam, ws_sheet, min_col_exclusive, voorkeur_rij, pv_rows, pauze_cols, open_spots_set):
        """
        Zoek, van rechts naar links, het laatste geldige korte kwartier ná
        min_col_exclusive. Probeert eerst voorkeur_rij, dan de andere rijen
        (vaste volgorde) -- zelfde 'eigen rij eerst'-principe als voorheen.
        Kwalificatie vereist waar mogelijk; lukt dat nergens binnen wat
        toch al geldig was, val terug op de eerste geldige optie, rood.
        Retourneert (pv, pv_row, col, conflict) of None.
        """
        pv_volgorde = (
            [(pv, pv_row) for pv, pv_row in pv_rows if pv_row == voorkeur_rij]
            + [(pv, pv_row) for pv, pv_row in pv_rows if pv_row != voorkeur_rij]
        )
        fallback = None
        for col in reversed(pauze_cols):
            if min_col_exclusive is not None and col <= min_col_exclusive:
                continue
            for pv, pv_row in pv_volgorde:
                if (pv_row, col) in open_spots_set:
                    continue
                if not pp2_is_beschikbaar(ws_sheet, pv_row, col):
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                    continue

                if fallback is None:
                    fallback = (pv, pv_row, col)

                venster = pp2_tijdvenster_pauze([col], ws_sheet)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    return (pv, pv_row, col, False)

        if fallback:
            return (*fallback, True)
        return None

    
    def pp2_verzamel_opties_alle_pvs(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Verzamelt per PV zijn eigen eerstvolgende tijdgeldige, vrije plek voor
        'naam' (nooit verder gezocht dan die ene plek per PV -> 'plakken').
        Sorteert op (vroegste tijd, moeilijkste PV).
        """
        opties = []
        for pv, pv_row in pv_rows:
            plek = pp2_eerste_vrije_blok_op_rij(naam, ws_sheet, pv_row, pauze_cols)
            if plek is None:
                continue
            col1, col2 = plek
            start_min = pp2_parse_kwartier_header(ws_sheet.cell(1, col1).value)
            opties.append((start_min, pp2_schaarste_pv(pv), pv, pv_row, col1, col2))
        opties.sort(key=lambda o: (o[0], o[1]))
        return opties


    def pp2_eerste_halfuur_start_min(naam, ws_sheet, pv_row, pauze_cols):
        """Vindt het starttijdstip (in minuten) van iemands geplaatste blok op deze rij."""
        for col in pauze_cols:
            if ws_sheet.cell(pv_row, col).value == naam:
                return pp2_parse_kwartier_header(ws_sheet.cell(1, col).value)
        return None


    def pp2_kan_2de_halfuur_na(naam, ws_sheet, pv_row, pauze_cols, na_tijdstip_min, extra_bezet_cols=None):
        """
        Simulatie (schrijft niets naar de sheet): kan het 2de halfuur van deze
        minderjarige op deze rij nog een geldig blok vinden dat niet eerder
        start dan na_tijdstip_min? extra_bezet_cols laat toe kolommen die we
        al 'gereserveerd' hebben in dezelfde simulatie ook als bezet te zien.
        """
        extra_bezet_cols = extra_bezet_cols or set()
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
        for col1, col2 in blokken:
            start_min = pp2_parse_kwartier_header(ws_sheet.cell(1, col1).value)
            if start_min is None or start_min < na_tijdstip_min:
                continue
            if col1 in extra_bezet_cols or col2 in extra_bezet_cols:
                continue
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
            return (col1, col2)
        return None
    
    
    def pp2_get_pauze_cols(ws_sheet):
        cols = []
        for col in range(2, ws_sheet.max_column + 1):
            header = ws_sheet.cell(1, col).value
            if header and "u" in str(header):
                cols.append(col)
        return cols
    
    def pp2_get_pv_rows(ws_sheet, selected):
        """
        Geeft lijst van tuples: (pv_dict, naam_rij)
        waarbij naam_rij de rij is waar de naam van de pauzevlinder staat.
        """
        rows = []
        for pv in selected:
            found = None
            for r in range(2, ws_sheet.max_row + 1):
                val = ws_sheet.cell(r, 1).value
                if val and str(val).strip() == str(pv["naam"]).strip():
                    found = r
                    break
            if found is not None:
                rows.append((pv, found))
        return rows
    
    def pp2_get_student_work_hours(naam):
        """
        Leest echte werkuren uit het werkblad Planning.
        """
        uren = set()
        for col in range(2, ws_planning.max_column + 1):
            header = ws_planning.cell(1, col).value
            uur = parse_header_uur(header)
            if uur is None:
                continue
            for row in range(2, ws_planning.max_row + 1):
                if ws_planning.cell(row, col).value == naam:
                    uren.add(uur)
                    break
        return sorted(uren)
    
    def pp2_is_first_or_last_work_hour(naam, kwartier_col, ws_sheet):
        """
        Checkt of dit kwartier in het eerste of laatste werkuur valt.
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return True
    
        header = ws_sheet.cell(1, kwartier_col).value
        pauze_uur = parse_header_uur(header)
        if pauze_uur is None:
            return True
    
        return pauze_uur == werk_uren[0] or pauze_uur == werk_uren[-1]
    
    def pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols):
        """
        Alle geldige kwartierkolommen voor korte pauze:
        - student werkt dat uur
        - niet in eerste of laatste werkuur
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if len(werk_uren) < 4:
            return []
    
        first_hour = werk_uren[0]
        last_hour = werk_uren[-1]
    
        candidates = []
        for col in pauze_cols:
            header = ws_sheet.cell(1, col).value
            uur = parse_header_uur(header)
            if uur is None:
                continue
            if uur in werk_uren and uur != first_hour and uur != last_hour:
                candidates.append(col)
    
        return candidates
    
    def pp2_choose_middle_col(naam, ws_sheet, pauze_cols):
        """
        Geef een gerangschikte lijst van geldige kwartierkolommen terug.
        Voorkeur: kolommen die op een half uur beginnen (:00 of :30), zodat
        een duo netjes één aaneengesloten halfuur-blok vormt. Daarbinnen:
        dichtst bij het midden van de shift eerst. Niet-uitgelijnde
        kolommen komen pas daarna, als terugval.
        """
        candidates = pp2_candidate_cols_for_student(naam, ws_sheet, pauze_cols)
        if not candidates:
            return []

        werk_uren = pp2_get_student_work_hours(naam)
        shift_start = min(werk_uren) * 60
        shift_end   = (max(werk_uren) + 1) * 60
        midpoint    = (shift_start + shift_end) / 2

        scored = []
        for col in candidates:
            mins = pp2_parse_kwartier_header(ws_sheet.cell(1, col).value)
            if mins is not None:
                niet_uitgelijnd = 1 if (mins % 30 != 0) else 0
                scored.append((niet_uitgelijnd, abs(mins - midpoint), col))

        scored.sort(key=lambda x: (x[0], x[1]))
        return [col for _, _, col in scored]
    
    def pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
        """
        Een korte pauze mag alleen als:
        - student werkt in dat kwartier
        - niet in eerste of laatste werkuur
        - student op dat kwartier nog nergens anders in het pauzerooster staat
        """
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
    
        if uur is None:
            return False
    
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return False
    
        if uur not in werk_uren:
            return False
    
        eerste_uur = werk_uren[0]
        laatste_uur = werk_uren[-1]
    
        if uur == eerste_uur or uur == laatste_uur:
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        return True
    
    def pp2_choose_middle_double_col_for_minor(naam, ws_sheet, pauze_cols):
        """
        Zoek startkolom voor 2 opeenvolgende kwartieren voor minderjarigen:
        - student werkt op beide kwartieren
        - student stopt om of voor 16u (dus laatste werkblok <= 15)
        - student werkt >4u en <=6u
        - start enkel op een half uur (:00 of :30)
        - zo vroeg mogelijk in de shift
        - beide cellen moeten geldig zijn volgens de gewone korte-pauze-regels
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return None
    
        if len(werk_uren) <= 4 or len(werk_uren) > 6:
            return None
    
        if max(werk_uren) > 15:
            return None
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            # moeten opeenvolgende kwartieren zijn
            if col2 != col1 + 1:
                continue
    
            header1 = ws_sheet.cell(1, col1).value
            uur1 = parse_header_uur(header1)
            if uur1 is None:
                continue
    
            # start enkel op heel uur of half uur
            header_text = str(header1).strip().lower()
            if not (header_text.endswith("u") or header_text.endswith("u30")):
                continue
    
            # beide kwartieren moeten geldig zijn volgens gewone korte-pauze-regels
            if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                continue
            if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                continue
    
            # beide kwartieren moeten effectief tijdens werkuren vallen
            uur2 = parse_header_uur(ws_sheet.cell(1, col2).value)
            if uur2 is None:
                continue
    
            if uur1 not in werk_uren or uur2 not in werk_uren:
                continue
    
            # eerste geldige optie meteen nemen
            return col1
    
        return None
    
    
    
    def pp2_same_halfhour(col_a, col_b, ws_sheet):
        mins_a = pp2_parse_kwartier_header(ws_sheet.cell(1, col_a).value)
        mins_b = pp2_parse_kwartier_header(ws_sheet.cell(1, col_b).value)
        if mins_a is None or mins_b is None:
            return False
        return (mins_a // 30) == (mins_b // 30)
    
    def pp2_choose_adjacent_same_halfhour(base_col, student_name, ws_sheet, pauze_cols, pv_name_row):
        """
        Tweede student van het duo moet naast de eerste zitten
        in hetzelfde halfuur, indien dat volgens de regels kan.
        """
        if base_col not in pauze_cols:
            return None
    
        idx = pauze_cols.index(base_col)
        opties = []
    
        if idx - 1 >= 0:
            opties.append(pauze_cols[idx - 1])
        if idx + 1 < len(pauze_cols):
            opties.append(pauze_cols[idx + 1])
    
        # Eerst alleen opties in hetzelfde halfuur
        opties = [c for c in opties if pp2_same_halfhour(base_col, c, ws_sheet)]
    
        for col in opties:
            # vak moet leeg zijn
            if ws_sheet.cell(pv_name_row, col).value not in [None, ""]:
                continue
            # niet in eerste/laatste werkuur van deze student
            if pp2_is_first_or_last_work_hour(student_name, col, ws_sheet):
                continue
            # student moet effectief dat uur werken
            uur = parse_header_uur(ws_sheet.cell(1, col).value)
            werk_uren = pp2_get_student_work_hours(student_name)
            if uur not in werk_uren:
                continue
            return col
    
        return None
    
    def pp2_write_name(ws_sheet, row_name, col, naam, conflict=False):
        """
        Schrijf in PP optie 2:
        - bovenste vak: attractie waarop student dat moment staat
        - onderste vak: naam van student
        - korte pauze = paars
        - lange pauze = groen (voor later bruikbaar)
        """
        lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    
        # bepaal uur van deze kolom
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)
    
        # attractie erboven invullen
        info_cel = ws_sheet.cell(row_name - 1, col)
        attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
        info_cel.value = attr if attr else ""
        info_cel.alignment = center_align
        info_cel.border = thin_border
    
        # naam invullen
        cel = ws_sheet.cell(row_name, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
    
        # check of dit een lange of korte pauze is
        is_lange_pauze = False
        if col - 1 >= 2 and ws_sheet.cell(row_name, col - 1).value == naam:
            is_lange_pauze = True
        if col + 1 <= ws_sheet.max_column and ws_sheet.cell(row_name, col + 1).value == naam:
            is_lange_pauze = True
    
        cel.fill = lichtgroen_fill if is_lange_pauze else lichtpaars_fill
        cel.fill = conflict_fill if conflict else (lichtgroen_fill if is_lange_pauze else lichtpaars_fill)
    
    def pp2_clear_pauze_grid(ws_sheet, pv_rows, pauze_cols):
        """
        Wis enkel de effectieve pauzevakken:
        - rij erboven: attractie/info
        - naamrij: naam
        Kolom A en extra info lager op het blad blijven behouden.
        """
        leeg_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
        for pv, naam_rij in pv_rows:
            info_rij = naam_rij - 1
            for col in pauze_cols:
                # bovenste rij leegmaken
                ws_sheet.cell(info_rij, col).value = None
                ws_sheet.cell(info_rij, col).alignment = center_align
                ws_sheet.cell(info_rij, col).border = thin_border
    
                # naamrij leegmaken
                ws_sheet.cell(naam_rij, col).value = None
                ws_sheet.cell(naam_rij, col).alignment = center_align
                ws_sheet.cell(naam_rij, col).border = thin_border
                ws_sheet.cell(naam_rij, col).fill = leeg_fill
    
    
    def pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_sheet, pv_rows):
        """
        True als deze student op deze kwartierkolom al ergens in het pauzerooster staat,
        ongeacht op welke pauzevlinder-rij.
        """
        for _pv, pv_row in pv_rows:
            if ws_sheet.cell(pv_row, col).value == naam:
                return True
        return False
    
    
    def pp2_student_heeft_al_lange_pauze_op_blok(naam, col1, col2, ws_sheet, pv_rows):
        """
        True als deze student deze 2 kwartieren al ergens als lange pauze heeft staan.
        """
        for _pv, pv_row in pv_rows:
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
        return False
    
    
    # -----------------------------
    # Vind de pauzevlinder-rijen in PP optie 2
    # -----------------------------
    pauze_cols_pp2 = pp2_get_pauze_cols(ws_pp2)
    pv_rows_pp2 = pp2_get_pv_rows(ws_pp2, selected)

    # Welke PV is de afgeknipte (indien van toepassing)? 
    # open spots i.p.v. het normale aandeel, want die dient toch geen
    # korte pauzes meer na de lange pauzes.
    pp2_afgeknipte_pv_naam = None
    if afgekapte_pv_uren:
        _afknip_pv_obj = pp2_bepaal_pv_voor_afknip(selected)
        if _afknip_pv_obj is not None:
            pp2_afgeknipte_pv_naam = _afknip_pv_obj["naam"]

    _aantal_afgeknipte_uren_pv = len(afgekapte_pv_uren)
    if _aantal_afgeknipte_uren_pv >= 4:
        pp2_max_open_spots_afgeknipt = 0
    elif _aantal_afgeknipte_uren_pv >= 2:
        pp2_max_open_spots_afgeknipt = 2
    else:
        pp2_max_open_spots_afgeknipt = min_open_spots_per_pv  # geen extra begrenzing
    pp2_open_spots_afgeknipte_teller = 0
    
    # Maak de grid leeg, maar behoud layout
    pp2_clear_pauze_grid(ws_pp2, pv_rows_pp2, pauze_cols_pp2)
    
    # -----------------------------
    # Closed spots PP optie 2
    # Afgekapte uren van laatste PV: worden hier gemarkeerd
    # zodat géén enkele pauzelogica er iets in plaatst
    # -----------------------------
    afgeknipt_fill_pp2 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    pp2_closed_spots = set()  # set van (naam_rij, col)
    
    if afgekapte_pv_uren and selected:
        laatste_pv = pp2_bepaal_pv_voor_afknip(selected)
        for pv, naam_rij in pv_rows_pp2:
            if pv["naam"] == laatste_pv["naam"]:
                for col in pauze_cols_pp2:
                    col_uur = parse_header_uur(ws_pp2.cell(1, col).value)
                    if col_uur in afgekapte_pv_uren:
                        pp2_closed_spots.add((naam_rij, col))
                        ws_pp2.cell(naam_rij, col).value = "X"
                        ws_pp2.cell(naam_rij, col).fill = afgeknipt_fill_pp2
                        ws_pp2.cell(naam_rij, col).alignment = center_align
                        ws_pp2.cell(naam_rij, col).border = thin_border
                        ws_pp2.cell(naam_rij - 1, col).value = None
                        ws_pp2.cell(naam_rij - 1, col).fill = afgeknipt_fill_pp2
                        ws_pp2.cell(naam_rij - 1, col).alignment = center_align
                        ws_pp2.cell(naam_rij - 1, col).border = thin_border
                break
    
    def pp2_is_beschikbaar(ws_sheet, rij, col):
        """Cel is beschikbaar als ze leeg is EN geen closed spot."""
        if (rij, col) in pp2_closed_spots:
            return False
        return ws_sheet.cell(rij, col).value in [None, ""]
    
    
    
    # -----------------------------
    # STAP 1:
    # Vroege stoppers (minstens 4u gewerkt en laatste werkblok <= 15)
    # - minderjarige vroege stoppers: eerst halfuur zo vroeg mogelijk,
    #   dan kwartier zo laat mogelijk, zo mogelijk op dezelfde PV-rij
    # - gewone vroege stoppers: duo-logica zoals voorheen
    # Excl. pauzevlinders zelf
    # -----------------------------
    pauzevlinder_namen_set = {pv["naam"] for pv in selected}
    
    vroege_stoppers_gewoon = []
    vroege_stoppers_minderjarig = []
    
    for s in studenten:
        naam = s["naam"]
    
        if naam in pauzevlinder_namen_set:
            continue
    
        werk_uren = pp2_get_student_work_hours(naam)
        if len(werk_uren) < 4:
            continue
    
        laatste_werkblok = max(werk_uren)
        startuur = min(werk_uren)
        aantal_uren = len(werk_uren)
    
        einduur_dag = max(open_uren) if open_uren else None
        if einduur_dag is None or einduur_dag > 16:
            if laatste_werkblok > 15:
                continue
        if werkduur_voor_pauze(naam) > 6:  # lange werkers horen in stap 2
            continue
    
        item = {
            "naam": naam,
            "werk_uren": werk_uren,
            "einduur": laatste_werkblok,
            "startuur": startuur,
            "aantal_uren": aantal_uren
        }
    
        if pp2_is_minderjarig(naam):
            vroege_stoppers_minderjarig.append(item)
        else:
            vroege_stoppers_gewoon.append(item)
    
    # Sorteervolgorde: vroegst stoppend, vroegst beginnend, alfabetisch
    vroege_stoppers_minderjarig.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))
    vroege_stoppers_gewoon.sort(key=lambda x: (x["einduur"], x["startuur"], x["naam"]))

    # Op korte dagen (einduur <= 17u) zijn er geen gewone vroege stoppers:
    # iedereen valt door naar stap 3 voor een beter gespreide verdeling.
    # Minderjarigen (stap 1a) blijven wel hun halfuur krijgen via de eigen logica.
    if open_uren and max(open_uren) <= 16:
        vroege_stoppers_gewoon = []
    
    pp2_geplaatste_pauzes = []
    pp2_niet_geplaatst = []
        
    # -----------------------------
    # STAP 1a: minderjarige vroege stoppers
    # Pauze 1: halfuur (2 opeenvolgende kwartieren) zo vroeg mogelijk
    # Pauze 2: kwartier zo laat mogelijk
    # Beide pauzes: niet in eerste of laatste werkuur
    # Pauze 2 zo mogelijk op dezelfde PV-rij als pauze 1
    # -----------------------------
    pp2_minderjarige_vroege_stopper_rij = {}
    
    if pv_rows_pp2:
        for idx, item in enumerate(vroege_stoppers_minderjarig):
            naam = item["naam"]
            werk_uren = item["werk_uren"]
            eerste_uur = werk_uren[0]
            laatste_uur = werk_uren[-1]
    
            pv_index = idx % len(pv_rows_pp2)
            pv, pv_name_row = pv_rows_pp2[pv_index]
            pv_label = pv["naam"]
    
            # -- Pauze 1: halfuur zo vroeg mogelijk --
            col1_gekozen = None
            for i in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[i]
                col2 = pauze_cols_pp2[i + 1]
    
                # opeenvolgende kwartieren
                if col2 != col1 + 1:
                    continue
    
                uur1 = parse_header_uur(ws_pp2.cell(1, col1).value)
                uur2 = parse_header_uur(ws_pp2.cell(1, col2).value)
    
                if uur1 is None or uur2 is None:
                    continue
    
                # niet in eerste of laatste werkuur
                if uur1 == eerste_uur or uur1 == laatste_uur:
                    continue
                if uur2 == eerste_uur or uur2 == laatste_uur:
                    continue
    
                # student moet beide uren werken
                if uur1 not in werk_uren or uur2 not in werk_uren:
                    continue
    
                # cellen moeten leeg zijn op deze PV-rij
                if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col1):
                    continue
                if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col2):
                    continue
    
                # eerste geldige optie nemen
                col1_gekozen = col1
                break
    
            if col1_gekozen is not None:
                col2_gekozen = col1_gekozen + 1
                pp2_write_name(ws_pp2, pv_name_row, col1_gekozen, naam)
                pp2_write_name(ws_pp2, pv_name_row, col2_gekozen, naam)
                pp2_minderjarige_vroege_stopper_rij[naam] = pv_name_row
    
                pp2_geplaatste_pauzes.append({
                    "naam": naam,
                    "pauzevlinder": pv_label,
                    "tijd": f"{ws_pp2.cell(1, col1_gekozen).value}-{ws_pp2.cell(1, col2_gekozen).value}",
                    "type": "minderjarig vroege stopper - halfuur"
                })
            else:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldig halfuur gevonden voor minderjarige vroege stopper (pauze 1)"
                })
    
            # -- Pauze 2: kwartier zo laat mogelijk, bij voorkeur zelfde PV-rij --
            vaste_rij = pp2_minderjarige_vroege_stopper_rij.get(naam)
    
            # Kandidaat-kolommen van achter naar voor
            kandidaten = list(reversed(pauze_cols_pp2))
    
            kort_geplaatst = False
            for gebruik_rij in ([vaste_rij] if vaste_rij else []) + [r for (_pv2, r) in pv_rows_pp2 if r != vaste_rij]:
                for col in kandidaten:
                    uur = parse_header_uur(ws_pp2.cell(1, col).value)
                    if uur is None:
                        continue
    
                    # niet in eerste of laatste werkuur
                    if uur == eerste_uur or uur == laatste_uur:
                        continue
    
                    # student moet dat uur werken
                    if uur not in werk_uren:
                        continue
    
                    # cel moet leeg zijn
                    if not pp2_is_beschikbaar(ws_pp2, gebruik_rij, col):
                        continue
    
                    # student mag op dit kwartier nog nergens staan
                    if pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_pp2, pv_rows_pp2):
                        continue
    
                    pp2_write_name(ws_pp2, gebruik_rij, col, naam)
    
                    pp2_geplaatste_pauzes.append({
                        "naam": naam,
                        "pauzevlinder": ws_pp2.cell(gebruik_rij, 1).value or f"rij {gebruik_rij}",
                        "tijd": ws_pp2.cell(1, col).value,
                        "type": "minderjarig vroege stopper - kort kwartier"
                    })
    
                    kort_geplaatst = True
                    break
    
                if kort_geplaatst:
                    break
    
            if not kort_geplaatst:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "geen geldig kwartier gevonden voor minderjarige vroege stopper (pauze 2)"
                })
    
    # -----------------------------
    # STAP 1b: gewone vroege stoppers
    # Inplannen per duo:
    # 1-2 bij PV1, 3-4 bij PV2, 5-6 bij PV3, ...
    # als er meer duo's zijn dan pauzevlinders, dan cyclisch verder
    # Als de voorkeurs-PV-rij al bezet is op de gekozen kolom,
    # worden andere PV-rijen geprobeerd.
    # -----------------------------
    duo_basis_col = {}
    duo_basis_pv_row = {}
    cluster_laatste_col = None
    cluster_laatste_pv_row = None
    
    if pv_rows_pp2:
        for idx, item in enumerate(vroege_stoppers_gewoon):
            naam = item["naam"]
    
            duo_nummer = idx // 2
            pv_index_voorkeur = duo_nummer % len(pv_rows_pp2)
    
            # Eerste van het duo
            if idx % 2 == 0:
                is_solo = (
                    idx == len(vroege_stoppers_gewoon) - 1
                    and len(vroege_stoppers_gewoon) % 2 == 1
                )

                aansluit_col = None
                if is_solo and cluster_laatste_col is not None:
                    kandidaat = cluster_laatste_col + 1
                    if (
                        kandidaat in pauze_cols_pp2
                        and pp2_is_beschikbaar(ws_pp2, cluster_laatste_pv_row, kandidaat)
                        and pp2_is_valid_short_break_for_student(naam, kandidaat, ws_pp2)
                    ):
                        aansluit_col = kandidaat

                if aansluit_col is not None:
                    gekozen_col_kandidaten = [aansluit_col]
                    pv_volgorde = (
                        [(pv, r) for pv, r in pv_rows_pp2 if r == cluster_laatste_pv_row]
                        + [(pv, r) for pv, r in pv_rows_pp2 if r != cluster_laatste_pv_row]
                    )
                else:
                    gekozen_col_kandidaten = pp2_choose_middle_col(naam, ws_pp2, pauze_cols_pp2)
                    pv_volgorde = (
                        [pv_rows_pp2[pv_index_voorkeur]]
                        + [r for i, r in enumerate(pv_rows_pp2) if i != pv_index_voorkeur]
                    )

                if not gekozen_col_kandidaten:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen geldige middenplek gevonden voor eerste van duo"
                    })
                    continue

                # Probeer elke kandidaatkolom (dichtst bij midden eerst,
                # of -- voor de partnerloze laatste -- aansluitend op de
                # bestaande cluster), en per kolom: eerst voorkeurs-PV-rij,
                # daarna de rest.
                geplaatst_eerste = False
                for gekozen_col in gekozen_col_kandidaten:
                    for pv, pv_name_row in pv_volgorde:
                        if not pp2_is_beschikbaar(ws_pp2, pv_name_row, gekozen_col):
                            continue

                        pp2_write_name(ws_pp2, pv_name_row, gekozen_col, naam)
                        duo_basis_col[duo_nummer]    = gekozen_col
                        duo_basis_pv_row[duo_nummer] = pv_name_row

                        if cluster_laatste_col is None or gekozen_col > cluster_laatste_col:
                            cluster_laatste_col = gekozen_col
                            cluster_laatste_pv_row = pv_name_row

                        pp2_geplaatste_pauzes.append({
                            "naam": naam,
                            "pauzevlinder": pv["naam"],
                            "tijd": ws_pp2.cell(1, gekozen_col).value,
                            "type": "eerste van duo"
                        })
                        geplaatst_eerste = True
                        break

                    if geplaatst_eerste:
                        break

                if not geplaatst_eerste:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen vrije plek gevonden voor eerste van duo (alle rijen/kolommen bezet)"
                    })
    
            # Tweede van het duo
            else:
                basis_col = duo_basis_col.get(duo_nummer)
                pv_name_row = duo_basis_pv_row.get(duo_nummer)
    
                if basis_col is None or pv_name_row is None:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen basisplek beschikbaar van eerste duo-genoot"
                    })
                    continue
    
                pv_label = next(
                    (pv["naam"] for pv, r in pv_rows_pp2 if r == pv_name_row),
                    f"rij {pv_name_row}"
                )
    
                buur_cols = []
                if basis_col + 1 in pauze_cols_pp2:
                    buur_cols.append(basis_col + 1)
                if basis_col - 1 in pauze_cols_pp2:
                    buur_cols.append(basis_col - 1)
    
                geplaatste_tweede = False
    
                for buur_col in buur_cols:
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, buur_col):
                        continue
    
                    if not pp2_is_valid_short_break_for_student(naam, buur_col, ws_pp2):
                        continue
    
                    pp2_write_name(ws_pp2, pv_name_row, buur_col, naam)

                    if cluster_laatste_col is None or buur_col > cluster_laatste_col:
                        cluster_laatste_col = buur_col
                        cluster_laatste_pv_row = pv_name_row
    
                    pp2_geplaatste_pauzes.append({
                        "naam": naam,
                        "pauzevlinder": pv_label,
                        "tijd": ws_pp2.cell(1, buur_col).value,
                        "type": "tweede van duo"
                    })
    
                    geplaatste_tweede = True
                    break
    
                if not geplaatste_tweede:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "geen geldige buurplek gevonden voor tweede van duo"
                    })
    
    
    #STAP 2 2222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222222
    
    # -----------------------------
    # STAP 2 PP optie 2:
    # lange pauzes invullen van links naar rechts,
    # per halfuurblok en per pauzevlinder
    # -----------------------------
    
    lichtgroen_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    def pp2_heeft_al_lange_pauze(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Check of naam al ergens een dubbele blok heeft in PP optie 2.
        """
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
                if (
                    ws_sheet.cell(pv_row, col1).value == naam and
                    ws_sheet.cell(pv_row, col2).value == naam
                ):
                    return True
        return False
    
    
    def pp2_lange_werkers_lijst():
        """
        Studenten die in stap 2 recht hebben op een halfuur pauze:
        - alle minderjarigen met minstens 4 uur werk
        - alle overige studenten met meer dan 6 uur werk
        - inclusief pauzevlinders indien ze eraan voldoen
        """
        result = []
        al_toegevoegd = set()
    
        for s in studenten:
            naam = s["naam"]
            gewerkte_uren = werkduur_voor_pauze(naam)
            is_minderjarig = "-18" in str(naam)
    
            if is_minderjarig and ((gewerkte_uren > 4) if PAUZE_STRIKT_BOVEN_4U else (gewerkte_uren >= 4)):
                if naam not in al_toegevoegd:
                    result.append(naam)
                    al_toegevoegd.add(naam)
            elif gewerkte_uren > 6:
                if naam not in al_toegevoegd:
                    result.append(naam)
                    al_toegevoegd.add(naam)
    
        return result
    
    
    
    def pp2_aantal_lange_pauzes_nodig_in_stap2(naam):
        """
        Hoeveel halfuren moet deze student in stap 2 krijgen?
        - minderjarige met < 4u werk => 0
        - minderjarige met >= 4u en <= 6u werk => 1
        - minderjarige met > 6u werk => 2
        - niet-minderjarige met > 6u werk => 1
        - anders => 0
        """
        gewerkte_uren = werkduur_voor_pauze(naam)
        is_minderjarig = "-18" in str(naam)
    
        if is_minderjarig:
            if PAUZE_STRIKT_BOVEN_4U:
                if gewerkte_uren <= 4:
                    return 0
            else:
                if gewerkte_uren < 4:
                    return 0
            if gewerkte_uren > 6:
                return 2
            return 1
    
        if gewerkte_uren > 6:
            return 1
    
        return 0
    
    
    def pp2_sort_step2_namen(namenlijst):
        """
        Sorteer voor stap 2:
        - eerst wie vroeger stopt
        - bij gelijke eindtijd:
            * wie tot het einduur van de dag werkt: wie eerder START,
              krijgt ook eerder pauze (geen willekeur meer)
            * anders: random volgorde (maakt niet uit wie eerst is)
        """
        dag_eind_uur = max(open_uren) if open_uren else None

        per_einduur = defaultdict(list)
        for naam in namenlijst:
            werk_uren = pp2_get_student_work_hours(naam)
            if werk_uren:
                einduur = max(werk_uren)
                per_einduur[einduur].append(naam)

        resultaat = []
        for einduur in sorted(per_einduur.keys()):
            groep = per_einduur[einduur][:]
            if dag_eind_uur is not None and einduur == dag_eind_uur:
                random.shuffle(groep)
                groep.sort(key=lambda n: min(pp2_get_student_work_hours(n)))
            else:
                random.shuffle(groep)
            resultaat.extend(groep)
        return resultaat
    
    def pp2_get_pv_row_for_name(naam, pv_rows):
        """
        Geef de naamrij terug van de pauzevlinder met deze naam.
        """
        for pv, pv_row in pv_rows:
            if pv["naam"] == naam:
                return pv_row
        return None


    def pp2_vind_rij_met_lange_pauze(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Zoekt op welke PV-rij deze student (niet per se zelf een pauzevlinder)
        al een lange pauze heeft staan. Geeft de rij terug, of None.
        """
        for _pv, pv_row in pv_rows:
            if pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
                return pv_row
        return None


    def pp2_lange_pauze_eindkolom_op_rij(naam, ws_sheet, pv_row, pauze_cols):
        """
        Eindkolom van deze student zijn lange pauze op deze rij (of None).
        Gebruikt als ankerpunt zodat de korte pauze nooit vóór de lange
        pauze van dezelfde persoon kan vallen.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
            if (
                ws_sheet.cell(pv_row, col1).value == naam
                and ws_sheet.cell(pv_row, col2).value == naam
            ):
                return col2
        return None
        
    
    def pp2_find_first_valid_long_block_any_row(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Zoekt het vroegst mogelijke geldige halfuur, met voorkeur voor een PV die
        de attracties in het venster (+30 min speling) effectief aankan.
        Geen enkele optie gekwalificeerd? Valt terug op de eerst tijdgeldige optie.
        Retourneert (pv_row, col1, col2, conflict) of None.
        """
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
        fallback = None
    
        for col1, col2 in blokken:
            for pv, pv_row in pv_rows:
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
                if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                    continue
    
                if fallback is None:
                    fallback = (pv_row, col1, col2)
    
                venster = pp2_tijdvenster_pauze([col1, col2], ws_sheet)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    return (pv_row, col1, col2, False)
    
        if fallback:
            return (*fallback, True)
        return None
    
    
    def pp2_find_first_valid_long_block_on_fixed_row(naam, ws_sheet, pv_row, pauze_cols, pv=None):
        """
        Zoek het vroegst mogelijke geldige halfuur op één vaste PV-rij.
        Retourneert (col1, col2, conflict) of None.
        """
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
        fallback = None
    
        for col1, col2 in blokken:
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
    
            if fallback is None:
                fallback = (col1, col2)
    
            if pv is not None:
                venster = pp2_tijdvenster_pauze([col1, col2], ws_sheet)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    return (col1, col2, False)
            else:
                return (col1, col2, False)
    
        if fallback:
            return (*fallback, True)
        return None
            
    
    
    def pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
        """
        Een lange pauze mag alleen als:
        - beide kwartieren samen exact 30 min vormen
        - student werkt in beide kwartieren
        - niet in eerste of laatste werkuur
        - student op geen van beide kwartieren al elders in het pauzerooster staat
        """
        header1 = ws_sheet.cell(1, col1).value
        header2 = ws_sheet.cell(1, col2).value
    
        mins1 = pp2_parse_kwartier_header(header1)
        mins2 = pp2_parse_kwartier_header(header2)
    
        if mins1 is None or mins2 is None:
            return False
    
        if mins2 - mins1 != 15:
            return False
    
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            return False
    
        uur1 = parse_header_uur(header1)
        uur2 = parse_header_uur(header2)
    
        if uur1 is None or uur2 is None:
            return False
    
        if uur1 not in werk_uren or uur2 not in werk_uren:
            return False
    
        eerste_uur = werk_uren[0]
        laatste_uur = werk_uren[-1]
    
        if uur1 == eerste_uur or uur1 == laatste_uur:
            return False
        if uur2 == eerste_uur or uur2 == laatste_uur:
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col1,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        if pp2_student_heeft_al_pauze_op_kolom(
            naam=naam,
            col=col2,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2
        ):
            return False
    
        return True
        
    
    def pp2_write_long_break(ws_sheet, pv_row, col1, col2, naam, leave_top_blank=False, conflict=False):
        """
        Schrijf een lange pauze van 2 kwartieren:
        - normaal: attractie erboven
        - voor pauzevlinder op eigen rij: bovenste cel leeg laten
        - naam in beide vakjes
        - groen kleuren
        """
        for col in [col1, col2]:
            info_cel = ws_sheet.cell(pv_row - 1, col)
            info_cel.alignment = center_align
            info_cel.border = thin_border
    
            if leave_top_blank:
                info_cel.value = ""
            else:
                header = ws_sheet.cell(1, col).value
                uur = parse_header_uur(header)
                attr = vind_attractie_op_uur(naam, uur) if uur is not None else None
                info_cel.value = attr if attr else ""
    
            info_cel.fill = conflict_fill if conflict else PatternFill(fill_type=None)
            naam_cel = ws_sheet.cell(pv_row, col)
            naam_cel.value = naam
            naam_cel.alignment = center_align
            naam_cel.border = thin_border
            naam_cel.fill = conflict_fill if conflict else lichtgroen_fill
            
    
    
    def pp2_halfuur_blokken(pauze_cols, ws_sheet):
        """
        Geeft alle mogelijke halfuurblokken terug, van links naar rechts.
        Flexibel:
        - mag starten op heel uur
        - mag ook starten op :15
        Dus bv.:
        (12u00, 12u15), (12u15, 12u30), (12u30, 12u45), ...
        zolang de cellen exact 15 minuten uit elkaar liggen.
        """
        blokken = []
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            mins1 = pp2_parse_kwartier_header(ws_sheet.cell(1, col1).value)
            mins2 = pp2_parse_kwartier_header(ws_sheet.cell(1, col2).value)
    
            if mins1 is None or mins2 is None:
                continue
    
            if mins2 - mins1 == 15:
                blokken.append((col1, col2))
    
        return blokken
    
    
    def pp2_place_long_break_for_pv_in_own_row(pv, pv_name_row, ws_sheet, pauze_cols, lange_pauze_ontvangers, lange_werkers_random):
        """
        Geef een langwerkende pauzevlinder verplicht een lange pauze in de eigen rij.
        We proberen de blokken strikt van links naar rechts.
        De cellen erboven blijven leeg.
        """
        naam = pv["naam"]
    
        if naam not in lange_werkers_random:
            return False
    
        if naam in lange_pauze_ontvangers:
            return False
    
        blokken = pp2_halfuur_blokken(pauze_cols, ws_sheet)
    
        for col1, col2 in blokken:
            # beide kwartieren moeten leeg zijn op eigen rij
            if ws_sheet.cell(pv_name_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_name_row, col2).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                continue
    
            pp2_write_long_break(
                ws_sheet=ws_sheet,
                pv_row=pv_name_row,
                col1=col1,
                col2=col2,
                naam=naam,
                leave_top_blank=True
            )
            lange_pauze_ontvangers.add(naam)
            return True
    
        return False
    
    
    # -----------------------------
    # Helpers voor speciale meerderjarige lange werkers
    # (theo <= 6u, echt > 6u, theo uren bevatten 12u én 13u)
    # -----------------------------

    def pp2_is_speciale_lange_werker(naam):
        """
        Meerderjarige student die theoretisch <= 6u werkt maar echt > 6u,
        en waarvan de theoretische uren zowel 12u als 13u bevatten.
        """
        if pp2_is_minderjarig(naam):
            return False
        if student_totalen.get(naam, 0) > 6:
            return False
        if werkduur_voor_pauze(naam) <= 6:
            return False
        theo_uren = pp2_get_student_work_hours(naam)
        return 12 in theo_uren and 13 in theo_uren

    def pp2_speciale_groep(naam):
        """
        1 = echte einduur - theo einduur >= 1.5u → enkel lange pauze 12-14u, geen korte
        2 = echte einduur - theo einduur < 1.5u  → lange + korte (minderjarige regeling)
        None = niet van toepassing
        """
        if not pp2_is_speciale_lange_werker(naam):
            return None
        student = next((s for s in studenten if s["naam"] == naam), None)
        if not student:
            return None
        echte_eind = student.get("eind_uur")
        if echte_eind is None:
            return None
        theo_uren = pp2_get_student_work_hours(naam)
        if not theo_uren:
            return None
        theo_eind_slot = max(theo_uren)
        theo_eind = theo_eind_slot + blok_durations.get(theo_eind_slot, 1.0)
        return 1 if (echte_eind - theo_eind) >= 1.5 else 2

    def pp2_is_valid_long_break_12_14(naam, col1, col2, ws_sheet):
        """Lange pauze geldig én uur van col1 moet 12 of 13 zijn."""
        if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
            return False
        uur1 = parse_header_uur(ws_sheet.cell(1, col1).value)
        return uur1 in (12, 13)

    # 1) Bouw de kandidatenlijsten voor stap 2
    pp2_step2_basis = pp2_lange_werkers_lijst()
    
    pp2_step2_minderjarigen = []
    pp2_step2_overige_lange_werkers = []
    
    for naam in pp2_step2_basis:
        if "-18" in str(naam):
            pp2_step2_minderjarigen.append(naam)
        else:
            pp2_step2_overige_lange_werkers.append(naam)
    
    pp2_step2_minderjarigen = pp2_sort_step2_namen(pp2_step2_minderjarigen)
    pp2_step2_overige_lange_werkers = pp2_sort_step2_namen(pp2_step2_overige_lange_werkers)
    
    # Verwijder speciale lange werkers uit de gewone lijst (eigen verwerking hieronder)
    pp2_speciale_groep1_namen = set()
    pp2_step2_overige_lange_werkers = [
        naam for naam in pp2_step2_overige_lange_werkers
        if pp2_speciale_groep(naam) is None
    ]
    # Groep 1 namen bijhouden voor korte pauze uitsluiting later
    for naam in pp2_step2_basis:
        if pp2_speciale_groep(naam) == 1:
            pp2_speciale_groep1_namen.add(naam)

    # Deze lijst houden we voor de bestaande latere logica aan
    pp2_lange_werkers_random = pp2_step2_minderjarigen + pp2_step2_overige_lange_werkers

    # 2) Houd bij wie al minstens één lange pauze kreeg
    pp2_lange_pauze_ontvangers = set()
    for naam in pp2_lange_werkers_random:
        if pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
            pp2_lange_pauze_ontvangers.add(naam)
    
    # Voor minderjarigen willen we onthouden op welke rij hun EERSTE halfuur kwam
    pp2_minderjarige_eerste_halfuur_rij = {}

    # 2b) Speciale meerderjarige lange werkers (groep 1 en 2)
    pp2_speciale_teller = 0
    for s in studenten:
        naam = s["naam"]
        groep = pp2_speciale_groep(naam)
        if groep is None:
            continue

        _is_pv = naam in pauzevlinder_namen_set
        _eigen_rij_speciaal = None
        if _is_pv:
            _eigen_rij_speciaal = next((r for p, r in pv_rows_pp2 if p["naam"] == naam), None)

        theo_uren = pp2_get_student_work_hours(naam)
        eerste_uur = theo_uren[0] if theo_uren else None
        laatste_uur = theo_uren[-1] if theo_uren else None

        if groep == 1:
            # Alleen lange pauze, verplicht tussen 12u en 14u.
            # Pauzevlinders proberen EERST hun eigen rij, anders elke rij.
            _pv_volgorde_g1 = (
                [(p, r) for p, r in pv_rows_pp2 if r == _eigen_rij_speciaal]
                + [(p, r) for p, r in pv_rows_pp2 if r != _eigen_rij_speciaal]
                if _eigen_rij_speciaal is not None else pv_rows_pp2
            )
            geplaatst = False
            for col1, col2 in pp2_halfuur_blokken(pauze_cols_pp2, ws_pp2):
                if geplaatst:
                    break
                for _pv, pv_name_row in _pv_volgorde_g1:
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col1):
                        continue
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col2):
                        continue
                    if not pp2_is_valid_long_break_12_14(naam, col1, col2, ws_pp2):
                        continue
                    pp2_write_long_break(
                        ws_sheet=ws_pp2, pv_row=pv_name_row,
                        col1=col1, col2=col2,
                        naam=naam, leave_top_blank=(pv_name_row == _eigen_rij_speciaal)
                    )
                    pp2_lange_pauze_ontvangers.add(naam)
                    geplaatst = True
                    break
            if not geplaatst:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "groep 1 speciale lange werker: geen geldig halfuur 12-14u gevonden"
                })

        elif groep == 2:
            # Lange pauze tussen 12u-14u + korte pauze zo ver mogelijk van de lange pauze.
            # Pauzevlinders proberen EERST hun eigen rij, anders round-robin zoals voorheen.
            if _eigen_rij_speciaal is not None:
                pv_name_row = _eigen_rij_speciaal
            else:
                pv_index = pp2_speciale_teller % len(pv_rows_pp2) if pv_rows_pp2 else 0
                _pv_g2, pv_name_row = pv_rows_pp2[pv_index] if pv_rows_pp2 else (None, None)

            # Pauze 1: lange pauze tussen 12u en 14u (zelfde logica als groep 1)
            col1_gekozen = None
            if pv_name_row:
                for col1, col2 in pp2_halfuur_blokken(pauze_cols_pp2, ws_pp2):
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col1):
                        continue
                    if not pp2_is_beschikbaar(ws_pp2, pv_name_row, col2):
                        continue
                    if not pp2_is_valid_long_break_12_14(naam, col1, col2, ws_pp2):
                        continue
                    col1_gekozen = col1
                    break

            # Eigen rij lukt niet (pauzevlinder) -> val terug op een andere rij.
            if col1_gekozen is None and _eigen_rij_speciaal is not None:
                for _p2, _r2 in pv_rows_pp2:
                    if _r2 == _eigen_rij_speciaal:
                        continue
                    for col1, col2 in pp2_halfuur_blokken(pauze_cols_pp2, ws_pp2):
                        if not pp2_is_beschikbaar(ws_pp2, _r2, col1):
                            continue
                        if not pp2_is_beschikbaar(ws_pp2, _r2, col2):
                            continue
                        if not pp2_is_valid_long_break_12_14(naam, col1, col2, ws_pp2):
                            continue
                        pv_name_row = _r2
                        col1_gekozen = col1
                        break
                    if col1_gekozen is not None:
                        break

            if col1_gekozen is not None:
                pp2_write_long_break(
                    ws_sheet=ws_pp2, pv_row=pv_name_row,
                    col1=col1_gekozen, col2=col1_gekozen + 1,
                    naam=naam, leave_top_blank=(pv_name_row == _eigen_rij_speciaal)
                )
                pp2_lange_pauze_ontvangers.add(naam)

                # Pauze 2: korte pauze zo ver mogelijk van de lange pauze
                # Sorteer kandidaten op afdalende afstand t.o.v. het midden van de lange pauze
                lange_pauze_midden = col1_gekozen + 0.5

                kandidaat_cols = [
                    col for col in pauze_cols_pp2
                    if parse_header_uur(ws_pp2.cell(1, col).value) not in (None, eerste_uur, laatste_uur)
                    and parse_header_uur(ws_pp2.cell(1, col).value) in theo_uren
                    and not pp2_student_heeft_al_pauze_op_kolom(naam, col, ws_pp2, pv_rows_pp2)
                ]
                kandidaat_cols.sort(key=lambda c: -abs(c - lange_pauze_midden))

                kort_geplaatst = False
                for gebruik_rij in ([pv_name_row] + [r for (_p2, r) in pv_rows_pp2 if r != pv_name_row]):
                    for col in kandidaat_cols:
                        if not pp2_is_beschikbaar(ws_pp2, gebruik_rij, col):
                            continue
                        pp2_write_name(ws_pp2, gebruik_rij, col, naam)
                        kort_geplaatst = True
                        break
                    if kort_geplaatst:
                        break

                if not kort_geplaatst:
                    pp2_niet_geplaatst.append({
                        "naam": naam,
                        "reden": "groep 2 speciale lange werker: geen geldig kwartier gevonden"
                    })
            else:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "groep 2 speciale lange werker: geen geldig halfuur 12-14u gevonden"
                })
        pp2_speciale_teller += 1

    # 3) Eerst: alle minderjarigen die in stap 2 recht hebben op een halfuur
    #    krijgen hun EERSTE halfuur zo vroeg mogelijk (zelfde rondes-systeem
    #    als de gewone lange werkers).
    pp2_wachtrij_1ste_halfuur = [
        naam for naam in pp2_step2_minderjarigen
        if pp2_aantal_lange_pauzes_nodig_in_stap2(naam) > 0
        and naam not in pp2_lange_pauze_ontvangers
    ]

    while pp2_wachtrij_1ste_halfuur:
        vooruitgang = False
        nog_niet_geplaatst = []

        for naam in pp2_wachtrij_1ste_halfuur:
            if naam in pp2_lange_pauze_ontvangers:
                continue

            opties = pp2_verzamel_opties_alle_pvs(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
            if not opties:
                continue

            geplaatst = False
            for start_min, _schaarste, pv, pv_row, col1, col2 in opties:
                venster = pp2_tijdvenster_pauze([col1, col2], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
                    leave_top_blank = eigen_pv_row == pv_row
                    pp2_write_long_break(
                        ws_sheet=ws_pp2, pv_row=pv_row,
                        col1=col1, col2=col2, naam=naam,
                        leave_top_blank=leave_top_blank
                    )
                    pp2_lange_pauze_ontvangers.add(naam)
                    pp2_minderjarige_eerste_halfuur_rij[naam] = pv_row
                    geplaatst = True
                    vooruitgang = True
                    break

            if not geplaatst:
                nog_niet_geplaatst.append(naam)

        pp2_wachtrij_1ste_halfuur = nog_niet_geplaatst
        if not vooruitgang:
            break

    for naam in pp2_wachtrij_1ste_halfuur:
        if naam in pp2_lange_pauze_ontvangers:
            continue

        opties = pp2_verzamel_opties_alle_pvs(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
        if not opties:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "stap 2 (minderjarige, eerste halfuur): geen enkele vrije plek gevonden bij een PV"
            })
            continue

        _start_min, _schaarste, pv, pv_row, col1, col2 = opties[0]
        eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
        leave_top_blank = eigen_pv_row == pv_row

        pp2_write_long_break(
            ws_sheet=ws_pp2, pv_row=pv_row,
            col1=col1, col2=col2, naam=naam,
            leave_top_blank=leave_top_blank, conflict=True
        )
        pp2_lange_pauze_ontvangers.add(naam)
        pp2_minderjarige_eerste_halfuur_rij[naam] = pv_row
        
    
    # 4) Daarna: bestaande logica voor overige lange pauzevlinders op eigen rij
    for pv, pv_name_row in pv_rows_pp2:
        pp2_place_long_break_for_pv_in_own_row(
            pv=pv,
            pv_name_row=pv_name_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers,
            lange_werkers_random=pp2_lange_werkers_random
        )

    # Fallback: pauzevlinders wiens eigen rij (deels) afgeknipt is, en die
    # dus geen lange pauze op hun eigen rij konden krijgen, alsnog een
    # lange pauze geven bij een ANDERE pauzevlinder -- net als een gewone
    # lange werker. Ze houden nog steeds recht op hun pauze.
    for pv, pv_name_row in pv_rows_pp2:
        naam = pv["naam"]
        if naam not in pp2_lange_werkers_random:
            continue
        if naam in pp2_lange_pauze_ontvangers:
            continue

        gevonden = pp2_find_first_valid_long_block_any_row(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2, pauze_cols=pauze_cols_pp2
        )
        if gevonden is None:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "pauzevlinder (eigen rij afgeknipt): geen lange pauze gevonden bij een andere PV"
            })
            continue

        _fallback_pv_row, col1, col2, conflict = gevonden
        pp2_write_long_break(
            ws_sheet=ws_pp2,
            pv_row=_fallback_pv_row,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=False,
            conflict=conflict
        )
        pp2_lange_pauze_ontvangers.add(naam)

    # ---- Harde 60%-afkapregel voor de NORMALE lange pauzes ----
    # Geen enkele normale lange pauze mag nog starten op of na 60% van de
    # pauzevlinderuren. Rijen waar minderjarigen al hun eerste halfuur
    # kregen, krijgen een vroegere, eigen drempel: 30 min per (nog
    # haalbaar) tweede halfuur dat op die rij later nog moet komen, zodat
    # er ruimte gereserveerd blijft voor stap 6.
    pp2_pv_start_min = pp2_parse_kwartier_header(ws_pp2.cell(1, pauze_cols_pp2[0]).value)
    pp2_pv_eind_min = pp2_parse_kwartier_header(ws_pp2.cell(1, pauze_cols_pp2[-1]).value) + 15
    pp2_drempel_tijd_globaal = pp2_pv_start_min + 0.6 * (pp2_pv_eind_min - pp2_pv_start_min)

    pp2_drempel_per_rij = {}
    for pv, pv_name_row in pv_rows_pp2:
        wachtende_minderjarigen = [
            naam for naam, rij in pp2_minderjarige_eerste_halfuur_rij.items()
            if rij == pv_name_row and pp2_aantal_lange_pauzes_nodig_in_stap2(naam) >= 2
        ]
        # Sorteer op het echte tijdstip van hun EERSTE halfuur -> wie het
        # vroegst zijn eerste pauze had, wordt hier ook het eerst gesimuleerd.
        wachtende_minderjarigen.sort(
            key=lambda n: pp2_eerste_halfuur_start_min(n, ws_pp2, pv_name_row, pauze_cols_pp2) or 0
        )

        kandidaat_na = pp2_drempel_tijd_globaal - 60
        extra_bezet = set()
        haalbaar = 0
        for naam in wachtende_minderjarigen:
            plek = pp2_kan_2de_halfuur_na(
                naam, ws_pp2, pv_name_row, pauze_cols_pp2, kandidaat_na, extra_bezet
            )
            if plek is not None:
                haalbaar += 1
                extra_bezet.add(plek[0])
                extra_bezet.add(plek[1])

        pp2_drempel_per_rij[pv_name_row] = pp2_drempel_tijd_globaal - (30 * haalbaar)
    
    # 5) Daarna: algemene verdeling van andere lange werkers
    # Rondes-systeem: per kandidaat de eerstvolgende plek van élke PV
    # verzamelen (nooit verder gezocht dan die ene plek per PV). Enkel
    # plaatsen bij een gekwalificeerde optie. Lukt dat voor niemand meer
    # in een volledige ronde, dan pas rood plaatsen op de eigen vroegste optie.

    pp2_wachtrij = [
        naam for naam in pp2_step2_overige_lange_werkers
        if naam not in pp2_lange_pauze_ontvangers
    ]

    while pp2_wachtrij:
        vooruitgang = False
        nog_niet_geplaatst = []

        for naam in pp2_wachtrij:
            if naam in pp2_lange_pauze_ontvangers:
                continue

            # 1) per PV zijn eigen eerstvolgende plek verzamelen
            opties = []
            for pv, pv_name_row in pv_rows_pp2:
                plek = pp2_eerste_vrije_blok_op_rij(naam, ws_pp2, pv_name_row, pauze_cols_pp2)
                if plek is not None:
                    col1, col2 = plek
                    start_min = pp2_parse_kwartier_header(ws_pp2.cell(1, col1).value)
                    if start_min >= pp2_drempel_per_rij.get(pv_name_row, pp2_drempel_tijd_globaal):
                        continue  # 60%-drempel (eventueel vervroegd) overschreden op deze rij
                    opties.append((start_min, pp2_schaarste_pv(pv), pv, pv_name_row, col1, col2))

            if not opties:
                continue  # deze kandidaat heeft nergens (nog) een vrije plek

            opties.sort(key=lambda o: (o[0], o[1]))

            # 3) eerste gekwalificeerde optie plaatsen
            geplaatst = False
            for start_min, _schaarste, pv, pv_name_row, col1, col2 in opties:
                venster = pp2_tijdvenster_pauze([col1, col2], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    pp2_write_long_break(
                        ws_sheet=ws_pp2, pv_row=pv_name_row,
                        col1=col1, col2=col2, naam=naam,
                        leave_top_blank=False
                    )
                    pp2_lange_pauze_ontvangers.add(naam)
                    geplaatst = True
                    vooruitgang = True
                    break

            if not geplaatst:
                nog_niet_geplaatst.append(naam)

        pp2_wachtrij = nog_niet_geplaatst

        if not vooruitgang:
            break  # niemand kon nog vooruit -> wachtrij zit muurvast

    # Iedereen die overblijft: de drempel voorkwam een plek binnen de
    # tijdslimiet. Check nu, zonder drempel, of de vroegst/moeilijkst
    # mogelijke PV de attractie (en de volgende, binnen het halfuur
    # speling) wel degelijk aankan -- enkel dan pas rood.
    for naam in pp2_wachtrij:
        if naam in pp2_lange_pauze_ontvangers:
            continue

        opties = []
        for pv, pv_name_row in pv_rows_pp2:
            plek = pp2_eerste_vrije_blok_op_rij(naam, ws_pp2, pv_name_row, pauze_cols_pp2)
            if plek is not None:
                col1, col2 = plek
                start_min = pp2_parse_kwartier_header(ws_pp2.cell(1, col1).value)
                opties.append((start_min, pp2_schaarste_pv(pv), pv, pv_name_row, col1, col2))

        if not opties:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "stap 2 (gewone lange werkers): geen enkele vrije plek gevonden bij een PV"
            })
            continue

        opties.sort(key=lambda o: (o[0], o[1]))
        _start_min, _schaarste, pv, pv_name_row, col1, col2 = opties[0]

        venster = pp2_tijdvenster_pauze([col1, col2], ws_pp2)
        attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
        conflict = not pp2_pv_kan_overname(pv, attrs)

        pp2_write_long_break(
            ws_sheet=ws_pp2, pv_row=pv_name_row,
            col1=col1, col2=col2, naam=naam,
            leave_top_blank=False, conflict=conflict
        )
        pp2_lange_pauze_ontvangers.add(naam)
    
    # 6) Helemaal als laatste: minderjarigen met > 6u krijgen nog een TWEEDE
    #    halfuur. EERST altijd proberen op exact dezelfde rij als het eerste
    #    halfuur (glued, geen gat). Lukt dat niet (geen plek, of niet
    #    gekwalificeerd), dan pas een andere PV -- ook daar telkens enkel
    #    de eigen eerstvolgende plek.
    pp2_wachtrij_2de_halfuur = [
        naam for naam in pp2_step2_minderjarigen
        if pp2_aantal_lange_pauzes_nodig_in_stap2(naam) >= 2
        and pp2_minderjarige_eerste_halfuur_rij.get(naam) is not None
    ]
    pp2_wachtrij_2de_halfuur.sort(
        key=lambda n: pp2_eerste_halfuur_start_min(
            n, ws_pp2, pp2_minderjarige_eerste_halfuur_rij[n], pauze_cols_pp2
        ) or 0
    )
    while pp2_wachtrij_2de_halfuur:
        vooruitgang = False
        nog_niet_geplaatst = []

        for naam in pp2_wachtrij_2de_halfuur:
            vaste_rij = pp2_minderjarige_eerste_halfuur_rij.get(naam)
            vaste_pv = next((p for p, r in pv_rows_pp2 if r == vaste_rij), None)
            geplaatst = False

            # Stap A: eerst proberen op dezelfde rij als het eerste halfuur
            plek = pp2_eerste_vrije_blok_op_rij(naam, ws_pp2, vaste_rij, pauze_cols_pp2)
            if plek is not None and vaste_pv is not None:
                col1, col2 = plek
                venster = pp2_tijdvenster_pauze([col1, col2], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(vaste_pv, attrs):
                    eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
                    leave_top_blank = eigen_pv_row == vaste_rij
                    pp2_write_long_break(
                        ws_sheet=ws_pp2, pv_row=vaste_rij,
                        col1=col1, col2=col2, naam=naam,
                        leave_top_blank=leave_top_blank
                    )
                    pp2_lange_pauze_ontvangers.add(naam)
                    geplaatst = True
                    vooruitgang = True

            # Stap B: eigen rij lukt niet -> andere PV proberen
            if not geplaatst:
                opties = pp2_verzamel_opties_alle_pvs(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
                for start_min, _schaarste, pv, pv_row, col1, col2 in opties:
                    venster = pp2_tijdvenster_pauze([col1, col2], ws_pp2)
                    attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                    if pp2_pv_kan_overname(pv, attrs):
                        eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
                        leave_top_blank = eigen_pv_row == pv_row
                        pp2_write_long_break(
                            ws_sheet=ws_pp2, pv_row=pv_row,
                            col1=col1, col2=col2, naam=naam,
                            leave_top_blank=leave_top_blank
                        )
                        pp2_lange_pauze_ontvangers.add(naam)
                        geplaatst = True
                        vooruitgang = True
                        break

            if not geplaatst:
                nog_niet_geplaatst.append(naam)

        pp2_wachtrij_2de_halfuur = nog_niet_geplaatst
        if not vooruitgang:
            break

    # Overblijvers: zelfde volgorde (eigen rij eerst, dan andere PV's),
    # maar nu zonder kwalificatie-eis -- op de vroegste optie, rood.
    for naam in pp2_wachtrij_2de_halfuur:
        vaste_rij = pp2_minderjarige_eerste_halfuur_rij.get(naam)

        plek = pp2_eerste_vrije_blok_op_rij(naam, ws_pp2, vaste_rij, pauze_cols_pp2)
        if plek is not None:
            col1, col2 = plek
            pv_row_gekozen = vaste_rij
        else:
            opties = pp2_verzamel_opties_alle_pvs(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
            if not opties:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "stap 2 (minderjarige, tweede halfuur): geen enkele vrije plek gevonden bij een PV"
                })
                continue
            _start_min, _schaarste, _pv, pv_row_gekozen, col1, col2 = opties[0]

        eigen_pv_row = pp2_get_pv_row_for_name(naam, pv_rows_pp2)
        leave_top_blank = eigen_pv_row == pv_row_gekozen

        pp2_write_long_break(
            ws_sheet=ws_pp2, pv_row=pv_row_gekozen,
            col1=col1, col2=col2, naam=naam,
            leave_top_blank=leave_top_blank, conflict=True
        )
        pp2_lange_pauze_ontvangers.add(naam)
    
    
    #STAP 3 333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333333
    
    # -----------------------------
    # STAP 3 PP optie 2:
    # open spots berekenen en verdelen
    # + korte pauzes van pauzevlinders zelf invullen
    # -----------------------------
    
    lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    naam_leeg_fill_pp2 = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    
    
    def pp2_benodigde_korte_kwartieren(naam):
        # Bestaande uitzondering: speciale groep 1
        if naam in pp2_speciale_groep1_namen:
            return 0
    
        theo_duur = student_totalen.get(naam, 0)
        student = next((s for s in studenten if s["naam"] == naam), None)
    
        if student:
            begin_uur = student.get("begin_uur")
            eind_uur  = student.get("eind_uur")
    
            if begin_uur is not None and eind_uur is not None:
                echte_duur = eind_uur - begin_uur
    
                # Uitzondering 2: theoretische shift <= 2u
                if theo_duur <= 2:
                    return 0
    
                # Uitzondering 1: echte einduur - theo einduur >= 1.5u
                theo_uren = pp2_get_student_work_hours(naam)
                if theo_uren:
                    theo_eind = max(theo_uren) + blok_durations.get(max(theo_uren), 1.0)
                    if (eind_uur - theo_eind) >= 1.5:
                        return 0
    
                # Iedereen die echt >= 4u werkt krijgt een korte pauze
                # (bij PAUZE_STRIKT_BOVEN_4U aangevinkt: pas vanaf MEER dan 4u)
                if (echte_duur > 4) if PAUZE_STRIKT_BOVEN_4U else (echte_duur >= 4):
                    return 1
    
        # Fallback naar bestaande logica
        if PAUZE_STRIKT_BOVEN_4U:
            if werkduur_voor_pauze(naam) <= 4:
                return 0
        else:
            if werkduur_voor_pauze(naam) < 4:
                return 0
        return 1
    
    
    def pp2_count_total_assigned_quarters_for_student(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Telt alle kwartiercellen in PP optie 2 waar deze naam al staat.
        Dit zijn dus ALLE reeds toegekende pauzekwartieren samen.
        """
        count = 0
        for _pv, pv_row in pv_rows:
            for col in pauze_cols:
                if ws_sheet.cell(pv_row, col).value == naam:
                    count += 1
        return count
    
    
    def pp2_count_al_toegekende_lange_kwartieren(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Telt hoeveel reeds toegekende kwartieren deel uitmaken van een LANGE pauze
        voor deze student in PP optie 2.
    
        Een lange pauze herkennen we als 2 opeenvolgende kwartieren op dezelfde rij
        met exact dezelfde naam.
    
        Voorbeelden:
        - 1 halfuur lange pauze => 2 kwartieren
        - 2 halve uren lange pauze => 4 kwartieren
        """
        count = 0
        gebruikte_cols_per_row = set()
    
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if (pv_row, col1) in gebruikte_cols_per_row or (pv_row, col2) in gebruikte_cols_per_row:
                    continue
    
                val1 = ws_sheet.cell(pv_row, col1).value
                val2 = ws_sheet.cell(pv_row, col2).value
    
                if val1 == naam and val2 == naam:
                    count += 2
                    gebruikte_cols_per_row.add((pv_row, col1))
                    gebruikte_cols_per_row.add((pv_row, col2))
    
        return count
    
    
    def pp2_count_al_toegekende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        Telt hoeveel KORTE kwartieren deze student al heeft.
    
        Nieuwe logica:
        - tel eerst alle reeds ingevulde kwartieren van deze student
        - trek daar alle kwartieren af die deel uitmaken van een lange pauze
        - wat overblijft, zijn korte kwartieren
    
        Hierdoor werkt dit ook correct voor minderjarigen die 2 halve uren kregen.
        """
        totaal = pp2_count_total_assigned_quarters_for_student(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols
        )
    
        lange_kwartieren = pp2_count_al_toegekende_lange_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols
        )
    
        return max(0, totaal - lange_kwartieren)
    
    
    def pp2_resterende_korte_kwartieren(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        Hoeveel korte kwartieren heeft deze student nog nodig?
        """
        nodig = pp2_benodigde_korte_kwartieren(naam)
        al_kort = pp2_count_al_toegekende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=lange_pauze_ontvangers
        )
        return max(0, nodig - al_kort)
    
    
    def pp2_heeft_al_voldoende_korte_pauze(naam, ws_sheet, pv_rows, pauze_cols, lange_pauze_ontvangers):
        """
        True als student al genoeg korte kwartieren heeft gekregen
        volgens de nieuwe PP2-regels.
        """
        return pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=lange_pauze_ontvangers
        ) == 0
    
    
    def pp2_korte_pauze_nodig_namen():
        """
        Iedereen met minstens 4 uur werk heeft recht op 1 kort kwartier,
        BEHALVE minderjarige vroege stoppers.
    
        Minderjarige vroege stoppers:
        - minderjarig
        - minstens 4u gewerkt
        - laatste werkuur <= 15
        """
        namen = []
    
        for s in studenten:
            naam = s["naam"]
            werk_uren = pp2_get_student_work_hours(naam)
    
            is_minor_early_stopper = (
                pp2_is_minderjarig(naam)
                and ((len(werk_uren) > 4) if PAUZE_STRIKT_BOVEN_4U else (len(werk_uren) >= 4))
                and werk_uren
                and max(werk_uren) <= 15
            )
    
            if is_minor_early_stopper:
                continue

            if naam in pp2_speciale_groep1_namen:
                continue

            if pp2_benodigde_korte_kwartieren(naam) > 0:
                namen.append(naam)
        return namen
    
    
    
    def pp2_count_remaining_empty_quarters(ws_sheet, pv_rows, pauze_cols):
        """
        Telt alle nog lege kwartiercellen in de naamrijen van PP optie 2.
        """
        count = 0
        for _pv, pv_row in pv_rows:
            for col in pauze_cols:
                if ws_sheet.cell(pv_row, col).value in [None, ""]:
                    count += 1
        return count
    
    
    def pp2_get_empty_cols_for_pv_row(ws_sheet, pv_row, pauze_cols, open_spots_set):
        """
        Geeft alle lege kwartierkolommen terug voor deze pauzevlinder-rij,
        exclusief reeds gemarkeerde open spots.
        """
        cols = []
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
            if ws_sheet.cell(pv_row, col).value in [None, ""]:
                cols.append(col)
        return cols
    
    
    def pp2_mark_open_spot(ws_sheet, pv_row, col):
        """
        Open spot blijft gewoon blauw en leeg.
        """
        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
    
        cel = ws_sheet.cell(pv_row, col)
        cel.value = ""
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = naam_leeg_fill_pp2
    
    
    
    
    def pp2_write_short_break_for_pv(ws_sheet, pv_row, col, naam, conflict=False):
        """
        Schrijf 1 kort kwartier voor een pauzevlinder zelf:
        - bovenliggende cel leeg
        - naam paars
        """
        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
    
        cel = ws_sheet.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border
        cel.fill = conflict_fill if conflict else lichtpaars_fill


    def pp2_write_short_break_regular(ws_sheet, pv_row, col, naam, conflict=False):
        """
        Korte pauze voor gewone student:
        - bovenliggende cel = attractie (rood bij conflict)
        - naamcel:
            * rood bij conflict (niet-gekwalificeerde PV)
            * lichtgeel voor minderjarigen die >4u werken
            * lichtpaars voor alle andere korte pauzes
        """
        header = ws_sheet.cell(1, col).value
        uur = parse_header_uur(header)

        attr = vind_attractie_op_uur(naam, uur) if uur is not None else None

        top_cel = ws_sheet.cell(pv_row - 1, col)
        top_cel.value = attr if attr else ""
        top_cel.alignment = center_align
        top_cel.border = thin_border
        top_cel.fill = conflict_fill if conflict else PatternFill(fill_type=None)

        cel = ws_sheet.cell(pv_row, col)
        cel.value = naam
        cel.alignment = center_align
        cel.border = thin_border

        if conflict:
            cel.fill = conflict_fill
        else:
            cel.fill = lichtpaars_fill
            
    
    
    def pp2_find_short_break_cols_for_pv(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, needed_quarters):
        """
        Zoek geldige kolom/kolommen voor de korte pauze van een pauzevlinder in de eigen rij.
    
        - needed_quarters == 1:
          neem het eerstvolgende geldige vrije kwartier
    
        - needed_quarters == 2:
          neem de eerste geldige set van 2 opeenvolgende kwartieren
        """
        if needed_quarters <= 0:
            return []
    
        if needed_quarters == 1:
            for col in pauze_cols:
                if (pv_row, col) in open_spots_set:
                    continue
                if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                    continue
                return [col]
            return []
    
        if needed_quarters == 2:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if col2 != col1 + 1:
                    continue
    
                if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                    continue
    
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
    
                if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                    continue
                if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                    continue
    
                return [col1, col2]
    
            return []
    
        return []
    
    
    
    # 1) Is dit een korte dag?
    #    Korte dag = openingsuren zijn 6 uur of minder
    pp2_is_korte_dag = len(open_uren) <= 6
    
    pp2_open_spots = set()
    pp2_pv_short_breaks_placed = []
    
    # -----------------------------
    # Hulploop: korte pauzes van pauzevlinders zelf invullen
    #
    # Nieuwe regel:
    # - pauzevlinders die GEEN lange werker zijn: hier al plaatsen
    # - pauzevlinders die WEL lange werker zijn: nog NIET hier plaatsen
    #   -> die komen later in stap 4, na de korte werkers
    # -----------------------------
    for pv, pv_row in pv_rows_pp2:
        naam = pv["naam"]
    
        # Lange pauzevlinders hier nog overslaan:
        # hun korte pauze moet pas later komen
        if naam in pp2_lange_werkers_lijst():
            continue
    
        resterend_nodig = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend_nodig <= 0:
            continue
    
        gekozen_cols = pp2_find_short_break_cols_for_pv(
            naam=naam,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots,
            needed_quarters=resterend_nodig
        )

        if gekozen_cols:
            for col in gekozen_cols:
                pp2_write_short_break_for_pv(ws_pp2, pv_row, col, naam)

            pp2_pv_short_breaks_placed.append({
                "naam": naam,
                "pv_row": pv_row,
                "kolommen": gekozen_cols,
                "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
            })
            continue

        # Eigen rij lukt niet -> val terug op een andere pauzevlinder,
        # net als bij lange pauzevlinders.
        if resterend_nodig == 1:
            _opties = pp2_verzamel_opties_alle_pvs_kort(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots)
            for _start_min, _schaarste, _ander_pv, _ander_pv_row, _col in _opties:
                _venster = pp2_tijdvenster_pauze([_col], ws_pp2)
                _attrs = pp2_attracties_in_venster(naam, *_venster) if _venster else set()
                if pp2_pv_kan_overname(_ander_pv, _attrs):
                    pp2_write_short_break_regular(ws_sheet=ws_pp2, pv_row=_ander_pv_row, col=_col, naam=naam)
                    pp2_pv_short_breaks_placed.append({
                        "naam": naam,
                        "pv_row": _ander_pv_row,
                        "kolommen": [_col],
                        "tijden": [ws_pp2.cell(1, _col).value]
                    })
                    break
    
    # -----------------------------
    # 2) Tellen hoeveel kwartierblokjes nog leeg zijn
    # -----------------------------
    pp2_remaining_empty_quarters = pp2_count_remaining_empty_quarters(
        ws_sheet=ws_pp2,
        pv_rows=pv_rows_pp2,
        pauze_cols=pauze_cols_pp2
    )
    
    # -----------------------------
    # 3) Tellen hoeveel KORTE kwartieren nog gegeven moeten worden
    #    Nieuwe telling:
    #    - gewone student meestal 1
    #    - minderjarige >4u = 2
    #    - ook minderjarige >6u met al lange pauze telt hier nog voor 2
    # -----------------------------
    pp2_korte_pauze_gerechtigden = pp2_korte_pauze_nodig_namen()
    
    pp2_remaining_short_quarters_needed = 0
    for naam in pp2_korte_pauze_gerechtigden:
        pp2_remaining_short_quarters_needed += pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
    # -----------------------------
    # 4) Open spots berekenen
    # -----------------------------
    pp2_open_spots_count = pp2_remaining_empty_quarters - pp2_remaining_short_quarters_needed
    if pp2_open_spots_count < 0:
        pp2_open_spots_count = 0
    
    # -----------------------------
    # 5) Open spots verdelen
    #
    # KORTE DAG:
    # - eerst korte pauzes geplaatst
    # - dus open spots vallen automatisch NA de korte pauzes
    #
    # LANGE DAG:
    # - script blijft exact hetzelfde gedrag houden als nu
    #   => open spots verdelen zoals nu
    # -----------------------------
    
    if not pp2_is_korte_dag:
        # ---------------------------------------------------
        # LANGE DAG:
        # - Open spots eerst verdelen (zoals in het originele script)
        # - Daarna enkel de korte pauzes van KORTE pauzevlinders plaatsen
        # - Lange pauzevlinders komen pas in stap 4 aan bod
        # ---------------------------------------------------
    
        # Reset eerst eventuele eerder geplaatste korte pauzes van pauzevlinders
        for item in pp2_pv_short_breaks_placed:
            naam = item["naam"]
            pv_row = item["pv_row"]
    
            for col in item["kolommen"]:
                top_cel = ws_pp2.cell(pv_row - 1, col)
                top_cel.value = ""
                top_cel.alignment = center_align
                top_cel.border = thin_border
    
                cel = ws_pp2.cell(pv_row, col)
                cel.value = ""
                cel.alignment = center_align
                cel.border = thin_border
                cel.fill = naam_leeg_fill_pp2
    
        pp2_pv_short_breaks_placed = []
    
        # ---------------------------------------------------
        # 1) Open spots verdelen
        # ---------------------------------------------------
        ronde_nummer = 0
    
        while len(pp2_open_spots) < pp2_open_spots_count:
            iets_geplaatst_deze_ronde = False
            vooraan = (ronde_nummer % 2 == 0)
    
            for _pv, pv_row in pv_rows_pp2:
                if len(pp2_open_spots) >= pp2_open_spots_count:
                    break

                if (
                    pp2_afgeknipte_pv_naam is not None
                    and _pv["naam"] == pp2_afgeknipte_pv_naam
                    and pp2_open_spots_afgeknipte_teller >= pp2_max_open_spots_afgeknipt
                ):
                    continue
    
                lege_cols = pp2_get_empty_cols_for_pv_row(
                    ws_sheet=ws_pp2,
                    pv_row=pv_row,
                    pauze_cols=pauze_cols_pp2,
                    open_spots_set=pp2_open_spots
                )
    
                if not lege_cols:
                    continue
    
                gekozen_col = lege_cols[0] if vooraan else lege_cols[-1]
    
                pp2_open_spots.add((pv_row, gekozen_col))
                pp2_mark_open_spot(ws_pp2, pv_row, gekozen_col)
                if pp2_afgeknipte_pv_naam is not None and _pv["naam"] == pp2_afgeknipte_pv_naam:
                    pp2_open_spots_afgeknipte_teller += 1
                iets_geplaatst_deze_ronde = True
    
            if not iets_geplaatst_deze_ronde:
                break
    
            ronde_nummer += 1
    
        # ---------------------------------------------------
        # 2) Enkel korte pauzes van KORTE pauzevlinders plaatsen
        # Lange pauzevlinders worden hier overgeslagen
        # ---------------------------------------------------
        for pv, pv_row in pv_rows_pp2:
            naam = pv["naam"]
    
            # Lange pauzevlinders hier overslaan
            if naam in pp2_lange_werkers_lijst():
                continue
    
            resterend_nodig = pp2_resterende_korte_kwartieren(
                naam=naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            )
    
            if resterend_nodig <= 0:
                continue
    
            gekozen_cols = pp2_find_short_break_cols_for_pv(
                naam=naam,
                pv_row=pv_row,
                ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots,
                needed_quarters=resterend_nodig
            )

            if gekozen_cols:
                for col in gekozen_cols:
                    pp2_write_short_break_for_pv(
                        ws_sheet=ws_pp2,
                        pv_row=pv_row,
                        col=col,
                        naam=naam
                    )

                pp2_pv_short_breaks_placed.append({
                    "naam": naam,
                    "pv_row": pv_row,
                    "kolommen": gekozen_cols,
                    "tijden": [
                        ws_pp2.cell(1, col).value for col in gekozen_cols
                    ]
                })
                continue

            # Eigen rij lukt niet -> val terug op een andere pauzevlinder.
            if resterend_nodig == 1:
                _opties = pp2_verzamel_opties_alle_pvs_kort(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots)
                for _start_min, _schaarste, _ander_pv, _ander_pv_row, _col in _opties:
                    _venster = pp2_tijdvenster_pauze([_col], ws_pp2)
                    _attrs = pp2_attracties_in_venster(naam, *_venster) if _venster else set()
                    if pp2_pv_kan_overname(_ander_pv, _attrs):
                        pp2_write_short_break_regular(ws_sheet=ws_pp2, pv_row=_ander_pv_row, col=_col, naam=naam)
                        pp2_pv_short_breaks_placed.append({
                            "naam": naam,
                            "pv_row": _ander_pv_row,
                            "kolommen": [_col],
                            "tijden": [ws_pp2.cell(1, _col).value]
                        })
                        break
    
    
    # ---------------------------------------------------
    # KORTE DAG: open spots verdelen vóór stap 4
    # Op een korte dag worden open spots niet in het
    # pauzevlinder-blok hierboven verdeeld.
    # We doen dat hier alsnog, vóór stap 4 korte pauzes
    # plaatst, zodat die cellen correct worden overgeslagen.
    # pp2_open_spots_count = remaining_empty - needed_short,
    # al berekend op basis van wat er na stap 1/2/2b over is.
    # ---------------------------------------------------
    if pp2_is_korte_dag:
        ronde_nummer = 0
        while len(pp2_open_spots) < pp2_open_spots_count:
            iets_geplaatst_deze_ronde = False
            for _pv, pv_row in pv_rows_pp2:
                if len(pp2_open_spots) >= pp2_open_spots_count:
                    break

                if (
                    pp2_afgeknipte_pv_naam is not None
                    and _pv["naam"] == pp2_afgeknipte_pv_naam
                    and pp2_open_spots_afgeknipte_teller >= pp2_max_open_spots_afgeknipt
                ):
                    continue

                lege_cols = pp2_get_empty_cols_for_pv_row(
                    ws_sheet=ws_pp2,
                    pv_row=pv_row,
                    pauze_cols=pauze_cols_pp2,
                    open_spots_set=pp2_open_spots
                )
                if not lege_cols:
                    continue
                # Altijd de laatste lege kolom nemen
                gekozen_col = lege_cols[-1]
                pp2_open_spots.add((pv_row, gekozen_col))
                pp2_mark_open_spot(ws_pp2, pv_row, gekozen_col)
                if pp2_afgeknipte_pv_naam is not None and _pv["naam"] == pp2_afgeknipte_pv_naam:
                    pp2_open_spots_afgeknipte_teller += 1
                iets_geplaatst_deze_ronde = True
            if not iets_geplaatst_deze_ronde:
                break
            ronde_nummer += 1

    
    #STAP 4 44444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444444
    
    
    # -----------------------------
    # STAP 4 PP optie 2:
    # korte pauzes voor:
    # 1) studenten die vroeger stoppen dan het einduur
    # 2) daarna lange pauzevlinders zelf (in eigen rij)
    # met nieuwe minderjarigenlogica
    # -----------------------------
    
    lichtpaars_fill = PatternFill(start_color="E6DAF7", end_color="E6DAF7", fill_type="solid")
    
    
    def pp2_get_day_end_hour():
        """
        Einduur van de dag op basis van open_uren.
        """
        if not open_uren:
            return None
        return max(open_uren)
    
    
    def pp2_get_students_stopping_before_end():
        """
        Studenten die vroeger stoppen dan het einduur van de dag
        en minstens 4 uur werken.
        """
        einduur_dag = pp2_get_day_end_hour()
        result = []
    
        if einduur_dag is None:
            return result
    
        for s in studenten:
            naam = s["naam"]
            werk_uren = pp2_get_student_work_hours(naam)
    
            if len(werk_uren) < 4:
                continue
    
            if max(werk_uren) < einduur_dag:
                result.append(naam)
    
        return result
    
    
    
    def pp2_get_long_break_owners_on_row(ws_sheet, pv_row, pauze_cols):
        """
        Geeft alle studenten terug die op deze rij een lange pauze hebben,
        gesorteerd op het ankerpunt voor hun korte pauze:
        - minderjarige lange werkers (>6u): gesorteerd op hun LAATSTE halfuur
        - alle anderen: gesorteerd op hun EERSTE halfuur (= volgorde van links naar rechts)
        """
        # Verzamel per student de eerste én laatste halfuur-eindkolom op deze rij
        eerste_col = {}
        laatste_col = {}
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            val1 = ws_sheet.cell(pv_row, col1).value
            val2 = ws_sheet.cell(pv_row, col2).value
    
            if val1 and val1 == val2:
                naam = str(val1).strip()
                if naam not in eerste_col:
                    eerste_col[naam] = col2
                laatste_col[naam] = col2
    
        owners = list(eerste_col.keys())
    
        def sorteersleutel(naam):
            is_minor_long_worker = (
                pp2_is_minderjarig(naam)
                and werkduur_voor_pauze(naam) > 6
            )
            if is_minor_long_worker:
                return laatste_col.get(naam, 0)
            else:
                return eerste_col.get(naam, 0)
    
        owners.sort(key=sorteersleutel)
        return owners
    
    def pp2_student_has_long_break_in_row(naam, ws_sheet, pv_row, pauze_cols):
        """
        Check of deze student een lange pauze heeft op precies deze rij.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            if (
                ws_sheet.cell(pv_row, col1).value == naam and
                ws_sheet.cell(pv_row, col2).value == naam
            ):
                return True
    
        return False
    
    
    def pp2_student_is_long_worker(naam):
        return naam in pp2_lange_werkers_lijst()
    
    
    def pp2_find_two_consecutive_valid_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
        """
        Zoek 2 opeenvolgende geldige kwartieren voor deze student op deze specifieke rij.
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            if col2 != col1 + 1:
                continue
    
            if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                continue
            if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col1, ws_sheet):
                continue
            if not pp2_is_valid_short_break_for_student(naam, col2, ws_sheet):
                continue
    
            return [col1, col2]
    
        return []
    
    
    def pp2_find_one_valid_col_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set):
        """
        Zoek 1 geldig kwartier voor deze student op deze specifieke rij.
        """
        for col in pauze_cols:
            if (pv_row, col) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
    
            return [col]
    
        return []
    
    
    def pp2_find_needed_short_cols_for_student_on_row(naam, pv_row, ws_sheet, pauze_cols, open_spots_set, min_col_exclusive=None, zoek_zo_laat_mogelijk=False):
        """
        Zoek het korte kwartier dat deze student nog nodig heeft op deze specifieke rij.
    
        - min_col_exclusive: zoek pas NA deze kolom
        - zoek_zo_laat_mogelijk: zoek van rechts naar links (voor minderjarige lange werkers)
        """
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_sheet,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend <= 0:
            return []
    
        kandidaat_cols = list(reversed(pauze_cols)) if zoek_zo_laat_mogelijk else list(pauze_cols)
    
        for col in kandidaat_cols:
            if min_col_exclusive is not None and col <= min_col_exclusive:
                continue
    
            if (pv_row, col) in open_spots_set:
                continue
    
            if ws_sheet.cell(pv_row, col).value not in [None, ""]:
                continue
    
            if not pp2_is_valid_short_break_for_student(naam, col, ws_sheet):
                continue
    
            return [col]
    
        return []
    
    
    
    def pp2_place_short_break_cols_on_row(naam, pv, pv_row, cols, conflict=False):
        """
        Schrijf 1 of 2 korte kwartieren voor gewone student op een bepaalde rij.
        """
        for col in cols:
            pp2_write_short_break_regular(
                ws_sheet=ws_pp2,
                pv_row=pv_row,
                col=col,
                naam=naam,
                conflict=conflict
            )

        pp2_regular_short_breaks_placed.append({
            "naam": naam,
            "pauzevlinder": pv["naam"],
            "tijden": [ws_pp2.cell(1, col).value for col in cols],
            "zelfde_rij_als_lange_pauze": pp2_student_has_long_break_in_row(
                naam, ws_pp2, pv_row, pauze_cols_pp2
            )
        })
    
    
    
    
    
    def pp2_student_heeft_nog_lange_pauze_nodig(naam, ws_sheet, pv_rows, pauze_cols):
        """
        Bepaal of deze student volgens de nieuwe regels nog minstens 1 lang halfuur mist.
    
        Regels minderjarigen:
        - < 4u gewerkt => 0 lange pauzes
        - 4u t.e.m. 6u => 1 lange pauze
        - > 6u => 2 lange pauzes
    
        Regels meerderjarigen:
        - > 6u => 1 lange pauze
        - anders => 0
        """
        gewerkte_uren = werkduur_voor_pauze(naam)
        is_minor = pp2_is_minderjarig(naam)
    
        if is_minor:
            if PAUZE_STRIKT_BOVEN_4U:
                if gewerkte_uren <= 4:
                    nodig = 0
                elif gewerkte_uren <= 6:
                    nodig = 1
                else:
                    nodig = 2
            else:
                if gewerkte_uren < 4:
                    nodig = 0
                elif gewerkte_uren <= 6:
                    nodig = 1
                else:
                    nodig = 2
        else:
            nodig = 1 if gewerkte_uren > 6 else 0
    
        # tel hoeveel lange halve uren al effectief ingepland zijn
        al = 0
        for _pv, pv_row in pv_rows:
            for idx in range(len(pauze_cols) - 1):
                col1 = pauze_cols[idx]
                col2 = pauze_cols[idx + 1]
    
                if (
                    ws_sheet.cell(pv_row, col1).value == naam and
                    ws_sheet.cell(pv_row, col2).value == naam
                ):
                    al += 1
    
        return al < nodig
    
    
    def pp2_find_first_valid_long_block_in_step4(naam, ws_sheet, pv_rows, pauze_cols, open_spots_set):
        """
        Zoek in stap 4 een geldig halfuur voor een student.
    
        Voor minderjarige vroege stoppers:
        - kies het EERSTE geldige halfuur (dus zo vroeg mogelijk)
    
        Voor alle anderen:
        - behoud ook het eerste geldige halfuur
        """
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            for _pv, pv_row in pv_rows:
                if (pv_row, col1) in open_spots_set or (pv_row, col2) in open_spots_set:
                    continue
    
                if ws_sheet.cell(pv_row, col1).value not in [None, ""]:
                    continue
                if ws_sheet.cell(pv_row, col2).value not in [None, ""]:
                    continue
    
                if not pp2_is_valid_long_break_for_student(naam, col1, col2, ws_sheet):
                    continue
    
                return pv_row, col1, col2
    
        return None
    
    # ---------------------------------------
    # 0) Eerst: minderjarigen die nog een LANGE pauze missen alsnog proberen plaatsen
    #    Dit vangt het geval op waarin een minderjarige laat start
    #    en stap 2 geen geldig halfuur vond.
    # ---------------------------------------
    pp2_step4_late_long_break_rescue = []
    pp2_regular_short_breaks_placed = []
    
    # ---------------------------------------
    # 0A) Eerst: minderjarige vroege stoppers
    #     die nog een LANGE pauze missen
    #     => zo vroeg mogelijk (links naar rechts)
    # ---------------------------------------
    pp2_minor_early_stoppers = [
        s["naam"] for s in studenten
        if (
            pp2_is_minderjarig(s["naam"])
            and ((len(pp2_get_student_work_hours(s["naam"])) > 4) if PAUZE_STRIKT_BOVEN_4U else (len(pp2_get_student_work_hours(s["naam"])) >= 4))
            and pp2_get_student_work_hours(s["naam"])
            and max(pp2_get_student_work_hours(s["naam"])) <= 15
        )
    ]
    
    for naam in pp2_minor_early_stoppers:
        if not pp2_student_heeft_nog_lange_pauze_nodig(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2
        ):
            continue

        gevonden = pp2_find_first_valid_long_block_any_row(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2, pauze_cols=pauze_cols_pp2
        )
        if gevonden is None:
            continue

        pv_row, col1, col2, conflict = gevonden

        pp2_write_long_break(
            ws_sheet=ws_pp2,
            pv_row=pv_row,
            col1=col1,
            col2=col2,
            naam=naam,
            leave_top_blank=False,
            conflict=conflict
        )

        pp2_lange_pauze_ontvangers.add(naam)

        pp2_step4_late_long_break_rescue.append({
            "naam": naam,
            "tijden": [ws_pp2.cell(1, col1).value, ws_pp2.cell(1, col2).value]
        })
    
    
    # ---------------------------------------
    # 0B) Daarna: exact diezelfde minderjarige
    #     vroege stoppers hun KORTE pauze
    #     => zo laat mogelijk, bij voorkeur op dezelfde rij
    # ---------------------------------------
    for naam in pp2_minor_early_stoppers:
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
        if resterend <= 0:
            continue

        laatste_lange_eindcol = None
        laatste_lange_rij = None
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if laatste_lange_eindcol is None or col2 > laatste_lange_eindcol:
                        laatste_lange_eindcol = col2
                        laatste_lange_rij = pv_row

        if laatste_lange_eindcol is None:
            continue

        gevonden = pp2_zoek_laatste_kort_kwartier_vanaf(
            naam=naam, ws_sheet=ws_pp2, min_col_exclusive=laatste_lange_eindcol,
            voorkeur_rij=laatste_lange_rij, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, open_spots_set=pp2_open_spots
        )
        if gevonden is None:
            continue

        pv, pv_row, col, conflict = gevonden
        pp2_place_short_break_cols_on_row(
            naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=conflict
        )


    # ---------------------------------------
    # KORTE-ONLY WERKERS (geen lange pauze, geen pauzevlinder zelf) --
    # ONGEACHT of ze voor of tot het einduur werken, BEHALVE wie de
    # laatste paar uur van de dag werkt en tot sluiting blijft -- die
    # blijft bewust laat (valt terug in de bestaande, late STAP5-3-staart).
    # ---------------------------------------
    _pauzevlinder_namen_vroeg = {pv["naam"] for pv, _ in pv_rows_pp2}

    _pv_eind_min_check = pp2_parse_kwartier_header(ws_pp2.cell(1, pauze_cols_pp2[-1]).value) + 15
    _dag_eind_uur_check = pp2_get_day_end_hour()
    _dag_eind_min_check = (_dag_eind_uur_check * 60) if _dag_eind_uur_check is not None else _pv_eind_min_check
    _gap_min_check = _dag_eind_min_check - _pv_eind_min_check
    _drempel_uren_laat = 4 if _gap_min_check <= 60 else 5

    def _werkt_tot_einduur_check(naam):
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren or not open_uren:
            return False
        return max(werk_uren) == max(open_uren)

    pp2_korte_only_kandidaten = [
        naam for naam in [s["naam"] for s in studenten]
        if naam not in pp2_minor_early_stoppers
        and naam not in pp2_lange_pauze_ontvangers
        and naam not in _pauzevlinder_namen_vroeg
        and not (
            pp2_is_minderjarig(naam)
            and ((len(pp2_get_student_work_hours(naam)) > 4) if PAUZE_STRIKT_BOVEN_4U else (len(pp2_get_student_work_hours(naam)) >= 4))
            and pp2_get_student_work_hours(naam)
            and max(pp2_get_student_work_hours(naam)) <= 15
        )
        and not (
            _werkt_tot_einduur_check(naam)
            and len(pp2_get_student_work_hours(naam)) <= _drempel_uren_laat
        )
        and pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        ) > 0
    ]

    random.shuffle(pp2_korte_only_kandidaten)
    pp2_wachtrij_korte_only = list(pp2_korte_only_kandidaten)

    while pp2_wachtrij_korte_only:
        vooruitgang = False
        nog_niet_geplaatst = []
        for naam in pp2_wachtrij_korte_only:
            if pp2_resterende_korte_kwartieren(
                naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            ) <= 0:
                continue
            opties = pp2_verzamel_opties_alle_pvs_kort(
                naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
            )
            if not opties:
                nog_niet_geplaatst.append(naam)
                continue
            geplaatst = False
            for start_min, _schaarste, pv, pv_row, col in opties:
                venster = pp2_tijdvenster_pauze([col], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                    geplaatst = True
                    vooruitgang = True
                    break
            if not geplaatst:
                nog_niet_geplaatst.append(naam)
        pp2_wachtrij_korte_only = nog_niet_geplaatst
        if not vooruitgang:
            break

    for naam in pp2_wachtrij_korte_only:
        if pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        ) <= 0:
            continue
        opties = pp2_verzamel_opties_alle_pvs_kort(
            naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
        )
        if not opties:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "korte-only werkers: geen enkele vrije plek gevonden bij een PV"
            })
            continue
        geplaatst = False
        for start_min, _schaarste, pv, pv_row, col in opties:
            venster = pp2_tijdvenster_pauze([col], ws_pp2)
            attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
            if pp2_pv_kan_overname(pv, attrs):
                pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                geplaatst = True
                break
        if geplaatst:
            continue
        _start_min, _schaarste, pv, pv_row, col = opties[0]
        pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=True)
        
    # ---------------------------------------
    # 1) Daarna pas: gewone korte werkers
    #    die vroeger stoppen dan het einde
    #    van de dag, maar GEEN minderjarige
    #    vroege stoppers zijn
    # ---------------------------------------
    pp2_students_before_end_all = pp2_get_students_stopping_before_end()
    
    pp2_pauzevlinder_namen_alle = {pv["naam"] for pv, _ in pv_rows_pp2}

    pp2_students_before_end_pending = [
        naam for naam in pp2_students_before_end_all
        if naam not in pp2_minor_early_stoppers
        and naam not in pp2_pauzevlinder_namen_alle
        and pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        ) > 0
    ]
    
    def pp2_get_last_long_break_end_col_for_sort(naam):
        """
        Geeft de eindkolom van het LAATSTE halfuur van deze student terug,
        over alle PV-rijen heen. Studenten zonder lange pauze krijgen -1,
        zodat ze vooraan komen in de sortering.
        """
        eindcol = -1
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if col2 > eindcol:
                        eindcol = col2
        return eindcol
    
    pp2_students_before_end_pending.sort(
        key=lambda naam: pp2_get_last_long_break_end_col_for_sort(naam)
    )

    # Willekeurige volgorde voor wie geen lange pauze heeft (zij staan
    # toch al vooraan, want hun sorteersleutel is overal -1) -- de
    # volgorde van wie al wél een lange pauze heeft, blijft gesorteerd
    # op eindkolom.
    _idx_eerste_met_lange = next(
        (i for i, n in enumerate(pp2_students_before_end_pending)
         if pp2_get_last_long_break_end_col_for_sort(n) != -1),
        len(pp2_students_before_end_pending)
    )
    _zonder_lange = pp2_students_before_end_pending[:_idx_eerste_met_lange]
    _met_lange = pp2_students_before_end_pending[_idx_eerste_met_lange:]
    random.shuffle(_zonder_lange)
    pp2_students_before_end_pending = _zonder_lange + _met_lange
    
    pp2_regular_short_breaks_placed = []
    
    
    
    for naam in pp2_students_before_end_pending[:]:
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
        if resterend <= 0:
            pp2_students_before_end_pending.remove(naam)
            continue

        eigen_rij = pp2_vind_rij_met_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
        is_minor_long_worker = (
            pp2_is_minderjarig(naam) and werkduur_voor_pauze(naam) > 6
        )

        geplaatst = False

        # ---------------------------------------------------
        # Stap A: EERST de eigen rij proberen -- zelfde ankerpunt/
        # zoekrichting als voorheen (zo laat mogelijk voor minderjarige
        # lange werkers, na hun laatste halfuur op deze rij).
        # ---------------------------------------------------
        if eigen_rij is not None and pp2_student_has_long_break_in_row(naam, ws_pp2, eigen_rij, pauze_cols_pp2):
            eigen_pv = next((p for p, r in pv_rows_pp2 if r == eigen_rij), None)

            ankercol = pp2_lange_pauze_eindkolom_op_rij(naam, ws_pp2, eigen_rij, pauze_cols_pp2)

            cols = pp2_find_needed_short_cols_for_student_on_row(
                naam=naam, pv_row=eigen_rij, ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2, open_spots_set=pp2_open_spots,
                min_col_exclusive=ankercol, zoek_zo_laat_mogelijk=is_minor_long_worker
            )

            if cols and eigen_pv is not None:
                venster = pp2_tijdvenster_pauze(cols, ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(eigen_pv, attrs):
                    pp2_place_short_break_cols_on_row(naam=naam, pv=eigen_pv, pv_row=eigen_rij, cols=cols)
                    geplaatst = True

        # ---------------------------------------------------
        # Stap B: eigen rij lukt niet -> andere PV (vroegste, bij
        # gelijkstand moeilijkste, geplakt).
        # ---------------------------------------------------
        if not geplaatst:
            opties = pp2_verzamel_opties_alle_pvs_kort(
                naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
            )
            for start_min, _schaarste, pv, pv_row, col in opties:
                venster = pp2_tijdvenster_pauze([col], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                    geplaatst = True
                    break

        # ---------------------------------------------------
        # Stap C: nergens gekwalificeerd -> beste beschikbare optie, rood.
        # ---------------------------------------------------
        if not geplaatst:
            opties = pp2_verzamel_opties_alle_pvs_kort(
                naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
            )
            if opties:
                _start_min, _schaarste, pv, pv_row, col = opties[0]
                pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=True)
                geplaatst = True
            else:
                pp2_niet_geplaatst.append({
                    "naam": naam,
                    "reden": "stap 4 (lange werkers): geen enkele vrije plek gevonden bij een PV"
                })

        if geplaatst and naam in pp2_students_before_end_pending:
            pp2_students_before_end_pending.remove(naam)
    
    # ---------------------------------------
    # 2) Daarna: lange pauzevlinders zelf
    #    - alleen die nog korte kwartieren nodig hebben
    #    - alleen in eigen rij
    #    - na korte werkers
    #    - voor andere lange werkers
    # ---------------------------------------
    pp2_lange_pv_short_breaks_placed = []
    
    for pv, pv_row in pv_rows_pp2:
        naam = pv["naam"]
    
        if not pp2_student_is_long_worker(naam):
            continue
    
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend <= 0:
            continue
    
        gekozen_cols = pp2_find_needed_short_cols_for_student_on_row(
            naam=naam,
            pv_row=pv_row,
            ws_sheet=ws_pp2,
            pauze_cols=pauze_cols_pp2,
            open_spots_set=pp2_open_spots
        )

        if gekozen_cols:
            for col in gekozen_cols:
                pp2_write_short_break_for_pv(
                    ws_sheet=ws_pp2,
                    pv_row=pv_row,
                    col=col,
                    naam=naam
                )

            pp2_lange_pv_short_breaks_placed.append({
                "naam": naam,
                "tijden": [ws_pp2.cell(1, col).value for col in gekozen_cols]
            })
            continue

        # Eigen rij lukt niet (bv. (deels) afgeknipt) -> val terug op een
        # andere pauzevlinder, net als een gewone lange werker.
        _opties_fallback = pp2_verzamel_opties_alle_pvs_kort(
            naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
        )
        _geplaatst_fallback = False
        for _start_min, _schaarste, _ander_pv, _ander_pv_row, _col in _opties_fallback:
            _venster = pp2_tijdvenster_pauze([_col], ws_pp2)
            _attrs = pp2_attracties_in_venster(naam, *_venster) if _venster else set()
            if pp2_pv_kan_overname(_ander_pv, _attrs):
                pp2_write_short_break_regular(ws_sheet=ws_pp2, pv_row=_ander_pv_row, col=_col, naam=naam)
                pp2_lange_pv_short_breaks_placed.append({
                    "naam": naam,
                    "tijden": [ws_pp2.cell(1, _col).value]
                })
                _geplaatst_fallback = True
                break

        if not _geplaatst_fallback:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "pauzevlinder (eigen rij afgeknipt): geen kort kwartier gevonden bij een andere PV"
            })
    
    
    # STAP 5 55555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555555

    
    
    # -----------------------------------
    # STAP 5 PP optie 2:
    # laatste resterende korte kwartieren invullen
    # - werkt met resterende kwartieren i.p.v. ja/nee
    # - minderjarigen >4u krijgen hier ook 2 opeenvolgende kwartieren
    # - eerst overige pending korte kwartieren
    # - pas daarna eindwerkers zonder lange pauze
    # -----------------------------------
    
    def pp2_get_long_break_students_on_row_in_order(ws_sheet, pv_row, pauze_cols):
        """
        Geef de studenten terug die op deze rij een lange pauze hebben,
        in dezelfde volgorde als de lange pauzes op de rij zelf:
        dus van links naar rechts.
    
        Dit zorgt ervoor dat korte pauzes later ook in een logische
        volgorde kunnen volgen, gelijklopend met de lange pauzes.
        """
        found = {}
    
        for idx in range(len(pauze_cols) - 1):
            col1 = pauze_cols[idx]
            col2 = pauze_cols[idx + 1]
    
            val1 = ws_sheet.cell(pv_row, col1).value
            val2 = ws_sheet.cell(pv_row, col2).value
    
            if val1 and val1 == val2:
                naam = str(val1).strip()
                # bewaar de startkolom van de eerste lange pauze op deze rij
                if naam not in found:
                    found[naam] = col1
    
        return [naam for naam, _col in sorted(found.items(), key=lambda x: x[1])]
    
    
    
    
    def pp2_student_works_until_day_end(naam):
        """
        True als student werkt tot het einduur van de dag.
        """
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren or not open_uren:
            return False
        return max(werk_uren) == max(open_uren)
    
    
    def pp2_build_step5_pending_groups():
        """
        Splits alle NIET-pauzevlinders die nog korte kwartieren nodig hebben in:
        A) overige pending korte kwartieren
        B) eindwerkers zonder lange pauze
    
        Minderjarige vroege stoppers horen hier ook NIET meer in:
        die werden al eerder apart behandeld en mogen in stap 5
        niet opnieuw een kort kwartier krijgen.
    
        Pauzevlinders zelf horen hier ook niet meer in:
        - korte pauzevlinders werden al eerder verwerkt
        - lange pauzevlinders kregen in stap 4 hun eigen aparte fase,
          enkel in hun eigen rij
        """
        pauzevlinder_namen_set = {pv["naam"] for pv in selected}
        all_pending = []
    
        for s in studenten:
            naam = s["naam"]
    
            # Pauzevlinders hier NIET meer meenemen
            if naam in pauzevlinder_namen_set:
                continue
    
            # Minderjarige vroege stoppers hier ook NIET meer meenemen
            if (
                pp2_is_minderjarig(naam)
                and ((len(pp2_get_student_work_hours(naam)) > 4) if PAUZE_STRIKT_BOVEN_4U else (len(pp2_get_student_work_hours(naam)) >= 4))
                and pp2_get_student_work_hours(naam)
                and max(pp2_get_student_work_hours(naam)) <= 15
            ):
                continue
    
            resterend = pp2_resterende_korte_kwartieren(
                naam=naam,
                ws_sheet=ws_pp2,
                pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2,
                lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            )
    
            if resterend > 0:
                all_pending.append(naam)
    
        endworkers_without_long_break = []
        other_pending_short_breaks = []
    
        for naam in all_pending:
            heeft_lange = (naam in pp2_lange_pauze_ontvangers)
            werkt_tot_einduur = pp2_student_works_until_day_end(naam)
    
            if werkt_tot_einduur and not heeft_lange:
                endworkers_without_long_break.append(naam)
            else:
                other_pending_short_breaks.append(naam)
    
        # Wie al een lange pauze heeft, blijft gesorteerd op het eindpunt
        # van die lange pauze (zelfde principe als STAP4-1) -- enkel wie
        # geen lange pauze heeft, wordt willekeurig geschud.
        def _eindcol_voor_sortering_stap5(naam):
            for _pv, pv_row in pv_rows_pp2:
                col = pp2_lange_pauze_eindkolom_op_rij(naam, ws_pp2, pv_row, pauze_cols_pp2)
                if col is not None:
                    return col
            return -1

        _met_lange_p5 = [n for n in other_pending_short_breaks if n in pp2_lange_pauze_ontvangers]
        _zonder_lange_p5 = [n for n in other_pending_short_breaks if n not in pp2_lange_pauze_ontvangers]
        _met_lange_p5.sort(key=_eindcol_voor_sortering_stap5)
        random.shuffle(_zonder_lange_p5)
        other_pending_short_breaks = _zonder_lange_p5 + _met_lange_p5

        random.shuffle(endworkers_without_long_break)
    
        return other_pending_short_breaks, endworkers_without_long_break
    
    
    def pp2_try_assign_from_candidate_list_on_row(candidate_list, pv, pv_row, shuffle_candidates=False):
        """
        Probeer op deze rij een kandidaat te plaatsen uit de opgegeven lijst.
        Werkt met 1 of 2 kwartieren, afhankelijk van wat nog nodig is.
    
        Belangrijk:
        - de volgorde van candidate_list blijft behouden als dat een prioriteitslijst is
          (bv. dezelfde volgorde als de lange pauzes op de rij)
        - voor gewone fallback-lijsten kan shuffle_candidates=True gebruikt worden
          zodat niet-minderjarigen daar opnieuw randomer verdeeld worden
        """
        kandidaten = candidate_list[:]
    
        if shuffle_candidates and len(kandidaten) > 1:
            random.shuffle(kandidaten)
    
        for kandidaat in kandidaten:
            cols = pp2_find_needed_short_cols_for_student_on_row(
                naam=kandidaat,
                pv_row=pv_row,
                ws_sheet=ws_pp2,
                pauze_cols=pauze_cols_pp2,
                open_spots_set=pp2_open_spots
            )
    
            if not cols:
                continue
    
            pp2_place_short_break_cols_on_row(
                naam=kandidaat,
                pv=pv,
                pv_row=pv_row,
                cols=cols
            )
    
            return kandidaat, cols
    
        return None, []
    
    
    pp2_other_pending_short_breaks, pp2_endworkers_without_long_break = pp2_build_step5_pending_groups()
    
    pp2_step5_short_breaks_placed = []
    
    
    
    
    # -----------------------------------
    # 1B) Speciale behandeling:
    #     minderjarigen die <= 6u werken én pas beginnen na 13u
    #     => korte pauze zo laat mogelijk, bij voorkeur op rij van lange pauze
    # -----------------------------------
    pp2_late_start_minors_handled = set()
    
    for naam in list(pp2_other_pending_short_breaks):
        werk_uren = pp2_get_student_work_hours(naam)
        if not werk_uren:
            continue

        is_minor = pp2_is_minderjarig(naam)
        werkt_kort = werkduur_voor_pauze(naam) <= 6
        begint_laat = min(werk_uren) > 13

        if not (is_minor and werkt_kort and begint_laat):
            continue

        lange_pauze_rij = None
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    lange_pauze_rij = pv_row
                    break
            if lange_pauze_rij is not None:
                break

        gevonden = pp2_zoek_laatste_kort_kwartier_vanaf(
            naam=naam, ws_sheet=ws_pp2, min_col_exclusive=None,
            voorkeur_rij=lange_pauze_rij, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, open_spots_set=pp2_open_spots
        )
        if gevonden is None:
            continue

        pv, pv_row, col, conflict = gevonden
        pp2_place_short_break_cols_on_row(
            naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=conflict
        )
        pp2_late_start_minors_handled.add(naam)
    
    # Verwijder deze studenten uit de gewone pending lijst
    pp2_other_pending_short_breaks = [
        naam for naam in pp2_other_pending_short_breaks
        if naam not in pp2_late_start_minors_handled
    ]

    # -----------------------------------             
    # 1C) Speciale behandeling:
    #     minderjarigen >6u die tot het einduur werken
    #     => korte pauze zo laat mogelijk, na het tweede halfuur
    # -----------------------------------
    pp2_minor_long_end_handled = set()

    for naam in list(pp2_other_pending_short_breaks):
        if not pp2_is_minderjarig(naam):
            continue
        if werkduur_voor_pauze(naam) <= 6:
            continue
        if not pp2_student_works_until_day_end(naam):
            continue

        ankercol = None
        laatste_lange_rij = None
        for _pv, pv_row in pv_rows_pp2:
            for idx in range(len(pauze_cols_pp2) - 1):
                col1 = pauze_cols_pp2[idx]
                col2 = pauze_cols_pp2[idx + 1]
                if (
                    ws_pp2.cell(pv_row, col1).value == naam
                    and ws_pp2.cell(pv_row, col2).value == naam
                ):
                    if ankercol is None or col2 > ankercol:
                        ankercol = col2
                        laatste_lange_rij = pv_row

        gevonden = pp2_zoek_laatste_kort_kwartier_vanaf(
            naam=naam, ws_sheet=ws_pp2, min_col_exclusive=ankercol,
            voorkeur_rij=laatste_lange_rij, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, open_spots_set=pp2_open_spots
        )
        if gevonden is None:
            continue

        pv, pv_row, col, conflict = gevonden
        pp2_place_short_break_cols_on_row(
            naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=conflict
        )
        pp2_minor_long_end_handled.add(naam)

    # Verwijder uit de algemene pending lijst
    pp2_other_pending_short_breaks = [
        naam for naam in pp2_other_pending_short_breaks
        if naam not in pp2_minor_long_end_handled
    ]

  
    # -----------------------------------
    # 2) Eerst alle "gewone" resterende korte kwartieren invullen
    #    (rondes-systeem: eigen rij eerst, anders vroegste/moeilijkste PV,
    #    anders wachtrij, anders rood)
    # -----------------------------------
    pp2_wachtrij_stap5 = list(pp2_other_pending_short_breaks)

    while pp2_wachtrij_stap5:
        vooruitgang = False
        nog_niet_geplaatst = []

        for naam in pp2_wachtrij_stap5:
            resterend = pp2_resterende_korte_kwartieren(
                naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            )
            if resterend <= 0:
                continue

            eigen_rij = pp2_vind_rij_met_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2)
            geplaatst = False

            if eigen_rij is not None and pp2_student_has_long_break_in_row(naam, ws_pp2, eigen_rij, pauze_cols_pp2):
                eigen_pv = next((p for p, r in pv_rows_pp2 if r == eigen_rij), None)
                ankercol = pp2_lange_pauze_eindkolom_op_rij(naam, ws_pp2, eigen_rij, pauze_cols_pp2)
                cols = pp2_find_needed_short_cols_for_student_on_row(
                    naam=naam, pv_row=eigen_rij, ws_sheet=ws_pp2,
                    pauze_cols=pauze_cols_pp2, open_spots_set=pp2_open_spots,
                    min_col_exclusive=ankercol
                )
                if cols and eigen_pv is not None:
                    venster = pp2_tijdvenster_pauze(cols, ws_pp2)
                    attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                    if pp2_pv_kan_overname(eigen_pv, attrs):
                        pp2_place_short_break_cols_on_row(naam=naam, pv=eigen_pv, pv_row=eigen_rij, cols=cols)
                        geplaatst = True
                        vooruitgang = True

            if not geplaatst:
                opties = pp2_verzamel_opties_alle_pvs_kort(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots)
                for start_min, _schaarste, pv, pv_row, col in opties:
                    venster = pp2_tijdvenster_pauze([col], ws_pp2)
                    attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                    if pp2_pv_kan_overname(pv, attrs):
                        pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                        geplaatst = True
                        vooruitgang = True
                        break

            if not geplaatst:
                nog_niet_geplaatst.append(naam)

        pp2_wachtrij_stap5 = nog_niet_geplaatst
        if not vooruitgang:
            break

    for naam in pp2_wachtrij_stap5:
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
        if resterend <= 0:
            continue

        opties = pp2_verzamel_opties_alle_pvs_kort(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots)
        if not opties:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "stap 5 (overige resterende korte kwartieren): geen enkele vrije plek gevonden bij een PV"
            })
            continue

        geplaatst = False
        for start_min, _schaarste, pv, pv_row, col in opties:
            venster = pp2_tijdvenster_pauze([col], ws_pp2)
            attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
            if pp2_pv_kan_overname(pv, attrs):
                pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col])
                geplaatst = True
                break

        if geplaatst:
            continue

        _start_min, _schaarste, pv, pv_row, col = opties[0]
        pp2_place_short_break_cols_on_row(naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=True)
    
    
    # -----------------------------------
    # 3) Pas daarna: studenten die tot het einduur werken én geen lange
    #    pauze kregen. Zonder lange pauze is er geen 'eigen rij' om aan
    #    vast te plakken -- dus zelfde eenvoudige systeem als de
    #    korte-only werkers: vroegste PV, bij gelijkstand moeilijkste,
    #    wachtrij, en pas rood als het écht nergens gekwalificeerd lukt.
    #    (Volgorde blijft willekeurig, zoals al gebeurt in
    #    pp2_build_step5_pending_groups.)
    # -----------------------------------
    pp2_wachtrij_eindwerkers = list(pp2_endworkers_without_long_break)

    while pp2_wachtrij_eindwerkers:
        vooruitgang = False
        nog_niet_geplaatst = []

        for naam in pp2_wachtrij_eindwerkers:
            if pp2_resterende_korte_kwartieren(
                naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
                pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
            ) <= 0:
                continue

            opties = pp2_verzamel_opties_alle_pvs_kort(
                naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
            )
            if not opties:
                nog_niet_geplaatst.append(naam)
                continue

            geplaatst = False
            for start_min, _schaarste, pv, pv_row, col in opties:
                venster = pp2_tijdvenster_pauze([col], ws_pp2)
                attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
                if pp2_pv_kan_overname(pv, attrs):
                    pp2_place_short_break_cols_on_row(
                        naam=naam, pv=pv, pv_row=pv_row, cols=[col]
                    )
                    geplaatst = True
                    vooruitgang = True
                    break

            if not geplaatst:
                nog_niet_geplaatst.append(naam)

        pp2_wachtrij_eindwerkers = nog_niet_geplaatst
        if not vooruitgang:
            break

    for naam in pp2_wachtrij_eindwerkers:
        if pp2_resterende_korte_kwartieren(
            naam=naam, ws_sheet=ws_pp2, pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2, lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        ) <= 0:
            continue

        opties = pp2_verzamel_opties_alle_pvs_kort(
            naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2, pp2_open_spots
        )
        if not opties:
            pp2_niet_geplaatst.append({
                "naam": naam,
                "reden": "stap 5 (eindwerkers zonder lange pauze): geen enkele vrije plek gevonden bij een PV"
            })
            continue

        geplaatst = False
        for start_min, _schaarste, pv, pv_row, col in opties:
            venster = pp2_tijdvenster_pauze([col], ws_pp2)
            attrs = pp2_attracties_in_venster(naam, *venster) if venster else set()
            if pp2_pv_kan_overname(pv, attrs):
                pp2_place_short_break_cols_on_row(
                    naam=naam, pv=pv, pv_row=pv_row, cols=[col]
                )
                geplaatst = True
                break

        if geplaatst:
            continue

        _start_min, _schaarste, pv, pv_row, col = opties[0]
        pp2_place_short_break_cols_on_row(
            naam=naam, pv=pv, pv_row=pv_row, cols=[col], conflict=True
        )
    
    
    
    
    #FEEDBACKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKKK
    # =============================
    # FEEDBACK SHEET - OPTIE 2
    # =============================
    ws_feedback2 = wb_arg.create_sheet("Feedback PP")
    
    groen_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    rood_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    row_fb2 = 1
    
    ws_feedback2.cell(row_fb2, 1, "Feedback pauzeplanning").font = Font(bold=True)
    row_fb2 += 2

    
    # -----------------------------------
    # 1) Lange pauzes controleren
    # Nieuwe logica PP optie 2:
    # - alleen studenten met >6 uur werk moeten een lange pauze hebben
    # -----------------------------------
    pp2_lange_pauze_ontbreekt = []
    
    for s in studenten:
        naam = s["naam"]
        gewerkte_uren = werkduur_voor_pauze(naam)
    
        if gewerkte_uren > 6:
            if not pp2_heeft_al_lange_pauze(naam, ws_pp2, pv_rows_pp2, pauze_cols_pp2):
                pp2_lange_pauze_ontbreekt.append(naam)
    
    if not pp2_lange_pauze_ontbreekt:
        cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle lange pauzes toegekend")
        cel.fill = groen_fill
        cel.font = Font(bold=True, color="006100")
        row_fb2 += 2
    else:
        cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende lange pauzes:")
        cel.fill = rood_fill
        cel.font = Font(bold=True)
        row_fb2 += 1
    
        for naam in sorted(pp2_lange_pauze_ontbreekt):
            ws_feedback2.cell(row_fb2, 1, naam)
            row_fb2 += 1
    
        row_fb2 += 1
    
    # -----------------------------------
    # 2) Korte kwartieren controleren
    # Gebruik exact dezelfde logica als de planner zelf:
    # - pp2_benodigde_korte_kwartieren(...)
    # - pp2_resterende_korte_kwartieren(...)
    # Dus geen aparte feedbacktelling meer
    # -----------------------------------
    pp2_korte_kwartieren_ontbreekt = []
    
    for s in studenten:
        naam = s["naam"]
    
        nodig = pp2_benodigde_korte_kwartieren(naam)
        if nodig <= 0:
            continue
    
        resterend = pp2_resterende_korte_kwartieren(
            naam=naam,
            ws_sheet=ws_pp2,
            pv_rows=pv_rows_pp2,
            pauze_cols=pauze_cols_pp2,
            lange_pauze_ontvangers=pp2_lange_pauze_ontvangers
        )
    
        if resterend > 0:
            pp2_korte_kwartieren_ontbreekt.append((naam, resterend))
    
    if not pp2_korte_kwartieren_ontbreekt:
        cel = ws_feedback2.cell(row_fb2, 1, "✓ Alle korte kwartieren toegekend")
        cel.fill = groen_fill
        cel.font = Font(bold=True, color="006100")
        row_fb2 += 2
    else:
        cel = ws_feedback2.cell(row_fb2, 1, "✗ Ontbrekende korte kwartieren:")
        cel.fill = rood_fill
        cel.font = Font(bold=True)
        row_fb2 += 1
    
        for naam, resterend in sorted(pp2_korte_kwartieren_ontbreekt, key=lambda x: x[0].lower()):
            if resterend == 1:
                ws_feedback2.cell(row_fb2, 1, f"{naam} - nog 1 kwartier tekort")
            else:
                ws_feedback2.cell(row_fb2, 1, f"{naam} - nog {resterend} kwartieren tekort")
            row_fb2 += 1
    
        row_fb2 += 1

    # -----------------------------------
    # Korte samenvatting + Aanpassingen-info, op de Pauzeplanning-sheet
    # zelf, onder de laatste pauzevlinderrij.
    # -----------------------------------
    if pv_rows_pp2:
        _laatste_pv_naam_rij = max(row for _, row in pv_rows_pp2)

        _samenvatting_lijnen = []
        if pp2_lange_pauze_ontbreekt:
            _namen_lang = sorted(pp2_lange_pauze_ontbreekt)
            _samenvatting_lijnen.append(
                f"{', '.join(_namen_lang)} kregen geen lange pauze"
                if len(_namen_lang) > 1
                else f"{_namen_lang[0]} kreeg geen lange pauze"
            )
        if pp2_korte_kwartieren_ontbreekt:
            _namen_kort = sorted(n for n, _ in pp2_korte_kwartieren_ontbreekt)
            _samenvatting_lijnen.append(
                f"{', '.join(_namen_kort)} kregen geen korte pauze"
                if len(_namen_kort) > 1
                else f"{_namen_kort[0]} kreeg geen korte pauze"
            )

        _start_rij_samenvatting = _laatste_pv_naam_rij + 2
        for i, _lijn in enumerate(_samenvatting_lijnen):
            _cel = ws_pp2.cell(row=_start_rij_samenvatting + i, column=1, value=_lijn)
            _cel.font = Font(bold=True)
            _cel.border = thin_border
            _cel.alignment = Alignment(horizontal="left", vertical="center")

        if _samenvatting_lijnen:
            _start_rij_aanpassingen = _start_rij_samenvatting + len(_samenvatting_lijnen) + 1
        else:
            _start_rij_aanpassingen = _laatste_pv_naam_rij + 3

        _vinkje_aanpassingen_lokaal = ws_aanpassingen.cell(row=3, column=22).value  # V3
        if _vinkje_aanpassingen_lokaal in [1, True, "WAAR", "X"]:
            _waarde_aanpassingen = ws_aanpassingen.cell(row=3, column=23).value  # W3
            if _waarde_aanpassingen:
                for i, _regel in enumerate(str(_waarde_aanpassingen).split("\n")):
                    _regel = _regel.strip()
                    if _regel:
                        _cel = ws_pp2.cell(row=_start_rij_aanpassingen + i, column=1, value=_regel)
                        _cel.fill = witte_fill
                        _cel.border = thin_border
                        _cel.alignment = Alignment(horizontal="left", vertical="center")

    pp2_rode_pauze_redenen = pp2_verzamel_rode_pauze_redenen(ws_pp2, pv_rows_pp2, pauze_cols_pp2)
    
    if pp2_rode_pauze_redenen:
        _titel_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        _kader_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
        _dun_rand = Border(
            left=Side(style="thin", color="F8CBAD"),
            right=Side(style="thin", color="F8CBAD"),
            top=Side(style="thin", color="F8CBAD"),
            bottom=Side(style="thin", color="F8CBAD"),
        )
        _laatste_kolom = 15  # niet verder dan kolom O

        ws_feedback2.cell(row_fb2, 1, "Rode pauzes: waarom?").font = Font(bold=True, size=13)
        row_fb2 += 2

        for naam, tekst in pp2_rode_pauze_redenen:
            ws_feedback2.merge_cells(start_row=row_fb2, start_column=1, end_row=row_fb2, end_column=_laatste_kolom)
            titel_cel = ws_feedback2.cell(row_fb2, 1, f"⚠  {naam}")
            titel_cel.font = Font(bold=True, color="9C0006")
            titel_cel.alignment = Alignment(vertical="center")
            for c in range(1, _laatste_kolom + 1):
                ws_feedback2.cell(row_fb2, c).fill = _titel_fill
                ws_feedback2.cell(row_fb2, c).border = _dun_rand
            row_fb2 += 1

            ws_feedback2.merge_cells(start_row=row_fb2, start_column=1, end_row=row_fb2, end_column=_laatste_kolom)
            tekst_cel = ws_feedback2.cell(row_fb2, 1, tekst)
            tekst_cel.alignment = Alignment(wrap_text=True, vertical="top")
            for c in range(1, _laatste_kolom + 1):
                ws_feedback2.cell(row_fb2, c).fill = _kader_fill
                ws_feedback2.cell(row_fb2, c).border = _dun_rand
            _geschatte_lijnen = max(1, (len(tekst) // 110) + 1)
            ws_feedback2.row_dimensions[row_fb2].height = max(15, _geschatte_lijnen * 15)
            row_fb2 += 1

            row_fb2 += 1  # lege regel tussen kaartjes

        row_fb2 += 1

    # -----------------------------------
    # kolombreedte en opmaak
    # -----------------------------------
    ws_feedback2.column_dimensions["A"].width = 45
    
    for row in ws_feedback2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

# Globals herstellen
    ws_planning     = _ws_planning_bak
    student_totalen = _student_totalen_bak


# ── Oorspronkelijke aanroep ──
def pp2_tel_rode_cellen_extern(wb, conflict_rgb="00FFC7CE"):
    """Telt het aantal rood-gemarkeerde (conflict) cellen in Pauzeplanning."""
    ws = wb["Pauzeplanning"]
    aantal = 0
    for row in ws.iter_rows():
        for cel in row:
            if cel.fill and cel.fill.start_color and cel.fill.start_color.rgb == conflict_rgb and cel.value:
                aantal += 1
    return aantal


PP2_MAX_POGINGEN = 5
_pp2_beste_aantal_rood = None

for _pp2_poging in range(PP2_MAX_POGINGEN):
    maak_pp2_sheets(wb_out, assigned_map)
    _pp2_aantal_rood = pp2_tel_rode_cellen_extern(wb_out)

    if _pp2_beste_aantal_rood is None or _pp2_aantal_rood < _pp2_beste_aantal_rood:
        _pp2_beste_aantal_rood = _pp2_aantal_rood

        for _naam in ["Pauzeplanning_beste", "Feedback PP_beste"]:
            if _naam in wb_out.sheetnames:
                wb_out.remove(wb_out[_naam])

        _pp2_kopie_pp = wb_out.copy_worksheet(wb_out["Pauzeplanning"])
        _pp2_kopie_pp.title = "Pauzeplanning_beste"
        _pp2_kopie_fb = wb_out.copy_worksheet(wb_out["Feedback PP"])
        _pp2_kopie_fb.title = "Feedback PP_beste"

    if _pp2_beste_aantal_rood == 0:
        break

# De beste poging terugzetten onder de juiste, definitieve naam
wb_out.remove(wb_out["Pauzeplanning"])
wb_out.remove(wb_out["Feedback PP"])

_pp2_beste_pp = wb_out["Pauzeplanning_beste"]
_pp2_beste_pp.title = "Pauzeplanning"
_pp2_beste_fb = wb_out["Feedback PP_beste"]
_pp2_beste_fb.title = "Feedback PP"


# PART 6 6666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666
# PART 6 666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666666

# -----------------------------
# DEEL 6: Wissels detecteren, classificeren en exporteren
# -----------------------------

from collections import defaultdict, deque
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# -----------------------------
# Helpers
# -----------------------------
def build_student_per_hour_map(assigned_map):
    student_per_uur = defaultdict(dict)
    for (uur, attr), namen in assigned_map.items():
        for naam in namen:
            student_per_uur[naam][uur] = attr
    return student_per_uur


def extract_hourly_changes(student_per_uur, open_uren):
    """
    Bouw per uur alle veranderingen op:
    - newcomers: studenten die op dit uur starten
    - movers: studenten die op dit uur van attractie wisselen
    - leavers: studenten die vorig uur wel werkten en nu niet meer
    - disappearing_sources: attractieplekken die verdwijnen tussen vorig uur en dit uur
    """
    changes_per_hour = {}

    def capaciteit_op_uur(uur, attr):
        if uur not in open_uren:
            return 0
        return max(0, aantallen.get(uur, {}).get(attr, 0))

    all_attrs = set()
    for uur2 in open_uren:
        all_attrs.update(aantallen.get(uur2, {}).keys())

    for uur in sorted(open_uren):
        # VERVANGEN DOOR:
        _sorted_open = sorted(open_uren)
        _idx = _sorted_open.index(uur)
        prev_uur = _sorted_open[_idx - 1] if _idx > 0 else None

        prev_students = {}
        curr_students = {}

        for naam, uren_dict in student_per_uur.items():
            if prev_uur in uren_dict:
                prev_students[naam] = uren_dict[prev_uur]
            if uur in uren_dict:
                curr_students[naam] = uren_dict[uur]

        newcomers = []
        movers = []
        leavers = []

        for naam, curr_attr in curr_students.items():
            if naam not in prev_students:
                newcomers.append({
                    "naam": naam,
                    "naar": curr_attr
                })
            else:
                prev_attr = prev_students[naam]
                if prev_attr != curr_attr:
                    movers.append({
                        "naam": naam,
                        "van": prev_attr,
                        "naar": curr_attr,
                        "uur": uur,
                        "type": "normaal"
                    })

        for naam, prev_attr in prev_students.items():
            if naam not in curr_students:
                leavers.append({
                    "naam": naam,
                    "van": prev_attr
                })

        disappearing_sources = []
        if prev_uur is not None and prev_uur in open_uren:
            for attr in sorted(all_attrs):
                prev_cap = capaciteit_op_uur(prev_uur, attr)
                curr_cap = capaciteit_op_uur(uur, attr)

                if curr_cap < prev_cap:
                    for pos in range(curr_cap + 1, prev_cap + 1):
                        disappearing_sources.append({
                            "attr": attr,
                            "pos": pos,
                            "reason": "capacity_drop"
                        })

        changes_per_hour[uur] = {
            "newcomers": newcomers,
            "movers": movers,
            "leavers": leavers,
            "disappearing_sources": disappearing_sources
        }

    return changes_per_hour




def classify_hourly_switches(uur, newcomers, movers, leavers=None, disappearing_sources=None):
    """
    Types:
    - volledig automatisch:
        een nieuwkomer komt toe op attractie A,
        daardoor kan een student van A weg,
        waardoor ketting verder loopt
    - half-automatisch:
        een ketting die start vanuit een verdwijnende plek
        of een logisch vervolg daarop is
    - normaal:
        losse wissels of resterende lussen zonder duidelijk startpunt

    Belangrijk:
    - de eerste edge van een ketting krijgt 'half-start'
    - de rest krijgt 'half-automatisch'
    - losse enkele wissels blijven 'normaal'

    Extra regels:
    - bij echte ronde lussen kiezen we het startpunt liefst op een attractie
      met 2 plekken op dit uur
    - niet-ronde kettingen komen vóór ronde lussen in de output
    - groene wissels starten altijd bij de attractie waar de nieuwkomer toekomt
    - enkel kettingen met lengte > 1 komen in de half-automatische output
      zodat er geen dubbels ontstaan
    """
    if not movers:
        return []

    if leavers is None:
        leavers = []

    if disappearing_sources is None:
        disappearing_sources = []

    # -----------------------------
    # Helpers
    # -----------------------------
    def stable_edge_key(edge):
        return (edge["van"], edge["naar"], edge["naam"])

    def next_edge_key(edge):
        return (edge["naar"], edge["naam"])

    def has_two_spots(attr):
        try:
            return aantallen[uur].get(attr, 1) >= 2
        except Exception:
            return False

    def roll_chain_from_start_edge(start_edge, edge_pool, used_ids):
        chain = []
        current = start_edge

        while current and current["id"] not in used_ids:
            chain.append(current)
            used_ids.add(current["id"])

            next_candidates = [
                e for e in edge_pool
                if e["id"] not in used_ids and e["van"] == current["naar"]
            ]
            next_candidates.sort(key=next_edge_key)
            current = next_candidates[0] if next_candidates else None

        return chain

    def classify_chain_shape(chain):
        """
        Geeft terug:
        - 'open' als begin en einde verschillen
        - 'cycle' als begin en einde terug sluiten
        """
        if len(chain) <= 1:
            return "single"

        eerste_van = chain[0]["van"]
        laatste_naar = chain[-1]["naar"]

        if eerste_van == laatste_naar:
            return "cycle"
        return "open"

    def add_chain_record_if_needed(chain_records, chain):
        """
        Enkel echte kettingen (lengte > 1) komen in chain_records.
        Singles blijven 'normaal' en worden later via normal_edges getoond.
        """
        if not chain:
            return

        shape = classify_chain_shape(chain)

        if len(chain) == 1:
            chain[0]["type"] = "normaal"
            return

        chain[0]["type"] = "half-start"
        for e in chain[1:]:
            e["type"] = "half-automatisch"

        chain_records.append({
            "shape": shape,
            "start_has_two_spots": has_two_spots(chain[0]["van"]),
            "edges": chain
        })

    # -----------------------------
    # Edges opbouwen
    # -----------------------------
    edges = []
    for idx, m in enumerate(movers):
        edges.append({
            "id": idx,
            "naam": m["naam"],
            "van": m["van"],
            "naar": m["naar"],
            "uur": uur,
            "type": "normaal"
        })

    # -----------------------------
    # Maps
    # -----------------------------
    outgoing = defaultdict(list)
    incoming = defaultdict(list)

    for e in edges:
        outgoing[e["van"]].append(e)
        incoming[e["naar"]].append(e)

    for attr in outgoing:
        outgoing[attr].sort(key=next_edge_key)
    for attr in incoming:
        incoming[attr].sort(key=lambda x: (x["van"], x["naam"]))

    # -----------------------------
    # 1. Volledig automatische kettingen
    # -----------------------------
    newcomers_by_attr = defaultdict(list)
    for n in newcomers:
        newcomers_by_attr[n["naar"]].append(n["naam"])

    auto_edge_ids = set()
    queue = deque()

    # Groen start ALTIJD bij de attractie waar de nieuwkomer toekomt
    # De nieuwkomer zet daar de ketting in gang.
    for attr in newcomers_by_attr.keys():
        for e in outgoing.get(attr, []):
            if e["id"] not in auto_edge_ids:
                auto_edge_ids.add(e["id"])
                queue.append(e)

    while queue:
        current = queue.popleft()
        next_attr = current["naar"]

        for next_edge in outgoing.get(next_attr, []):
            if next_edge["id"] not in auto_edge_ids:
                auto_edge_ids.add(next_edge["id"])
                queue.append(next_edge)

    for e in edges:
        if e["id"] in auto_edge_ids:
            e["type"] = "volledig automatisch"

    # -----------------------------
    # 2. Resterende edges
    # -----------------------------
    remaining_edges = [e for e in edges if e["id"] not in auto_edge_ids]

    if not remaining_edges:
        auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]

        ordered_auto = []
        used_auto = set()

        # Volg exact de volgorde van newcomers, niet alfabetisch op attractie
        for newcomer in newcomers:
            start_attr = newcomer["naar"]

            start_candidates = [
                e for e in auto_edges
                if e["id"] not in used_auto and e["van"] == start_attr
            ]
            start_candidates.sort(key=next_edge_key)

            for start in start_candidates:
                current = start
                while current and current["id"] not in used_auto:
                    ordered_auto.append(current)
                    used_auto.add(current["id"])

                    next_candidates = [
                        e for e in auto_edges
                        if e["id"] not in used_auto and e["van"] == current["naar"]
                    ]
                    next_candidates.sort(key=next_edge_key)
                    current = next_candidates[0] if next_candidates else None

        leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
        leftovers_auto.sort(key=stable_edge_key)
        ordered_auto.extend(leftovers_auto)

        return ordered_auto

    source_attrs = [x["attr"] for x in disappearing_sources]

    chain_records = []
    used_ids = set()

    # -----------------------------
    # 3. Eerst kettingen vanuit verdwijnende plekken
    # -----------------------------
    for start_attr in source_attrs:
        start_candidates = [
            e for e in remaining_edges
            if e["id"] not in used_ids and e["van"] == start_attr
        ]
        start_candidates.sort(key=stable_edge_key)

        for start_edge in start_candidates:
            if start_edge["id"] in used_ids:
                continue

            chain = roll_chain_from_start_edge(start_edge, remaining_edges, used_ids)
            add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 4. Restjes groeperen in componenten
    # -----------------------------
    leftovers = [e for e in remaining_edges if e["id"] not in used_ids]

    if leftovers:
        remaining_by_id = {e["id"]: e for e in leftovers}
        adjacency = defaultdict(set)

        for e1 in leftovers:
            for e2 in leftovers:
                if e1["id"] == e2["id"]:
                    continue
                if e1["naar"] == e2["van"] or e2["naar"] == e1["van"]:
                    adjacency[e1["id"]].add(e2["id"])
                    adjacency[e2["id"]].add(e1["id"])

        visited = set()
        components = []

        for e in leftovers:
            if e["id"] in visited:
                continue

            stack = [e["id"]]
            comp_ids = []

            while stack:
                curr = stack.pop()
                if curr in visited:
                    continue
                visited.add(curr)
                comp_ids.append(curr)

                for nb in adjacency[curr]:
                    if nb not in visited:
                        stack.append(nb)

            components.append([remaining_by_id[i] for i in comp_ids])

        for comp_edges in components:
            if not comp_edges:
                continue

            comp_used = set()

            start_candidates = []
            for e in comp_edges:
                has_prev = any(
                    other["id"] != e["id"] and other["naar"] == e["van"]
                    for other in comp_edges
                )
                if not has_prev:
                    start_candidates.append(e)

            # ---------------------------------
            # NIET-RONDE KETTINGEN
            # ---------------------------------
            if start_candidates:
                start_candidates.sort(key=stable_edge_key)

                for start_edge in start_candidates:
                    if start_edge["id"] in comp_used:
                        continue

                    chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)

                for edge in rest:
                    if edge["id"] in comp_used:
                        continue
                    chain = roll_chain_from_start_edge(edge, comp_edges, comp_used)
                    add_chain_record_if_needed(chain_records, chain)

            # ---------------------------------
            # ECHTE RONDE LUS
            # ---------------------------------
            else:
                two_spot_candidates = [e for e in comp_edges if has_two_spots(e["van"])]

                if two_spot_candidates:
                    two_spot_candidates.sort(key=stable_edge_key)
                    start_edge = two_spot_candidates[0]
                else:
                    comp_edges.sort(key=stable_edge_key)
                    start_edge = comp_edges[0]

                chain = roll_chain_from_start_edge(start_edge, comp_edges, comp_used)

                rest = [e for e in comp_edges if e["id"] not in comp_used]
                rest.sort(key=stable_edge_key)
                chain.extend(rest)

                add_chain_record_if_needed(chain_records, chain)

    # -----------------------------
    # 5. Definitieve volgorde
    # -----------------------------
    auto_edges = [e for e in edges if e["type"] == "volledig automatisch"]
    normal_edges = [e for e in edges if e["type"] == "normaal"]

    ordered_auto = []
    used_auto = set()

    # Groen start ALTIJD vanuit de attractie van de nieuwkomer
    # en volgt dan pas de ketting verder.
    for newcomer in newcomers:
        start_attr = newcomer["naar"]

        start_candidates = [
            e for e in auto_edges
            if e["id"] not in used_auto and e["van"] == start_attr
        ]
        start_candidates.sort(key=next_edge_key)

        for start in start_candidates:
            current = start
            while current and current["id"] not in used_auto:
                ordered_auto.append(current)
                used_auto.add(current["id"])

                next_candidates = [
                    e for e in auto_edges
                    if e["id"] not in used_auto and e["van"] == current["naar"]
                ]
                next_candidates.sort(key=next_edge_key)
                current = next_candidates[0] if next_candidates else None

    leftovers_auto = [e for e in auto_edges if e["id"] not in used_auto]
    leftovers_auto.sort(key=stable_edge_key)
    ordered_auto.extend(leftovers_auto)

    # niet-ronde kettingen eerst, dan ronde lussen
    chain_records.sort(
        key=lambda rec: (
            0 if rec["shape"] == "open" else 1,
            0 if rec["shape"] == "cycle" and rec["start_has_two_spots"] else 1,
            stable_edge_key(rec["edges"][0]) if rec["edges"] else ("", "", "")
        )
    )

    ordered_half = []
    for rec in chain_records:
        ordered_half.extend(rec["edges"])

    normal_edges.sort(key=stable_edge_key)

    # Extra veiligheid tegen dubbels
    seen_ids = set()
    final_order = []

    for e in ordered_auto + ordered_half + normal_edges:
        if e["id"] in seen_ids:
            continue
        seen_ids.add(e["id"])
        final_order.append(e)

    return final_order



# ─────────────────────────────────────────────────────────────────
# DEEL 6 uitvoer als herbruikbare functie
# ─────────────────────────────────────────────────────────────────
def maak_wisselplanning_sheet(wb_arg, am_arg):
    """
    Bouw het 'Wissels'-werkblad op basis van am_arg (assigned_map).
    Vervangt het bestaande sheet als het al bestaat.
    """
    # Verwijder oud sheet indien aanwezig
    if "Wissels" in wb_arg.sheetnames:
        del wb_arg["Wissels"]

    # Stap 1: student → uur → attractie
    student_per_uur = build_student_per_hour_map(am_arg)

    # Stap 2: veranderingen per uur opbouwen
    changes_per_hour = extract_hourly_changes(student_per_uur, open_uren)

    # Stap 3: per uur classificeren en ordenen
    wissels_per_uur = {}
    for uur in sorted(open_uren):
        newcomers        = changes_per_hour[uur]["newcomers"]
        movers           = changes_per_hour[uur]["movers"]
        leavers          = changes_per_hour[uur]["leavers"]
        disappearing_sources = changes_per_hour[uur]["disappearing_sources"]

        ordered_switches = classify_hourly_switches(
            uur, newcomers, movers, leavers, disappearing_sources
        )
        if ordered_switches:
            wissels_per_uur[uur] = ordered_switches

    # KPI berekenen
    totaal_wissels = 0
    aantal_auto    = 0
    for uur in wissels_per_uur:
        for w in wissels_per_uur[uur]:
            totaal_wissels += 1
            if w["type"] == "volledig automatisch":
                aantal_auto += 1
    niet_groen = totaal_wissels - aantal_auto

    # Stap 4: werkblad aanmaken
    ws_wissels = wb_arg.create_sheet(title="Wissels")

    # KPI rechts van de tabel (kolom G)
    ws_wissels.cell(1, 7, "KPI Wissels").font = Font(bold=True)
    ws_wissels.cell(2, 7, "Totaal wissels:")
    ws_wissels.cell(2, 8, totaal_wissels)
    ws_wissels.cell(3, 7, "Volledig automatisch:")
    ws_wissels.cell(3, 8, aantal_auto)
    ws_wissels.cell(4, 7, "Niet-groen (KPI):")
    ws_wissels.cell(4, 8, niet_groen)
    ws_wissels.cell(4, 8).font = Font(bold=True)

    _center = Alignment(horizontal="center", vertical="center")
    _border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    current_row = 1
    for uur in sorted(wissels_per_uur.keys()):
        # Titelrij per uur
        title_cell = ws_wissels.cell(current_row, 1, f"Wissels om {formatteer_uur(uur)}")
        title_cell.font      = Font(bold=True)
        title_cell.alignment = _center
        current_row += 1

        # Headers
        for col_idx, header in enumerate(["Student", "Van", "Naar"], start=1):
            cell           = ws_wissels.cell(current_row, col_idx, header)
            cell.font      = Font(bold=True)
            cell.alignment = _center
            cell.border    = _border
        current_row += 1

        # Wissels
        for w in wissels_per_uur[uur]:
            ws_wissels.cell(current_row, 1, w["naam"])
            ws_wissels.cell(current_row, 2, w["van"])
            ws_wissels.cell(current_row, 3, w["naar"])

            for col_idx in range(1, 4):
                cell           = ws_wissels.cell(current_row, col_idx)
                cell.alignment = _center
                cell.border    = _border

            if w["type"] == "volledig automatisch":
                ws_wissels.cell(current_row, 2).fill = green_fill
                ws_wissels.cell(current_row, 3).fill = green_fill
            elif w["type"] == "half-automatisch":
                ws_wissels.cell(current_row, 2).fill = yellow_fill
                ws_wissels.cell(current_row, 3).fill = yellow_fill
            elif w["type"] == "half-start":
                ws_wissels.cell(current_row, 2).fill = orange_fill
                ws_wissels.cell(current_row, 3).fill = orange_fill

            current_row += 1
        current_row += 1  # lege rij tussen uren

    # Stap 5: kolombreedtes
    for col_idx, breedte in {1: 22, 2: 25, 3: 25, 7: 24, 8: 18}.items():
        ws_wissels.column_dimensions[get_column_letter(col_idx)].width = breedte


# ── oorspronkelijke aanroep (vervangt de oude losse code) ──
maak_wisselplanning_sheet(wb_out, assigned_map)

# -----------------------------
# Werkblad Heropleidingen
# -----------------------------
from openpyxl.styles.proxy import StyleProxy
from copy import copy

ws_bron = wb.worksheets[[ws.title for ws in wb.worksheets].index("Heropleidingen")] if "Heropleidingen" in wb.sheetnames else None
if ws_bron:
    ws_hero = wb_out.create_sheet(title="Heropleidingen")
    for rij in ws_bron.iter_rows():
        for cel in rij:
            nieuwe_cel = ws_hero.cell(row=cel.row, column=cel.column, value=cel.value)
            if cel.has_style:
                nieuwe_cel.font = copy(cel.font)
                nieuwe_cel.fill = copy(cel.fill)
                nieuwe_cel.border = copy(cel.border)
                nieuwe_cel.alignment = copy(cel.alignment)
                nieuwe_cel.number_format = cel.number_format
    for kol, breedte in ws_bron.column_dimensions.items():
        ws_hero.column_dimensions[kol].width = breedte.width
        ws_hero.column_dimensions["A"].width = 11
    for rij, hoogte in ws_bron.row_dimensions.items():
        ws_hero.row_dimensions[rij].height = hoogte.height



#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW
#NIEUWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWWW

# -----------------------------
# Werkbladen altijd verbergen
# -----------------------------
for bladnaam in ["Pauzevlinders", "Feedback"]:
    if bladnaam in wb_out.sheetnames:
        ws_hide = wb_out[bladnaam]
        ws_hide.sheet_state = "veryHidden" 

# Snapshot voor last-minute (zonder dringende heropleidingen)
if "lm_base_bytes" not in st.session_state:
    _buf = BytesIO()
    wb_out.save(_buf)
    st.session_state["lm_base_bytes"] = _buf.getvalue()

output_lm_base = BytesIO(st.session_state["lm_base_bytes"])

# -----------------------------
# Dringende heropleidingen in Planning
# -----------------------------
ws_plan = wb_out["Planning"]
laatste_rij = ws_plan.max_row
invoegrij = laatste_rij + 2

for rij in ws_hero.iter_rows():
    if rij[0].value == "Belangrijk!":
        naam = rij[1].value or ""
        omschrijving = rij[2].value or ""
        ws_plan.cell(invoegrij, 1).value = f"Dringende heropleiding: {naam}: {omschrijving}"
        invoegrij += 1



#ooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooooo


# -----------------------------
# Opslaan in hetzelfde unieke bestand als DEEL 3
# -----------------------------
output = BytesIO()
wb_out.save(output)
output.seek(0)
# st.success("Planning gegenereerd!")
st.download_button(
    "Download planning",
    data=output.getvalue(),
    file_name=f"Planning_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)





#DEELLL 8 OFZOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOO
# ============================================================
# LAST-MINUTE AFWEZIGEN V5
# VOLLEDIGE VERVANGING van alle vorige last-minute patches
# Plakken ONDER de bestaande st.download_button("Download planning", ...)
# ============================================================

import copy
import random
from collections import defaultdict, Counter
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------------------------------------
# Basishelpers
# ------------------------------------------------------------
def lm5_split_display_label(label):
    if not label:
        return "", 1
    s = str(label).strip()
    parts = s.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0].strip(), int(parts[1])
    return s, 1

def lm5_is_pv_row(label):
    return str(label).strip().lower().startswith("pauzevlinder")

def lm5_is_extra_row(label):
    return str(label).strip().lower().startswith("extra")

def lm5_parse_output_hour(header):
    if not header:
        return None
    s = str(header).strip().lower()
    if s == "9u30-11u":
        return min(open_uren) if open_uren else 10
    if s == "18u-19u30":
        return max(open_uren) if open_uren else 18
    # Strip label-gedeelte tussen haakjes, bv. "9u30 (0,5h)" → "9u30"
    kern = s.split("(")[0].strip()
    if "u" in kern:
        delen = kern.split("u")
        try:
            uren = int(delen[0])
            min_str = delen[1].strip() if len(delen) > 1 else ""
            minuten = int(min_str) if min_str.isdigit() else 0
            return uren + minuten / 60
        except:
            return None
    if ":" in kern:
        delen = kern.split(":")
        try:
            return int(delen[0]) + int(delen[1]) / 60
        except:
            return None
    try:
        return float(kern)
    except:
        return None
        

def lm5_student_lookup():
    return {str(s["naam"]).strip(): s for s in studenten}

def lm5_get_student(naam):
    return lm5_student_lookup().get(str(naam).strip())

def lm5_copy_student_state(student):
    return {
        "naam": str(student["naam"]).strip(),
        "uren_beschikbaar": list(student["uren_beschikbaar"]),
        "attracties": list(student["attracties"]),
        "aantal_attracties": student["aantal_attracties"],
        "is_pauzevlinder": student["is_pauzevlinder"],
        "pv_number": student["pv_number"],
        "assigned_attracties": set(),
        "assigned_hours": []
    }

def lm5_student_can_attr(student, attr):
    if not student:
        return False
    return student_kan_attr(student, attr)

def lm5_get_hours_on_attr(ctx, student_name, attr):
    uren = []
    for (uur, a), namen in ctx["assigned_map"].items():
        if a == attr and student_name in namen:
            uren.append(uur)
    return sorted(set(uren))

def lm5_count_switches(ctx, student_name):
    uur_attr = []
    for uur in sorted(set(ctx["student_states"][student_name]["assigned_hours"])):
        attr = lm5_student_current_attr_on_hour(ctx, student_name, uur)
        if attr:
            uur_attr.append((uur, attr))
    if not uur_attr:
        return 0
    switches = 0
    prev = uur_attr[0][1]
    for _, attr in uur_attr[1:]:
        if attr != prev:
            switches += 1
        prev = attr
    return switches

# ------------------------------------------------------------
# Input + originele planning lezen
# ------------------------------------------------------------
def lm5_extract_base_maps(base_bytes):
    wb_tmp = load_workbook(BytesIO(base_bytes))
    ws_plan = wb_tmp["Planning"]

    uur_to_col = {}
    for col in range(2, ws_plan.max_column + 1):
        uur = lm5_parse_output_hour(ws_plan.cell(1, col).value)
        if uur in open_uren and uur not in uur_to_col:
            uur_to_col[uur] = col

    if not uur_to_col:
        raise ValueError("Geen geldige uurkolommen gevonden in blad 'Planning'.")

    attr_rows = []
    pv_rows = []
    extra_rows = []
    student_hour_attr = {}
    student_hour_row = defaultdict(str)
    row_labels = []

    for row in range(2, ws_plan.max_row + 1):
        label = ws_plan.cell(row, 1).value
        if not label:
            continue
        label = str(label).strip()
        row_labels.append((row, label))

        if lm5_is_pv_row(label):
            pv_rows.append((row, label))
        elif lm5_is_extra_row(label):
            extra_rows.append((row, label))
        else:
            attr_rows.append((row, label))

        for uur, col in uur_to_col.items():
            naam = ws_plan.cell(row, col).value
            if naam and str(naam).strip():
                naam = str(naam).strip()
                attr, _pos = lm5_split_display_label(label)
                student_hour_attr[(naam, uur)] = attr
                student_hour_row[(naam, uur)] = label

    return {
        "wb": wb_tmp,
        "ws_plan": ws_plan,
        "uur_to_col": uur_to_col,
        "attr_rows": attr_rows,
        "pv_rows": pv_rows,
        "extra_rows": extra_rows,
        "student_hour_attr": student_hour_attr,
        "student_hour_row": student_hour_row,
        "row_labels": row_labels,
    }

def lm5_working_students_today(base_maps):
    out = set()
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if uur in open_uren:
            out.add(naam)
    return sorted(out)

def lm5_present_students_on_hour(base_maps, uur, absentees_set):
    out = []
    seen = set()
    for (naam, uur2), _attr in base_maps["student_hour_attr"].items():
        if uur2 == uur and naam not in absentees_set and naam not in seen:
            out.append(naam)
            seen.add(naam)
    return out

def lm5_pv_names():
    return [str(pv["naam"]).strip() for pv in selected]

def lm5_bereken_pauze_counts(absentees_set, base_maps):
    """
    Bereken lange_pauzes en korte_pauzes voor de last-minute planning
    op basis van werkelijke werkuren (excl. afwezigen).

    lange_pauzes = (studenten > 6u) + (minderjarige studenten >= 4u)
    korte_pauzes = studenten >= 4u
    """
    uren_per_student = defaultdict(set)
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if naam not in absentees_set:
            uren_per_student[naam].add(uur)

    lange_pauzes = 0
    korte_pauzes = 0
    for naam, uren in uren_per_student.items():
        n = len(uren)
        is_minor = "-18" in str(naam)
        if n > 6:
            lange_pauzes += 1
        if is_minor and ((n > 4) if PAUZE_STRIKT_BOVEN_4U else (n >= 4)):
            lange_pauzes += 1
        if (n > 4) if PAUZE_STRIKT_BOVEN_4U else (n >= 4):
            korte_pauzes += 1

    return lange_pauzes, korte_pauzes


def lm5_extract_capacity_actions():
    result = []

    # L3:M12 in Aanpassingen
    for rij in range(3, 13):
        left_source = ws_aanpassingen.cell(rij, 12).value   # kolom L
        right_source = ws_aanpassingen.cell(rij, 13).value  # kolom M

        left = str(left_source).strip() if left_source is not None and str(left_source).strip() != "" else ""
        right = str(right_source).strip() if right_source is not None and str(right_source).strip() != "" else ""

        if not left and not right:
            continue

        if left and not right:
            result.append({"type": "disable", "attr": left, "source_row": rij})

        elif left and right:
            result.append({"type": "merge", "groep": [left, right], "source_row": rij})

    return result

def lm5_all_single_attrs():
    return [a for a in attracties_te_plannen if " + " not in str(a)]

def lm5_full_capable_students():
    singles = lm5_all_single_attrs()
    out = []
    for s in studenten:
        if all(lm5_student_can_attr(s, attr) for attr in singles):
            out.append(str(s["naam"]).strip())
    return out

# ------------------------------------------------------------
# Pauzevlinder-vervangers
# ------------------------------------------------------------
def lm5_pick_pv_replacements(absent_pv_names, start_uur, base_maps, absentees_set):
    full_capable = set(lm5_full_capable_students())
    pv_name_set = set(lm5_pv_names())

    future_workers = set()
    for (naam, uur), _attr in base_maps["student_hour_attr"].items():
        if uur >= start_uur and naam not in absentees_set:
            future_workers.add(naam)

    chosen = {}
    used = set()

    for pvnaam in absent_pv_names:
        kandidaten = []
        for naam in sorted(future_workers):
            if naam in pv_name_set:
                continue
            if naam in used:
                continue
            if naam not in full_capable:
                continue

            orig_extra_count = 0
            orig_total_count = 0
            for (n, uur), rowlabel in base_maps["student_hour_row"].items():
                if n == naam and uur >= start_uur:
                    orig_total_count += 1
                    if "extra" in str(rowlabel).lower():
                        orig_extra_count += 1

            student = lm5_get_student(naam)
            breedte = student["aantal_attracties"] if student else 999

            kandidaten.append((
                -orig_extra_count,
                orig_total_count,
                -breedte,
                naam
            ))

        if kandidaten:
            kandidaten.sort()
            top = [x[3] for x in kandidaten[:min(3, len(kandidaten))]]
            chosen[pvnaam] = random.choice(top)
            used.add(chosen[pvnaam])

    return chosen

def lm5_active_pv_assignment_for_hour(ctx, uur):
    if uur not in required_pauze_hours:
        return {}

    vrijgegeven_uren = ctx.get("vrijgegeven_pv_uren", set())
    vrijgegeven_pv   = ctx.get("vrijgegeven_pv_naam", None)

    result = {}
    for pvnaam in lm5_pv_names():
        if uur in vrijgegeven_uren and pvnaam == vrijgegeven_pv:
            continue
        if pvnaam in ctx["abs_set"]:
            vervanger = ctx["pv_replacements"].get(pvnaam)
            if vervanger:
                result[pvnaam] = vervanger
        else:
            result[pvnaam] = pvnaam
    return result

def lm5_present_attraction_students_on_hour(ctx, uur):
    bruto = lm5_present_students_on_hour(ctx["base_maps"], uur, ctx["abs_set"])
    pv_assignment = lm5_active_pv_assignment_for_hour(ctx, uur)
    bezette_pv_mensen = set(pv_assignment.values())
    return [naam for naam in bruto if naam not in bezette_pv_mensen]

# ------------------------------------------------------------
# Uurstaat herberekenen
# ------------------------------------------------------------
def lm5_rebuild_hour_state(uur, available_attraction_students, capacity_actions):
    counts = {}
    active = set()
    debug_actions = []

    # -----------------------------
    # Basis actieve attracties opbouwen voor DIT uur
    # -----------------------------
    for attr in attracties_te_plannen:
        if " + " in str(attr):
            continue

        if uur in dichte_uren_per_attr.get(normalize_attr(attr), set()):
            counts[attr] = 0
        else:
            if aantallen_raw.get(attr, 0) >= 1:
                counts[attr] = 1
                active.add(attr)
            else:
                counts[attr] = 0

    # -----------------------------
    # Vaste samenvoegingen van DIT uur eerst toepassen
    # -----------------------------
    groepen = []
    for groep in uur_samenvoegingen.get(uur, []):
        g = [str(x).strip() for x in groep if x and str(x).strip()]
        if len(g) < 2:
            continue

        groepen.append(g)
        sameng = " + ".join(g)
        counts[sameng] = 1
        active.add(sameng)

        for onderdeel in g:
            counts[onderdeel] = 0
            if onderdeel in active:
                active.remove(onderdeel)

    def min_spots():
        return sum(1 for a in active if counts.get(a, 0) >= 1)

    def samengestelde_naam_van_groep(g):
        return " + ".join(g)

    def groep_is_al_actief(g):
        sameng = samengestelde_naam_van_groep(g)
        return sameng in active and all(counts.get(a, 0) == 0 for a in g)

    def merge_is_mogelijk_en_geeft_reductie(g):
        # Alleen als ALLE losse onderdelen nog actief zijn
        # Dan krijg je een echte reductie van minstens 1 plek
        if len(g) < 2:
            return False
        if not all(a in active for a in g):
            return False

        before = min_spots()
        after = before - len(g) + 1
        return after < before

    def disable_is_mogelijk_en_geeft_reductie(attr):
        if attr not in active:
            return False
        if counts.get(attr, 0) < 1:
            return False

        before = min_spots()
        after = before - 1
        return after < before

    debug_actions.append(
        f"Uur {uur}: START | available={available_attraction_students} | min_spots={min_spots()} | active_start={sorted(active)}"
    )

    # -----------------------------
    # PER UUR: rij voor rij door CD/CE
    # -----------------------------
    while min_spots() > available_attraction_students:
        found_reduction = False

        for entry in capacity_actions:
            before = min_spots()

            if before <= available_attraction_students:
                debug_actions.append(
                    f"Uur {uur}: STOP acties | min_spots={before} <= available={available_attraction_students}"
                )
                break

            if entry["type"] == "merge":
                g = [str(x).strip() for x in entry["groep"] if x and str(x).strip()]
                sameng = samengestelde_naam_van_groep(g)
                source_row = entry.get("source_row", "?")

                if groep_is_al_actief(g):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE SKIP | groep={g} | reden=al_actief"
                    )
                    continue

                if not merge_is_mogelijk_en_geeft_reductie(g):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE SKIP | groep={g} | reden=niet_mogelijk_op_dit_uur | current_active={sorted(active)}"
                    )
                    continue

                for onderdeel in g:
                    counts[onderdeel] = 0
                    if onderdeel in active:
                        active.remove(onderdeel)

                counts[sameng] = 1
                active.add(sameng)

                if g not in groepen:
                    groepen.append(g)

                after = min_spots()

                if after < before:
                    found_reduction = True
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE OK | groep={g} | min_spots {before}->{after} | new_active={sorted(active)}"
                    )
                    break
                else:
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} MERGE GEEN EFFECT | groep={g}"
                    )

            elif entry["type"] == "disable":
                attr = str(entry["attr"]).strip()
                source_row = entry.get("source_row", "?")

                if not disable_is_mogelijk_en_geeft_reductie(attr):
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE SKIP | attr={attr} | reden=niet_mogelijk_op_dit_uur"
                    )
                    continue

                counts[attr] = 0
                if attr in active:
                    active.remove(attr)

                after = min_spots()

                if after < before:
                    found_reduction = True
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE OK | attr={attr} | min_spots {before}->{after} | new_active={sorted(active)}"
                    )
                    break
                else:
                    debug_actions.append(
                        f"Uur {uur}: RIJ {source_row} DISABLE GEEN EFFECT | attr={attr}"
                    )

        if not found_reduction:
            debug_actions.append(
                f"Uur {uur}: GEEN VERDERE REDUCTIE MOGELIJK | min_spots={min_spots()} | available={available_attraction_students}"
            )
            break

    # -----------------------------
    # Tweede plaatsen pas NA merge/disable
    # -----------------------------
    second_spot_blocked_lm = set()
    base_spots = min_spots()
    extra_spots = available_attraction_students - base_spots

    debug_actions.append(
        f"Uur {uur}: VOOR 2DE PLEKKEN | base_spots={base_spots} | extra_spots={extra_spots}"
    )

    for attr in second_priority_order:
        if attr in active and aantallen_raw.get(attr, 0) == 2:
            if extra_spots > 0:
                counts[attr] = 2
                extra_spots -= 1
                debug_actions.append(
                    f"Uur {uur}: 2DE PLEK OPEN | attr={attr} | resterend_extra_spots={extra_spots}"
                )
            else:
                counts[attr] = 1
                second_spot_blocked_lm.add(attr)
                debug_actions.append(
                    f"Uur {uur}: 2DE PLEK DICHT | attr={attr}"
                )

    # -----------------------------
    # Red spots opnieuw opbouwen
    # -----------------------------
    red_spots_lm = set()

    samengestelde_actief = set(" + ".join(g) for g in groepen)
    losse_in_samenvoeging = set(a for g in groepen for a in g)

    for attr in losse_in_samenvoeging:
        red_spots_lm.add(attr)

    for samengestelde_attr in samengevoegde_attracties:
        if samengestelde_attr not in samengestelde_actief:
            red_spots_lm.add(samengestelde_attr)

    for attr in list(counts.keys()):
        if " + " not in str(attr):
            if uur in dichte_uren_per_attr.get(normalize_attr(attr), set()):
                red_spots_lm.add(attr)

    debug_actions.append(
        f"Uur {uur}: EINDE | active_end={sorted(active)} | counts={{{', '.join(f'{k}: {v}' for k, v in sorted(counts.items()))}}}"
    )

    return {
        "counts": counts,
        "active": active,
        "groepen": groepen,
        "red_spots": red_spots_lm,
        "second_spot_blocked": second_spot_blocked_lm,
        "debug_actions": debug_actions,
    }


# ------------------------------------------------------------
# Toewijzingshelpers
# ------------------------------------------------------------
def lm5_student_current_attr_on_hour(ctx, naam, uur):
    for (h, attr), namen in ctx["assigned_map"].items():
        if h == uur and naam in namen:
            return attr
    return None

def lm5_can_place_student_on_attr(ctx, student, attr, uren):
    if not lm5_student_can_attr(student, attr):
        return False

    for uur in uren:
        hstate = ctx["hour_states"].get(uur)
        if not hstate:
            return False

        if attr not in hstate["active"]:
            return False
        if attr in hstate["red_spots"]:
            return False

        max_spots = hstate["counts"].get(attr, 0)
        if attr in hstate["second_spot_blocked"]:
            max_spots = min(max_spots, 1)

        if ctx["per_hour_assigned_counts"][uur][attr] >= max_spots:
            return False

        if uur in student["assigned_hours"]:
            return False

        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        if student["naam"] in set(pv_assignment_now.values()):
            return False

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(uren))

    if len(totaal) > 6:
        return False

    return True

def lm5_place_student_on_attr(ctx, student, attr, uren):
    for uur in uren:
        ctx["assigned_map"][(uur, attr)].append(student["naam"])
        ctx["per_hour_assigned_counts"][uur][attr] += 1
        if uur not in student["assigned_hours"]:
            student["assigned_hours"].append(uur)
    student["assigned_attracties"].add(attr)
    ctx["prev_attr"][student["naam"]] = attr

def lm5_remove_student_from_attr_hours(ctx, student, attr, uren):
    for uur in uren:
        namen = ctx["assigned_map"].get((uur, attr), [])
        if student["naam"] in namen:
            namen.remove(student["naam"])
            if ctx["per_hour_assigned_counts"][uur][attr] > 0:
                ctx["per_hour_assigned_counts"][uur][attr] -= 1

def lm5_rebuild_student_attracties(ctx, student):
    attrs = set()
    for (uur, attr), namen in ctx["assigned_map"].items():
        if student["naam"] in namen:
            attrs.add(attr)
    for a in list(student["assigned_attracties"]):
        if a in ["Extra", "Pauzevlinder-vervanging"]:
            attrs.add(a)
    student["assigned_attracties"] = attrs

def lm5_original_attr_score(ctx, naam, attr, uren):
    same = 0
    mismatches = 0
    for uur in uren:
        orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
        if normalize_attr(orig_attr) == normalize_attr(attr):
            same += 1
        else:
            mismatches += 1
    return (-same, mismatches)

def lm5_candidate_attr_score(ctx, student, attr, uren):
    naam = student["naam"]
    orig_score = lm5_original_attr_score(ctx, naam, attr, uren)

    prev_penalty = 0
    prev_attr = ctx["prev_attr"].get(naam, "")
    if prev_attr and normalize_attr(prev_attr) != normalize_attr(attr):
        prev_penalty = 1

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if naam in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(uren))

    run_penalty = 0 if len(uren) >= 3 else (1 if len(uren) == 2 else 2)
    over4_penalty = 1 if len(totaal) > 4 else 0

    return (
        run_penalty,
        prev_penalty,
        over4_penalty,
        orig_score[0],
        orig_score[1],
        ctx["changes_count"][naam],
        attr
    )

def lm5_try_place_best_block(ctx, student, future_hours, start_idx):
    if start_idx >= len(future_hours):
        return False, start_idx + 1

    for block_size in [3, 2, 1]:
        uren = future_hours[start_idx:start_idx + block_size]
        if len(uren) < block_size:
            continue
        if any(uur in student["assigned_hours"] for uur in uren):
            continue

        candidate_attrs = []
        for attr in attracties_te_plannen:
            if lm5_can_place_student_on_attr(ctx, student, attr, uren):
                candidate_attrs.append(attr)

        if candidate_attrs:
            candidate_attrs.sort(key=lambda a: lm5_candidate_attr_score(ctx, student, a, uren))
            best_attr = candidate_attrs[0]
            lm5_place_student_on_attr(ctx, student, best_attr, uren)
            return True, start_idx + block_size

    return False, start_idx + 1

# ------------------------------------------------------------
# Uur-per-uur seed
# ------------------------------------------------------------
def lm5_seed_same_place_first(ctx, uur, target_slots, present_attraction_students):
    used_students = set()
    assigned_rows = set()

    for attr, pos, rijlabel in target_slots:
        orig_names = []
        for (naam, uur2), rowlabel in ctx["base_maps"]["student_hour_row"].items():
            if uur2 == uur and rowlabel == rijlabel and naam in present_attraction_students:
                orig_names.append(naam)

        for naam in orig_names:
            if naam in used_students:
                continue
            student = ctx["student_states"][naam]
            if lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                lm5_place_student_on_attr(ctx, student, attr, [uur])
                used_students.add(naam)
                assigned_rows.add(rijlabel)
                break

    return used_students, assigned_rows

def lm5_fill_remaining_hour(ctx, uur, target_slots, present_attraction_students, used_students, assigned_rows):
    remaining_slots = [(attr, pos, rijlabel) for attr, pos, rijlabel in target_slots if rijlabel not in assigned_rows]
    remaining_students = [n for n in present_attraction_students if n not in used_students]

    for attr, pos, rijlabel in remaining_slots:
        kandidaten = []
        for naam in remaining_students:
            student = ctx["student_states"][naam]
            if not lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                continue

            orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
            same_orig = 0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1
            same_prev = 0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1

            kandidaten.append((
                same_orig,
                same_prev,
                ctx["changes_count"][naam],
                student["aantal_attracties"],
                naam
            ))

        if kandidaten:
            kandidaten.sort()
            gekozen = kandidaten[0][4]
            student = ctx["student_states"][gekozen]
            lm5_place_student_on_attr(ctx, student, attr, [uur])

            orig_attr = ctx["base_maps"]["student_hour_attr"].get((gekozen, uur), "")
            if normalize_attr(orig_attr) != normalize_attr(attr):
                ctx["changes_count"][gekozen] += 1

            used_students.add(gekozen)
            remaining_students.remove(gekozen)

# ------------------------------------------------------------
# Vrijgekomen studenten + lege plaatsen
# ------------------------------------------------------------
def lm5_collect_released_students_and_missing_slots(ctx, base_maps, start_uur):
    released_students = defaultdict(list)
    missing_slots_by_hour = defaultdict(list)

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)
        hstate = ctx["hour_states"][uur]
        target_slots, _inactive = lm5_build_target_slots_for_hour(base_maps["attr_rows"], hstate)

        for attr, pos, rijlabel in target_slots:
            namen = ctx["assigned_map"].get((uur, attr), [])
            if len(namen) < pos:
                missing_slots_by_hour[uur].append((attr, pos, rijlabel))

        for naam in present_attraction_students:
            if uur not in ctx["student_states"][naam]["assigned_hours"]:
                released_students[uur].append(naam)

    return released_students, missing_slots_by_hour

# ------------------------------------------------------------
# Eerst directe invulling
# ------------------------------------------------------------
def lm5_try_direct_fill_from_released_students(ctx, released_students, missing_slots_by_hour):
    any_change = False

    for uur in sorted(missing_slots_by_hour.keys()):
        slots = list(missing_slots_by_hour[uur])
        remaining_released = list(released_students.get(uur, []))
        new_slots = []

        for (attr, pos, rijlabel) in slots:
            gevuld = False
            kandidaten = []

            for naam in remaining_released:
                student = ctx["student_states"][naam]
                if lm5_can_place_student_on_attr(ctx, student, attr, [uur]):
                    orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
                    kandidaten.append((
                        0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1,
                        0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1,
                        ctx["changes_count"][naam],
                        naam
                    ))

            if kandidaten:
                kandidaten.sort()
                gekozen = kandidaten[0][3]
                student = ctx["student_states"][gekozen]
                lm5_place_student_on_attr(ctx, student, attr, [uur])

                if gekozen in remaining_released:
                    remaining_released.remove(gekozen)

                orig_attr = ctx["base_maps"]["student_hour_attr"].get((gekozen, uur), "")
                if normalize_attr(orig_attr) != normalize_attr(attr):
                    ctx["changes_count"][gekozen] += 1

                gevuld = True
                any_change = True

            if not gevuld:
                new_slots.append((attr, pos, rijlabel))

        released_students[uur] = remaining_released
        missing_slots_by_hour[uur] = new_slots

    return any_change

# ------------------------------------------------------------
# Kettingwissels op exact hetzelfde blok
# ------------------------------------------------------------
def lm5_student_has_exact_block_on_attr(ctx, naam, attr, block_hours):
    for uur in block_hours:
        huidige_attr = lm5_student_current_attr_on_hour(ctx, naam, uur)
        if huidige_attr != attr:
            return False
    return True

def lm5_find_same_block_students(ctx, block_hours, exclude_names=None):
    if exclude_names is None:
        exclude_names = set()

    kandidaten = []
    alle_namen = sorted(ctx["student_states"].keys())

    for naam in alle_namen:
        if naam in exclude_names:
            continue

        attrs = set()
        ok = True
        for uur in block_hours:
            attr = lm5_student_current_attr_on_hour(ctx, naam, uur)
            if not attr:
                ok = False
                break
            attrs.add(attr)

        if ok and len(attrs) == 1:
            kandidaten.append((naam, list(attrs)[0]))

    return kandidaten

def lm5_can_student_take_attr_block(ctx, student, attr, block_hours):
    if not lm5_student_can_attr(student, attr):
        return False

    for uur in block_hours:
        hstate = ctx["hour_states"].get(uur)
        if not hstate:
            return False
        if attr not in hstate["active"]:
            return False
        if attr in hstate["red_spots"]:
            return False

        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        if student["naam"] in set(pv_assignment_now.values()):
            return False

        huidige_attr = lm5_student_current_attr_on_hour(ctx, student["naam"], uur)
        if huidige_attr and huidige_attr != attr:
            return False

        max_spots = hstate["counts"].get(attr, 0)
        if attr in hstate["second_spot_blocked"]:
            max_spots = min(max_spots, 1)

        current_count = ctx["per_hour_assigned_counts"][uur][attr]
        already_here = student["naam"] in ctx["assigned_map"].get((uur, attr), [])
        effective_count = current_count if already_here else current_count + 1

        if effective_count > max_spots:
            return False

    bestaande = sorted([
        h for h in set(student["assigned_hours"])
        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
    ])
    totaal = sorted(set(bestaande) | set(block_hours))
    if len(totaal) > 6:
        return False
    return True

def lm5_try_chain_swap_for_block(ctx, released_student_name, target_attr, block_hours):
    released_student = ctx["student_states"][released_student_name]

    kandidaten = lm5_find_same_block_students(
        ctx=ctx,
        block_hours=block_hours,
        exclude_names={released_student_name}
    )

    for andere_naam, attr_b in kandidaten:
        if attr_b == target_attr:
            continue

        andere_student = ctx["student_states"][andere_naam]

        if not lm5_student_can_attr(released_student, attr_b):
            continue
        if not lm5_student_can_attr(andere_student, target_attr):
            continue

        orig_switches_r = lm5_count_switches(ctx, released_student_name)
        orig_switches_o = lm5_count_switches(ctx, andere_naam)

        saved_assigned_map = copy.deepcopy(ctx["assigned_map"])
        saved_counts = copy.deepcopy(ctx["per_hour_assigned_counts"])
        saved_hours_r = list(released_student["assigned_hours"])
        saved_hours_o = list(andere_student["assigned_hours"])
        saved_attrs_r = set(released_student["assigned_attracties"])
        saved_attrs_o = set(andere_student["assigned_attracties"])
        saved_prev = dict(ctx["prev_attr"])

        lm5_remove_student_from_attr_hours(ctx, andere_student, attr_b, block_hours)

        if not lm5_can_student_take_attr_block(ctx, released_student, attr_b, block_hours):
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        lm5_place_student_on_attr(ctx, released_student, attr_b, block_hours)

        if not lm5_can_student_take_attr_block(ctx, andere_student, target_attr, block_hours):
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        lm5_place_student_on_attr(ctx, andere_student, target_attr, block_hours)

        lm5_rebuild_student_attracties(ctx, released_student)
        lm5_rebuild_student_attracties(ctx, andere_student)

        # harde 6u/4u check op betrokken attracties
        if len(lm5_get_hours_on_attr(ctx, released_student_name, attr_b)) > 6:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        if len(lm5_get_hours_on_attr(ctx, andere_naam, target_attr)) > 6:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        new_switches_r = lm5_count_switches(ctx, released_student_name)
        new_switches_o = lm5_count_switches(ctx, andere_naam)
        extra_switches = (new_switches_r - orig_switches_r) + (new_switches_o - orig_switches_o)

        if extra_switches > 1:
            ctx["assigned_map"] = saved_assigned_map
            ctx["per_hour_assigned_counts"] = saved_counts
            released_student["assigned_hours"] = saved_hours_r
            andere_student["assigned_hours"] = saved_hours_o
            released_student["assigned_attracties"] = saved_attrs_r
            andere_student["assigned_attracties"] = saved_attrs_o
            ctx["prev_attr"] = saved_prev
            continue

        ctx["changes_count"][released_student_name] += 1
        ctx["changes_count"][andere_naam] += 1
        return True

    return False

def lm5_try_fill_missing_with_chain_swaps(ctx, released_students, missing_slots_by_hour, start_uur):
    any_change = False
    future_hours = [u for u in sorted(open_uren) if u >= start_uur]

    for block_size in [3, 2, 1]:
        for i in range(len(future_hours) - block_size + 1):
            block_hours = future_hours[i:i + block_size]

            target_attrs = None
            for uur in block_hours:
                attrs_this_hour = set(attr for attr, _pos, _rijlabel in missing_slots_by_hour.get(uur, []))
                if target_attrs is None:
                    target_attrs = attrs_this_hour
                else:
                    target_attrs &= attrs_this_hour

            if not target_attrs:
                continue

            for target_attr in sorted(target_attrs):
                kandidaten_released = []
                for naam in sorted(set(n for uur in block_hours for n in released_students.get(uur, []))):
                    if all(naam in released_students.get(uur, []) for uur in block_hours):
                        kandidaten_released.append(naam)

                for released_name in kandidaten_released:
                    released_student = ctx["student_states"][released_name]

                    # eerst directe blokinvulling
                    if lm5_can_student_take_attr_block(ctx, released_student, target_attr, block_hours):
                        lm5_place_student_on_attr(ctx, released_student, target_attr, block_hours)
                        for uur in block_hours:
                            if released_name in released_students[uur]:
                                released_students[uur].remove(released_name)
                            removed = False
                            nieuwe = []
                            for item in missing_slots_by_hour[uur]:
                                if item[0] == target_attr and not removed:
                                    removed = True
                                    continue
                                nieuwe.append(item)
                            missing_slots_by_hour[uur] = nieuwe
                        any_change = True
                        break

                    # anders kettingwissel
                    if lm5_try_chain_swap_for_block(ctx, released_name, target_attr, block_hours):
                        for uur in block_hours:
                            if released_name in released_students[uur]:
                                released_students[uur].remove(released_name)
                            removed = False
                            nieuwe = []
                            for item in missing_slots_by_hour[uur]:
                                if item[0] == target_attr and not removed:
                                    removed = True
                                    continue
                                nieuwe.append(item)
                            missing_slots_by_hour[uur] = nieuwe
                        any_change = True
                        break

    return any_change

# ------------------------------------------------------------
# Nadien blokken 3/2/1 maken voor overige gaten
# ------------------------------------------------------------
def lm5_assign_future_blocks(ctx, start_uur):
    future_students = sorted({
        naam for (naam, uur), _attr in ctx["base_maps"]["student_hour_attr"].items()
        if uur >= start_uur and naam not in ctx["abs_set"]
    })

    for naam in future_students:
        student = ctx["student_states"][naam]
        future_hours = sorted({
            uur for (n, uur), _attr in ctx["base_maps"]["student_hour_attr"].items()
            if n == naam and uur >= start_uur
        })

        if not future_hours:
            continue

        i = 0
        while i < len(future_hours):
            uur = future_hours[i]

            if uur in student["assigned_hours"]:
                i += 1
                continue

            pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
            if naam in set(pv_assignment_now.values()):
                student["assigned_hours"].append(uur)
                student["assigned_attracties"].add("Pauzevlinder-vervanging")
                ctx["prev_attr"][naam] = "Pauzevlinder-vervanging"
                i += 1
                continue

            placed, next_i = lm5_try_place_best_block(ctx, student, future_hours, i)
            if placed:
                i = next_i
            else:
                i += 1

# ------------------------------------------------------------
# Exact 1 keer per gewerkt uur
# ------------------------------------------------------------
def lm5_force_exactly_one_assignment_per_hour(ctx, start_uur):
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_students = lm5_present_students_on_hour(ctx["base_maps"], uur, ctx["abs_set"])
        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        pv_reserved_now = set(pv_assignment_now.values())

        for naam in present_students:
            student = ctx["student_states"][naam]

            if uur in student["assigned_hours"]:
                continue

            if naam in pv_reserved_now:
                student["assigned_hours"].append(uur)
                student["assigned_attracties"].add("Pauzevlinder-vervanging")
                ctx["prev_attr"][naam] = "Pauzevlinder-vervanging"
                continue

            ctx["extra_assignments"][uur].append(naam)
            student["assigned_hours"].append(uur)
            student["assigned_attracties"].add("Extra")
            ctx["prev_attr"][naam] = "Extra"

# ------------------------------------------------------------
# Extra -> attractie verplaatsing
# ------------------------------------------------------------
def lm5_try_fill_empty_slots_from_extras(ctx, start_uur):
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        hstate = ctx["hour_states"][uur]
        extras_now = list(ctx["extra_assignments"][uur])

        for attr in hstate["active"]:
            max_spots = hstate["counts"].get(attr, 0)
            if attr in hstate["second_spot_blocked"]:
                max_spots = min(max_spots, 1)

            current = list(ctx["assigned_map"].get((uur, attr), []))

            while len(current) < max_spots:
                kandidaten = []

                for naam in extras_now:
                    student = ctx["student_states"][naam]

                    if not lm5_student_can_attr(student, attr):
                        continue

                    bestaande = sorted([
                        h for h in set(student["assigned_hours"])
                        if student["naam"] in ctx["assigned_map"].get((h, attr), [])
                    ])
                    totaal = sorted(set(bestaande) | {uur})

                    if len(totaal) > 6:
                        continue

                    orig_attr = ctx["base_maps"]["student_hour_attr"].get((naam, uur), "")
                    kandidaten.append((
                        0 if normalize_attr(orig_attr) == normalize_attr(attr) else 1,
                        0 if normalize_attr(ctx["prev_attr"].get(naam, "")) == normalize_attr(attr) else 1,
                        ctx["changes_count"][naam],
                        naam
                    ))

                if not kandidaten:
                    break

                kandidaten.sort()
                gekozen = kandidaten[0][3]

                ctx["extra_assignments"][uur].remove(gekozen)
                extras_now.remove(gekozen)

                ctx["assigned_map"][(uur, attr)].append(gekozen)
                ctx["per_hour_assigned_counts"][uur][attr] += 1
                ctx["student_states"][gekozen]["assigned_attracties"].add(attr)
                ctx["prev_attr"][gekozen] = attr

                current = list(ctx["assigned_map"].get((uur, attr), []))

# ------------------------------------------------------------
# Complete build
# ------------------------------------------------------------
def lm5_build_target_slots_for_hour(attr_rows, hour_state):
    slots = []
    inactive_rows = set()

    for _row, rijlabel in attr_rows:
        attr, pos = lm5_split_display_label(rijlabel)
        allowed = hour_state["counts"].get(attr, 0)

        if attr in hour_state["red_spots"]:
            allowed = 0

        if attr in hour_state["second_spot_blocked"] and pos == 2:
            allowed = 0

        if allowed >= pos:
            slots.append((attr, pos, rijlabel))
        else:
            inactive_rows.add(rijlabel)

    return slots, inactive_rows


# ------------------------------------------------------------
# Context
# ------------------------------------------------------------
def lm5_init_context(base_maps, absentees, start_uur):
    abs_set = {str(x).strip() for x in absentees}
    absent_pv = [n for n in abs_set if n in set(lm5_pv_names())]

    pv_replacements = lm5_pick_pv_replacements(
        absent_pv_names=absent_pv,
        start_uur=start_uur,
        base_maps=base_maps,
        absentees_set=abs_set
    )

    student_states = {}
    for s in studenten:
        student_states[str(s["naam"]).strip()] = lm5_copy_student_state(s)

    return {
        "base_maps": base_maps,
        "abs_set": abs_set,
        "absent_pv": absent_pv,
        "pv_replacements": pv_replacements,
        "student_states": student_states,
        "assigned_map": defaultdict(list),
        "extra_assignments": defaultdict(list),
        "per_hour_assigned_counts": {uur: defaultdict(int) for uur in open_uren},
        "hour_states": {},
        "changes_count": defaultdict(int),
        "prev_attr": {},
    }


def lm5_vrijgeven_afgekapte_pv_uren(ctx, start_uur):
    """
    Last-minute only: als de afgekapte PV-uren minstens 2 aaneensluitende
    uren bevatten, worden die uren vrijgegeven zodat de laatste PV (of diens
    vervanger) in die uren gewoon op de planning kan worden ingezet.
    """
    if not afgekapte_pv_uren or not selected:
        return

    # Alleen runs van >= 2 aaneensluitende uren vrijgeven
    runs = contiguous_runs(sorted(afgekapte_pv_uren))
    geschikte_uren = set()
    for run in runs:
        if len(run) >= 2:
            geschikte_uren.update(run)

    if not geschikte_uren:
        return

    _afknip_pv = pp2_bepaal_pv_voor_afknip(selected)
    laatste_pv_naam = str(_afknip_pv["naam"]).strip() if _afknip_pv else ""
    # Bepaal de effectieve persoon (zelf of vervanger)
    effectief_naam = ctx["pv_replacements"].get(laatste_pv_naam, laatste_pv_naam)

    # Sla op in ctx zodat lm5_active_pv_assignment_for_hour het kan gebruiken
    ctx["vrijgegeven_pv_uren"] = geschikte_uren
    ctx["vrijgegeven_pv_naam"] = laatste_pv_naam   # originele PV-naam, niet vervanger

    # Voeg de uren terug toe aan uren_beschikbaar van de effectieve persoon
    if effectief_naam in ctx["student_states"]:
        s = ctx["student_states"][effectief_naam]
        for uur in sorted(geschikte_uren):
            if uur >= start_uur and uur not in s["uren_beschikbaar"]:
                s["uren_beschikbaar"].append(uur)


def lm5_seed_hours_before_start(ctx, start_uur):
    for (naam, uur), attr in ctx["base_maps"]["student_hour_attr"].items():
        if uur >= start_uur:
            continue
        if naam in ctx["abs_set"]:
            continue

        ctx["assigned_map"][(uur, attr)].append(naam)
        ctx["per_hour_assigned_counts"][uur][attr] += 1
        ctx["student_states"][naam]["assigned_hours"].append(uur)
        ctx["student_states"][naam]["assigned_attracties"].add(attr)
        ctx["prev_attr"][naam] = attr


# ------------------------------------------------------------
# Complete build
# ------------------------------------------------------------

def lm5_extend_attr_rows_with_dynamic_merges(base_maps, ctx, start_uur):
    attr_rows = list(base_maps["attr_rows"])
    pv_rows = list(base_maps.get("pv_rows", []))
    extra_rows = list(base_maps.get("extra_rows", []))

    bestaande_attr_pos = set()
    for row, rijlabel in attr_rows:
        attr, pos = lm5_split_display_label(rijlabel)
        bestaande_attr_pos.add((attr, pos))

    # Verzamel alle dynamische samengestelde attracties die effectief voorkomen
    dynamische_attrs = {}

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue
        if uur not in ctx["hour_states"]:
            continue

        hstate = ctx["hour_states"][uur]

        for attr in hstate["active"]:
            if " + " not in str(attr):
                continue

            max_pos = hstate["counts"].get(attr, 0)
            if attr in hstate["second_spot_blocked"]:
                max_pos = min(max_pos, 1)
            max_pos = max(1, max_pos)

            dynamische_attrs[attr] = max(dynamische_attrs.get(attr, 0), max_pos)

        for attr, count in hstate["counts"].items():
            if " + " not in str(attr):
                continue
            if count < 1:
                continue

            max_pos = count
            if attr in hstate["second_spot_blocked"]:
                max_pos = min(max_pos, 1)
            max_pos = max(1, max_pos)

            dynamische_attrs[attr] = max(dynamische_attrs.get(attr, 0), max_pos)

    if not dynamische_attrs:
        return

    # Bepaal waar de nieuwe rijen moeten komen:
    # liefst ONDER de extra-rijen, anders onder de pauzevlinderrijen,
    # anders onder de bestaande attractierijen.
    alle_bestaande_rijen = [row for row, _ in attr_rows] + [row for row, _ in pv_rows] + [row for row, _ in extra_rows]
    if alle_bestaande_rijen:
        insert_after = max(alle_bestaande_rijen)
    else:
        insert_after = 1

    nieuwe_attr_rows = []

    current_row = insert_after + 1

    for attr in sorted(dynamische_attrs.keys(), key=lambda x: str(x).lower()):
        nodig = dynamische_attrs[attr]

        for pos in range(1, nodig + 1):
            if (attr, pos) in bestaande_attr_pos:
                continue

            current_row += 1

            if nodig > 1:
                rijlabel = f"{attr} {pos}"
            else:
                rijlabel = attr

            nieuwe_attr_rows.append((current_row, rijlabel))
            bestaande_attr_pos.add((attr, pos))

    # Voeg de nieuwe rijen toe aan attr_rows
    attr_rows.extend(nieuwe_attr_rows)

    # Sorteer op fysieke rij
    attr_rows.sort(key=lambda x: x[0])

    # Schrijf terug naar base_maps
    base_maps["attr_rows"] = attr_rows


def lm5_extend_extra_rows_if_needed(base_maps, ctx):
    """
    Als ctx meer extra-studenten heeft dan er extra-rijen bestaan,
    voeg dan de ontbrekende 'Extra N'-rijen toe aan base_maps.
    """
    extra_rows = list(base_maps.get("extra_rows", []))

    # Hoeveel extra-plekken zijn er maximaal nodig op één uur?
    max_nodig = 0
    for uur, namen in ctx["extra_assignments"].items():
        max_nodig = max(max_nodig, len(namen))

    extra_tekort = max_nodig - len(extra_rows)
    if extra_tekort <= 0:
        return

    # Bepaal na welke rij de nieuwe extra-rijen komen:
    # 1 lege rij na de laagste bestaande rij in het sheet
    attr_rows  = list(base_maps.get("attr_rows", []))
    pv_rows    = list(base_maps.get("pv_rows", []))
    alle_rijen = [row for row, _ in attr_rows + pv_rows + extra_rows]
    insert_after = max(alle_rijen) if alle_rijen else 1

    # +2: 1 lege rij spatie, dan de eerste nieuwe extra-rij
    current_row = insert_after + 1

    for i in range(extra_tekort):
        nieuw_idx = len(extra_rows) + i + 1
        label = f"Extra {nieuw_idx}"
        extra_rows.append((current_row, label))
        current_row += 1

    base_maps["extra_rows"] = extra_rows


# ------------------------------------------------------------
# Post-processing: 5/6-uursblokken wisselen (last-minute versie)
# ------------------------------------------------------------
def lm5_postprocess_long_blocks(ctx, start_uur):
    """
    Equivalent van de hoofdplanning-postprocessing, maar voor de ctx-structuur.
    Werkt alleen op uren >= start_uur (uren daarvoor zijn ingezaaid en vast).
    """

    am = ctx["assigned_map"]

    # --- Hulpfuncties die werken op ctx ---

    def lm5_pp_get_attr_on_hour(naam, uur):
        for (u, a), namen in am.items():
            if u == uur and naam in namen:
                return a
        return None

    def lm5_pp_get_hours_on_attr(naam, attr):
        return sorted(u for (u, a), namen in am.items() if a == attr and naam in namen)

    def lm5_pp_get_runs_on_attr(naam, attr):
        return contiguous_runs(lm5_pp_get_hours_on_attr(naam, attr))

    def lm5_pp_count_attr_switches(naam):
        uur_attr = sorted(
            (u, lm5_pp_get_attr_on_hour(naam, u))
            for u in sorted(set(ctx["student_states"][naam]["assigned_hours"]))
            if lm5_pp_get_attr_on_hour(naam, u)
        )
        if not uur_attr:
            return 0
        switches = 0
        prev = uur_attr[0][1]
        for _, a in uur_attr[1:]:
            if a != prev:
                switches += 1
            prev = a
        return switches

    def lm5_pp_count_problem_attrs(naam):
        return sum(
            1 for (_, a), namen in am.items()
            if naam in namen and len(lm5_pp_get_hours_on_attr(naam, a)) > 4
        )

    def lm5_pp_total_overflow(naam):
        seen = set()
        overflow = 0
        for (_, a), namen in am.items():
            if naam in namen and a not in seen:
                seen.add(a)
                uren = len(lm5_pp_get_hours_on_attr(naam, a))
                if uren > 4:
                    overflow += uren - 4
        return overflow

    def lm5_pp_remove(naam, uur, attr):
        key = (uur, attr)
        if naam in am.get(key, []):
            am[key].remove(naam)
        s = ctx["student_states"][naam]
        if uur in s["assigned_hours"]:
            s["assigned_hours"].remove(uur)

    def lm5_pp_add(naam, uur, attr):
        am[(uur, attr)].append(naam)
        ctx["student_states"][naam]["assigned_hours"].append(uur)
        ctx["student_states"][naam]["assigned_attracties"].add(attr)

    def lm5_pp_rebuild_attrs(naam):
        s = ctx["student_states"][naam]
        s["assigned_attracties"] = {
            lm5_pp_get_attr_on_hour(naam, u)
            for u in set(s["assigned_hours"])
            if lm5_pp_get_attr_on_hour(naam, u)
        }

    def lm5_pp_is_valid(naam, attr, uren):
        """Student mag attr op al die uren doen, en uren >= start_uur."""
        if not all(u >= start_uur for u in uren):
            return False
        s = ctx["student_states"].get(naam)
        if not s or not lm5_student_can_attr(s, attr):
            return False
        for u in uren:
            hstate = ctx["hour_states"].get(u, {})
            if attr not in hstate.get("active", set()):
                return False
        return True

    def lm5_pp_respects_rules(naam, attr):
        uren = lm5_pp_get_hours_on_attr(naam, attr)
        if len(uren) > 6:
            return False
        if max_consecutive_hours(uren) > 4:
            return False
        return True

    def lm5_pp_is_swap_target(naam, attr, block_hours):
        """Staat naam op exact al die uren op attr?"""
        return all(naam in am.get((u, attr), []) for u in block_hours)

    def lm5_pp_try_swap_block(naam_a, attr_a, block_hours, all_names):
        if len(block_hours) not in [2, 3]:
            return False

        orig_sw_a  = lm5_pp_count_attr_switches(naam_a)
        orig_pr_a  = lm5_pp_count_problem_attrs(naam_a)
        orig_ov_a  = lm5_pp_total_overflow(naam_a)
        eerste_uur = block_hours[0]

        for naam_b in all_names:
            if naam_b == naam_a:
                continue

            attr_b = lm5_pp_get_attr_on_hour(naam_b, eerste_uur)
            if not attr_b or attr_b == attr_a:
                continue
            if not lm5_pp_is_swap_target(naam_b, attr_b, block_hours):
                continue
            if not lm5_pp_is_valid(naam_a, attr_b, block_hours):
                continue
            if not lm5_pp_is_valid(naam_b, attr_a, block_hours):
                continue

            orig_sw_b = lm5_pp_count_attr_switches(naam_b)
            orig_pr_b = lm5_pp_count_problem_attrs(naam_b)
            orig_ov_b = lm5_pp_total_overflow(naam_b)

            # Wissel uitvoeren
            for u in block_hours:
                lm5_pp_remove(naam_a, u, attr_a)
                lm5_pp_remove(naam_b, u, attr_b)
            for u in block_hours:
                lm5_pp_add(naam_a, u, attr_b)
                lm5_pp_add(naam_b, u, attr_a)
            lm5_pp_rebuild_attrs(naam_a)
            lm5_pp_rebuild_attrs(naam_b)

            # Geen geïsoleerde 1-uursblokken laten ontstaan
            def heeft_geisoleerd_uur(naam, attr):
                runs = lm5_pp_get_runs_on_attr(naam, attr)
                return any(len(r) == 1 for r in runs)

            valid = all(
                lm5_pp_respects_rules(n, a)
                for n, a in [(naam_a, attr_a), (naam_a, attr_b),
                             (naam_b, attr_a), (naam_b, attr_b)]
            )

            if heeft_geisoleerd_uur(naam_a, attr_a): valid = False
            if heeft_geisoleerd_uur(naam_a, attr_b): valid = False
            if heeft_geisoleerd_uur(naam_b, attr_a): valid = False
            if heeft_geisoleerd_uur(naam_b, attr_b): valid = False

            extra_sw = (
                (lm5_pp_count_attr_switches(naam_a) - orig_sw_a) +
                (lm5_pp_count_attr_switches(naam_b) - orig_sw_b)
            )
            if extra_sw > 2:
                valid = False

            new_pr = lm5_pp_count_problem_attrs(naam_a) + lm5_pp_count_problem_attrs(naam_b)
            new_ov = lm5_pp_total_overflow(naam_a)      + lm5_pp_total_overflow(naam_b)
            orig_pr = orig_pr_a + orig_pr_b
            orig_ov = orig_ov_a + orig_ov_b

            if new_pr > orig_pr:
                valid = False
            if new_pr == orig_pr and new_ov > orig_ov:
                valid = False

            verbetering = (
                new_pr < orig_pr
                or (new_pr == orig_pr and new_ov < orig_ov)
            )
            if not verbetering:
                valid = False

            if valid:
                return True

            # Rollback
            for u in block_hours:
                lm5_pp_remove(naam_a, u, attr_b)
                lm5_pp_remove(naam_b, u, attr_a)
            for u in block_hours:
                lm5_pp_add(naam_a, u, attr_a)
                lm5_pp_add(naam_b, u, attr_b)
            lm5_pp_rebuild_attrs(naam_a)
            lm5_pp_rebuild_attrs(naam_b)

        return False

    def lm5_pp_try_swap_long_attr(naam, attr, all_names):
        if len(lm5_pp_get_hours_on_attr(naam, attr)) <= 4:
            return False

        runs = lm5_pp_get_runs_on_attr(naam, attr)
        if not runs:
            return False

        def kandidaat_blokken(run):
            blokken = []
            if len(run) >= 3:
                blokken.append(run[-3:])
            if len(run) >= 2:
                blokken.append(run[-2:])
            if len(run) >= 3 and run[:3] != run[-3:]:
                blokken.append(run[:3])
            if len(run) >= 2 and run[:2] != run[-2:]:
                blokken.append(run[:2])
            return blokken

        laatste_run = runs[-1]
        eerste_run  = runs[0]

        for blok in kandidaat_blokken(laatste_run):
            if lm5_pp_try_swap_block(naam, attr, blok, all_names):
                return True
        if eerste_run != laatste_run:
            for blok in kandidaat_blokken(eerste_run):
                if lm5_pp_try_swap_block(naam, attr, blok, all_names):
                    return True
        return False

    # --- Hoofdlus ---
    all_names = [
        naam for naam, s in ctx["student_states"].items()
        if any(u >= start_uur for u in s["assigned_hours"])
        and not s.get("is_pauzevlinder")
    ]

    for _ in range(7):
        wijziging = False
        for naam in all_names:
            s = ctx["student_states"][naam]
            probleem_attrs = sorted(
                {a for a in s["assigned_attracties"]
                 if len(lm5_pp_get_hours_on_attr(naam, a)) > 4},
                key=lambda a: -len(lm5_pp_get_hours_on_attr(naam, a))
            )
            for attr in probleem_attrs:
                if lm5_pp_try_swap_long_attr(naam, attr, all_names):
                    wijziging = True
                    break
        if not wijziging:
            break



def lm5_build_lastminute_context(base_bytes, absentees, start_uur):
    base_maps = lm5_extract_base_maps(base_bytes)
    ctx = lm5_init_context(base_maps, absentees, start_uur)
    herbereken_afgekapte_pv_uren(absentees_set=ctx["abs_set"], base_maps=base_maps)
    lm5_vrijgeven_afgekapte_pv_uren(ctx, start_uur)
    lm5_seed_hours_before_start(ctx, start_uur)
    capacity_actions = lm5_extract_capacity_actions()


    # STAP 1: per uur echte capaciteit herberekenen
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)

        hour_state = lm5_rebuild_hour_state(
            uur=uur,
            available_attraction_students=len(present_attraction_students),
            capacity_actions=capacity_actions
        )

        ctx["hour_states"][uur] = hour_state
  

    # STAP 2: eerst zoveel mogelijk exact dezelfde plek houden
    for uur in sorted(open_uren):
        if uur < start_uur:
            continue

        hstate = ctx["hour_states"][uur]
        target_slots, _inactive_rows = lm5_build_target_slots_for_hour(base_maps["attr_rows"], hstate)
        present_attraction_students = lm5_present_attraction_students_on_hour(ctx, uur)

        used_students, assigned_rows = lm5_seed_same_place_first(
            ctx=ctx,
            uur=uur,
            target_slots=target_slots,
            present_attraction_students=present_attraction_students
        )

        lm5_fill_remaining_hour(
            ctx=ctx,
            uur=uur,
            target_slots=target_slots,
            present_attraction_students=present_attraction_students,
            used_students=used_students,
            assigned_rows=assigned_rows
        )

    # STAP 3: vrijgekomen studenten en lege plekken verzamelen
    released_students, missing_slots_by_hour = lm5_collect_released_students_and_missing_slots(
        ctx=ctx,
        base_maps=base_maps,
        start_uur=start_uur
    )

    # STAP 4: directe invulling
    lm5_try_direct_fill_from_released_students(
        ctx=ctx,
        released_students=released_students,
        missing_slots_by_hour=missing_slots_by_hour
    )

    # STAP 5: kettingwissels op blokken van 3/2/1
    for _ in range(5):
        changed = lm5_try_fill_missing_with_chain_swaps(
            ctx=ctx,
            released_students=released_students,
            missing_slots_by_hour=missing_slots_by_hour,
            start_uur=start_uur
        )
        if not changed:
            break

    # STAP 6: resterende gaten met blokvoorkeur 3/2/1
    lm5_assign_future_blocks(ctx, start_uur)

    # STAP 7: exact 1 keer per gewerkt uur
    lm5_force_exactly_one_assignment_per_hour(ctx, start_uur)

    # STAP 8: lege attractieplekken opvullen door Extra -> attractie
    lm5_try_fill_empty_slots_from_extras(ctx, start_uur)

    # STAP 9: post-processing 5/6-uursblokken wegwisselen
    lm5_postprocess_long_blocks(ctx, start_uur)

    lm5_extend_extra_rows_if_needed(base_maps, ctx)
    lm5_extend_attr_rows_with_dynamic_merges(base_maps, ctx, start_uur)

    return ctx, base_maps


# ------------------------------------------------------------
# Output schrijven
# ------------------------------------------------------------

def lm5_reconstruct_studenten(ctx_assigned_map):
    """Maak een kopie van studenten met assigned_hours vanuit de nieuwe ctx."""
    import copy as _copy
    hours_per_student = defaultdict(list)
    for (uur, attr), namen in ctx_assigned_map.items():
        for naam in namen:
            hours_per_student[str(naam).strip()].append(uur)

    lm_studenten = _copy.deepcopy(studenten)
    for s in lm_studenten:
        s["assigned_hours"] = hours_per_student.get(str(s["naam"]).strip(), [])
    return lm_studenten

def lm5_write_lastminute_workbook(base_bytes, ctx, base_maps, start_uur, absentees):
    wb_lm = load_workbook(BytesIO(base_bytes))
    ws_plan = wb_lm["Planning"]
    # Last-minute planning toont altijd datum van vandaag, ongeacht W4
    ws_plan.cell(1, 1).value = vandaag_altijd_vandaag
    ws_pauze_lm = wb_lm["Pauzevlinders"]
    ws_pauze_lm.cell(1, 1).value = vandaag_altijd_vandaag

    gray_fill  = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    uur_to_col = base_maps["uur_to_col"]
    attr_rows  = list(base_maps["attr_rows"])
    pv_rows    = base_maps["pv_rows"]
    extra_rows = base_maps["extra_rows"]
    # Samengevoegde attractierijen fysiek boven pauzevlinders plaatsen
    if pv_rows and attr_rows:
        eerste_pv_rij = min(row for row, _ in pv_rows)
        nieuwe_attr = [(row, label) for row, label in attr_rows if row >= eerste_pv_rij]
        if nieuwe_attr:
            aantal = len(nieuwe_attr)
            ws_plan.insert_rows(eerste_pv_rij, amount=aantal + 1)
            verschuiving = aantal + 1
            pv_rows    = [(row + verschuiving, label) for row, label in pv_rows]
            extra_rows = [(row + verschuiving, label) for row, label in extra_rows]
            hernummer  = {old: eerste_pv_rij + i for i, (old, _) in enumerate(nieuwe_attr)}
            attr_rows  = [(hernummer[row], label) if row in hernummer else (row, label) for row, label in attr_rows]
            base_maps["pv_rows"]    = pv_rows
            base_maps["extra_rows"] = extra_rows
            base_maps["attr_rows"]  = attr_rows

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Zorg dat alle dynamische attractierijen echt bestaan in kolom A
    for row, rijlabel in attr_rows:
        huidige_waarde = ws_plan.cell(row, 1).value
        if huidige_waarde is None or str(huidige_waarde).strip() == "":
            ws_plan.cell(row, 1).value     = rijlabel
            ws_plan.cell(row, 1).font      = Font(bold=True)
            ws_plan.cell(row, 1).fill      = white_fill
            ws_plan.cell(row, 1).alignment = center_align
            ws_plan.cell(row, 1).border    = thin_border

        for uur in sorted(open_uren):
            if uur not in uur_to_col:
                continue
            col  = uur_to_col[uur]
            cell = ws_plan.cell(row, col)
            cell.alignment = center_align
            cell.border    = thin_border
            if cell.fill.fill_type is None:
                cell.fill = white_fill

    # Zorg dat ook nieuwe extra-rijen in kolom A staan
    for row, rijlabel in extra_rows:
        huidige_waarde = ws_plan.cell(row, 1).value
        if huidige_waarde is None or str(huidige_waarde).strip() == "":
            ws_plan.cell(row, 1).value     = rijlabel
            ws_plan.cell(row, 1).font      = Font(bold=True)
            ws_plan.cell(row, 1).fill      = white_fill
            ws_plan.cell(row, 1).alignment = left_align
            ws_plan.cell(row, 1).border    = thin_border

            for uur in sorted(open_uren):
                if uur not in uur_to_col:
                    continue
                col  = uur_to_col[uur]
                cell = ws_plan.cell(row, col)
                cell.alignment = center_align
                cell.border    = thin_border
                cell.fill      = white_fill

    for uur in sorted(open_uren):
        if uur < start_uur:
            continue
        if uur not in uur_to_col:
            continue

        col    = uur_to_col[uur]
        hstate = ctx["hour_states"][uur]

        # Attractierijen
        for row, rijlabel in attr_rows:
            cell       = ws_plan.cell(row, col)
            attr, pos  = lm5_split_display_label(rijlabel)
            allowed    = hstate["counts"].get(attr, 0)
            inactive   = False

            if attr in hstate["red_spots"]:
                inactive = True
            if attr in hstate["second_spot_blocked"] and pos == 2:
                inactive = True
            if allowed < pos:
                inactive = True

            namen = list(ctx["assigned_map"].get((uur, attr), []))
            naam  = namen[pos - 1] if pos - 1 < len(namen) else ""

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if inactive:
                cell.fill = gray_fill
            elif naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

        # PV-rijen
        pv_assignment_now = lm5_active_pv_assignment_for_hour(ctx, uur)
        for idx, (row, _label) in enumerate(pv_rows, start=1):
            cell   = ws_plan.cell(row, col)
            pvnaam = str(selected[idx - 1]["naam"]).strip() if idx <= len(selected) else ""
            naam   = pv_assignment_now.get(pvnaam, "")

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

        # Extra-rijen
        extras_now = list(ctx["extra_assignments"][uur])
        for idx, (row, _label) in enumerate(extra_rows):
            cell = ws_plan.cell(row, col)
            naam = extras_now[idx] if idx < len(extras_now) else ""

            cell.value     = naam
            cell.alignment = center_align
            cell.border    = thin_border

            if naam and naam in student_kleuren:
                cell.fill = PatternFill(
                    start_color=student_kleuren[naam],
                    end_color=student_kleuren[naam],
                    fill_type="solid"
                )
            else:
                cell.fill = white_fill

    # Reconstruct assigned_hours per student vanuit ctx
    import copy as _copy
    _hours_per_student = defaultdict(list)
    for (uur, attr), namen in ctx["assigned_map"].items():
        for naam in namen:
            _hours_per_student[str(naam).strip()].append(uur)

    studenten_lm = _copy.deepcopy(studenten)
    for s in studenten_lm:
        s["assigned_hours"] = _hours_per_student.get(str(s["naam"]).strip(), [])

    # Herwerk Analyse
    # Bouw actieve attracties per uur vanuit last-minute hour_states
    lm_actieve_attracties = {}
    for uur in open_uren:
        if uur in ctx["hour_states"]:
            lm_actieve_attracties[uur] = ctx["hour_states"][uur]["active"]
        else:
            lm_actieve_attracties[uur] = actieve_attracties_per_uur.get(uur, set())
    maak_analyse_sheet(
        wb_lm,
        ctx["assigned_map"],
        ctx["extra_assignments"],
        studenten_lm,
        actieve_attracties_override=lm_actieve_attracties
    )
    # Herwerk Wisselplanning
    maak_wisselplanning_sheet(wb_lm, ctx["assigned_map"])

    # Herwerk PP optie 2:
    # Vervang eerst kolom A in Pauzevlinders én tijdelijk selected,
    # zodat maak_pp2_sheets de vervanger als echte PV behandelt
    # (juiste werkuren, eigen pauze op eigen rij).
    selected_bak = [dict(pv) for pv in selected]
    if ctx.get("pv_replacements"):
        ws_pauze_lm = wb_lm["Pauzevlinders"]
        for i, pv in enumerate(selected):
            pvnaam    = str(pv["naam"]).strip()
            vervanger = ctx["pv_replacements"].get(pvnaam)
            if not vervanger:
                continue
            vervanger_student = next(
                (s for s in studenten if str(s["naam"]).strip() == vervanger), None
            )
            if not vervanger_student:
                continue
            # Kolom A in Pauzevlinders bijwerken
            for r in range(2, ws_pauze_lm.max_row + 1):
                if str(ws_pauze_lm.cell(r, 1).value or "").strip() == pvnaam:
                    ws_pauze_lm.cell(r, 1).value = vervanger
                    break
            # selected tijdelijk aanpassen zodat maak_pp2_sheets
            # de vervanger als PV herkent en diens werkuren gebruikt
            selected[i] = dict(vervanger_student)
            selected[i]["is_pauzevlinder"] = True
            selected[i]["pv_number"]       = pv.get("pv_number", i + 1)

    herbereken_afgekapte_pv_uren(absentees_set=ctx["abs_set"], base_maps=base_maps)
    maak_pp2_sheets(wb_lm, ctx["assigned_map"])

    for i in range(len(selected)):
        if i < len(selected_bak):
            selected[i] = selected_bak[i]
    herbereken_afgekapte_pv_uren()  # ← zonder args: terug naar originele staat

    return wb_lm
    
    
# ------------------------------------------------------------
# UI
# ------------------------------------------------------------
st.markdown("### Last-minute afwezigen")


@st.cache_data
def _cached_base_maps(base_bytes):
    return lm5_extract_base_maps(base_bytes)
    
base_bytes_lm5 = st.session_state["lm_base_bytes"]
base_maps_lm5 = _cached_base_maps(base_bytes_lm5)   # ← slaat nu WEL aan
werkende_studenten_vandaag_lm5 = lm5_working_students_today(base_maps_lm5)

with st.expander("Last-minute afwezigen", expanded=False):
    with st.form("lm5_form"):
        gekozen_afwezigen_lm5 = st.multiselect(
            "Kies 1 tot 5 afwezige studenten",
            options=werkende_studenten_vandaag_lm5,
            default=[],
        )
        start_uur_lm5 = st.selectbox(
            "Vanaf welk uur moet de nieuwe planning starten?",
            options=sorted(open_uren),
            format_func=formatteer_uur,
        )
        submitted = st.form_submit_button("Maak last-minute planning")

    if submitted:
        if not gekozen_afwezigen_lm5:
            st.warning("Kies eerst minstens 1 afwezige student.")
        elif len(gekozen_afwezigen_lm5) > 5:
            st.warning("Je mag maximaal 5 studenten kiezen.")
        else:
            try:
                ctx_lm5, base_maps_lm5_build = lm5_build_lastminute_context(
                    base_bytes=base_bytes_lm5,
                    absentees=gekozen_afwezigen_lm5,
                    start_uur=start_uur_lm5
                )

                lm5_result = lm5_write_lastminute_workbook(
                    base_bytes=base_bytes_lm5,
                    ctx=ctx_lm5,
                    base_maps=base_maps_lm5_build,
                    start_uur=start_uur_lm5,
                    absentees=gekozen_afwezigen_lm5
                )

                if isinstance(lm5_result, (bytes, bytearray)):
                    lm5_file_bytes = bytes(lm5_result)
                else:
                    lm5_output = BytesIO()
                    lm5_result.save(lm5_output)
                    lm5_output.seek(0)
                    lm5_file_bytes = lm5_output.getvalue()

                st.session_state["lm5_result_bytes"] = lm5_file_bytes
                st.session_state["lm5_result_filename"] = f"Planning_last_minute_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            except Exception as e:
                st.error(f"Fout in last-minute planner: {e}")

    if "lm5_result_bytes" in st.session_state:
        st.success("Last-minute planning gemaakt.")
        st.download_button(
            "Download last-minute planning",
            data=st.session_state["lm5_result_bytes"],
            file_name=st.session_state.get("lm5_result_filename", "Planning_last_minute.xlsx"),
            key="lm5_download_button"
        )
